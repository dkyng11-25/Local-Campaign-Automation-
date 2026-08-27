import os
import copy
import json
import random
import re
import threading
import time as time_module
import warnings

import pandas as pd
import requests

from dotenv import load_dotenv
from collections import Counter, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta, timezone
from email.utils import parsedate_to_datetime
from enum import Enum
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

# openpyxl이 날짜 서식이 적용된 비정상적인 큰 숫자를 읽을 때 발생시키는
# 특정 UserWarning만 숨긴다. 다른 openpyxl 경고와 일반 경고는 유지한다.
# read_only=True에서는 셀을 실제 순회할 때 경고가 늦게 발생할 수 있으므로
# load_workbook() 호출 구간만이 아니라 모듈 전체에 선택적으로 적용한다.
warnings.filterwarnings(
    "ignore",
    message=(
        r"Cell .* is marked as a date but the serial value .* "
        r"is outside the limits for dates\. "
        r"The cell will be treated as an error\."
    ),
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader",
)

# =============================================================================
# 프로젝트 경로 설정
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_ROOT / "output"
PAYLOAD_DIR = PROJECT_ROOT / "payload"

# Buzz Volume은 날짜별/차수별 파이프라인 폴더와 분리된
# 공용 작업 폴더를 사용한다.
BUZZ_VOLUME_DIR = OUTPUT_DIR / "Buzz_Volume"
BUZZ_VOLUME_COMPLETED_DIR = (
    BUZZ_VOLUME_DIR / "completed"
)
BUZZ_VOLUME_ARTIFACTS_DIR = (
    BUZZ_VOLUME_DIR / "artifacts"
)
FAILED_RESULTS_DIR = (
    BUZZ_VOLUME_ARTIFACTS_DIR
    / "failed_results"
)
RESPONSE_SAMPLE_DIR = (
    BUZZ_VOLUME_ARTIFACTS_DIR
    / "response_samples"
)

HEADER_ROW = 4
ROW_ID_COLUMN_NAME = "#"
CAMPAIGN_DATE_COLUMN_NAME = "Campaign Date"
BUZZ_VOLUME_COLUMN_NAME = "Buzz Volume"
SAVE_FAILED_RESPONSES = True

# 병렬 처리 및 안정성 설정
BATCH_SIZE = 100

# Worker는 동시에 대기할 수 있는 API 작업 수다.
# 실제 요청 시작 속도는 MAX_REQUESTS_PER_SECOND가 별도로 제한한다.
MAX_WORKERS = 7
MAX_REQUESTS_PER_SECOND = 5
RATE_LIMIT_PERIOD_SECONDS = 1.0

# Sprinklr가 403 Developer Over Rate 또는 429를 반환했는데
# Retry-After 헤더가 없을 때 모든 스레드가 함께 쉬는 최소/최대 시간이다.
RATE_LIMIT_BASE_COOLDOWN_SECONDS = 15.0
RATE_LIMIT_MAX_COOLDOWN_SECONDS = 300.0

# MAX_RETRIES는 최초 요청 이후 추가로 재시도하는 횟수다.
# 예: MAX_RETRIES=3이면 최대 4회 요청한다.
MAX_RETRIES = 3
RETRY_BASE_DELAY_SECONDS = 2.0
RETRY_MAX_DELAY_SECONDS = 60.0
FINAL_FAILED_RETRY_ROUNDS = 1

REQUEST_CONNECT_TIMEOUT_SECONDS = 10
REQUEST_READ_TIMEOUT_SECONDS = 180

SAVE_CHECKPOINT_AFTER_EACH_BATCH = True
API_FAILED_VALUE = "API_Failed"

RETRYABLE_STATUS_CODES = frozenset({
    429,
    500,
    502,
    503,
    504,
})

"""사용자 설정 필요"""
SPRINKLR_BASE_URL = "https://api3.sprinklr.com/prod"
ENDPOINT = "/api/v2/reports/query"

# 보안상 실제 값은 코드에 직접 쓰기보다 환경변수로 관리
load_dotenv(
    dotenv_path=PROJECT_ROOT/".env"
)
API_KEY = os.getenv("SPRINKLR_API_KEY")
ACCESS_TOKEN = os.getenv("SPRINKLR_ACCESS_TOKEN")



# =============================================================================
# 병렬 API 호출 공통 도구
# =============================================================================


class SprinklrRequestError(RuntimeError):
    """Sprinklr API 요청 실패 정보를 보존하는 예외."""

    def __init__(
        self,
        message: str,
        *,
        attempt_count: int,
        status_code: int | None = None,
        response_text: str | None = None,
    ) -> None:
        super().__init__(message)
        self.attempt_count = attempt_count
        self.status_code = status_code
        self.response_text = response_text


@dataclass(frozen=True)
class SprinklrFetchResult:
    response_json: dict[str, Any]
    attempt_count: int


class SlidingWindowRateLimiter:
    """
    모든 작업 스레드가 공유하는 요청 시작 속도 제한기.

    max_calls=5, period_seconds=1.0이면 최근 1초 구간에
    최대 5개의 요청 시작만 허용한다.

    Sprinklr가 rate limit 응답을 반환하면 defer()로 모든 스레드에
    공통 cooldown을 적용한다. 한 스레드만 sleep하는 방식보다
    동시에 재시도하는 thundering herd를 줄일 수 있다.
    """

    def __init__(
        self,
        max_calls: int,
        period_seconds: float,
    ) -> None:
        if max_calls <= 0:
            raise ValueError(
                "max_calls는 1 이상이어야 합니다."
            )

        if period_seconds <= 0:
            raise ValueError(
                "period_seconds는 0보다 커야 합니다."
            )

        self.max_calls = max_calls
        self.period_seconds = period_seconds
        self._timestamps: deque[float] = deque()
        self._blocked_until = 0.0
        self._lock = threading.Lock()

    def acquire(self) -> None:
        while True:
            sleep_seconds = 0.0

            with self._lock:
                current_time = time_module.monotonic()

                if current_time < self._blocked_until:
                    sleep_seconds = (
                        self._blocked_until - current_time
                    )
                else:
                    while (
                        self._timestamps
                        and current_time - self._timestamps[0]
                        >= self.period_seconds
                    ):
                        self._timestamps.popleft()

                    if len(self._timestamps) < self.max_calls:
                        self._timestamps.append(current_time)
                        return

                    sleep_seconds = max(
                        0.0,
                        self.period_seconds
                        - (current_time - self._timestamps[0]),
                    )

            if _STOP_EVENT.is_set():
                raise InterruptedError(
                    "사용자 중단 요청으로 API 호출 대기를 종료합니다."
                )

            if sleep_seconds > 0:
                _STOP_EVENT.wait(
                    timeout=min(sleep_seconds, 1.0)
                )

    def defer(self, cooldown_seconds: float) -> float:
        """모든 스레드의 다음 요청 가능 시각을 뒤로 미룬다."""
        if cooldown_seconds <= 0:
            return 0.0

        with self._lock:
            current_time = time_module.monotonic()
            requested_until = current_time + cooldown_seconds
            self._blocked_until = max(
                self._blocked_until,
                requested_until,
            )
            return max(
                0.0,
                self._blocked_until - current_time,
            )


RATE_LIMITER = SlidingWindowRateLimiter(
    max_calls=MAX_REQUESTS_PER_SECOND,
    period_seconds=RATE_LIMIT_PERIOD_SECONDS,
)

_THREAD_LOCAL = threading.local()
_SESSION_REGISTRY: list[requests.Session] = []
_SESSION_REGISTRY_LOCK = threading.Lock()
_PRINT_LOCK = threading.Lock()
_STOP_EVENT = threading.Event()


def safe_print(*values: Any) -> None:
    """병렬 작업 로그가 서로 섞이지 않도록 출력한다."""
    with _PRINT_LOCK:
        print(*values)


def get_thread_session() -> requests.Session:
    """각 작업 스레드에 전용 requests.Session을 하나씩 생성한다."""
    session = getattr(
        _THREAD_LOCAL,
        "sprinklr_session",
        None,
    )

    if session is None:
        session = requests.Session()

        adapter = requests.adapters.HTTPAdapter(
            pool_connections=1,
            pool_maxsize=1,
            max_retries=0,
        )
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        _THREAD_LOCAL.sprinklr_session = session

        with _SESSION_REGISTRY_LOCK:
            _SESSION_REGISTRY.append(session)

    return session


def close_registered_sessions() -> None:
    """현재 배치에서 생성된 스레드별 Session을 닫는다."""
    with _SESSION_REGISTRY_LOCK:
        sessions = list(_SESSION_REGISTRY)
        _SESSION_REGISTRY.clear()

    for session in sessions:
        session.close()


def parse_retry_after_seconds(
    retry_after_header: str | None,
) -> float | None:
    """
    Retry-After의 두 표준 형식을 모두 지원한다.

    - 초 단위 숫자: "120"
    - HTTP 날짜: "Wed, 21 Oct 2015 07:28:00 GMT"
    """
    if not retry_after_header:
        return None

    normalized_header = retry_after_header.strip()

    try:
        seconds = float(normalized_header)
    except ValueError:
        seconds = None

    if seconds is not None:
        return max(0.0, seconds)

    try:
        retry_datetime = parsedate_to_datetime(
            normalized_header
        )
    except (TypeError, ValueError, OverflowError):
        return None

    if retry_datetime.tzinfo is None:
        retry_datetime = retry_datetime.replace(
            tzinfo=timezone.utc
        )

    now_utc = datetime.now(timezone.utc)
    return max(
        0.0,
        (retry_datetime.astimezone(timezone.utc) - now_utc).total_seconds(),
    )


def calculate_retry_delay(
    attempt_number: int,
    retry_after_header: str | None,
) -> float:
    """일반 재시도용 Retry-After 또는 exponential backoff+jitter."""
    retry_after_seconds = parse_retry_after_seconds(
        retry_after_header
    )

    if retry_after_seconds is not None:
        return min(
            retry_after_seconds,
            RETRY_MAX_DELAY_SECONDS,
        )

    exponential_delay = (
        RETRY_BASE_DELAY_SECONDS
        * (2 ** max(attempt_number - 1, 0))
    )
    jitter = random.uniform(0.0, 1.0)

    return min(
        exponential_delay + jitter,
        RETRY_MAX_DELAY_SECONDS,
    )


def calculate_rate_limit_delay(
    attempt_number: int,
    retry_after_header: str | None,
) -> float:
    """Rate limit 전용으로 더 보수적인 공통 cooldown을 계산한다."""
    retry_after_seconds = parse_retry_after_seconds(
        retry_after_header
    )

    if retry_after_seconds is not None:
        return min(
            retry_after_seconds,
            RATE_LIMIT_MAX_COOLDOWN_SECONDS,
        )

    exponential_delay = (
        RATE_LIMIT_BASE_COOLDOWN_SECONDS
        * (2 ** max(attempt_number - 1, 0))
    )
    jitter = random.uniform(0.0, 3.0)

    return min(
        exponential_delay + jitter,
        RATE_LIMIT_MAX_COOLDOWN_SECONDS,
    )


def is_developer_over_rate_response(
    response: requests.Response,
) -> bool:
    """Sprinklr 고유의 403 Developer Over Rate 응답인지 판별한다."""
    return (
        response.status_code == 403
        and "developer over rate" in response.text.casefold()
    )


def is_rate_limit_response(
    response: requests.Response,
) -> bool:
    return (
        response.status_code == 429
        or is_developer_over_rate_response(response)
    )

# =============================================================================
# Buzz Volume 공용 폴더 및 Excel 경로 생성
# =============================================================================


def ensure_buzz_volume_directories() -> None:
    """
    Buzz Volume 결과 및 보조 산출물 폴더가 존재하도록 보장한다.

    입력 폴더인 output/Buzz_Volume 자체가 없으면 사용자의
    run_pipeline.py 실행 또는 프로젝트 폴더 구성을 확인하도록
    명확한 오류를 발생시킨다.

    completed 및 artifacts 하위 폴더는 이 모듈에서도 안전하게
    생성하여 단독 실행을 지원한다.
    """

    if not BUZZ_VOLUME_DIR.exists():
        raise FileNotFoundError(
            "Buzz Volume 공용 입력 폴더를 찾을 수 없습니다.\n"
            f"확인 경로: {BUZZ_VOLUME_DIR}\n"
            "먼저 run_pipeline.py를 한 번 실행하거나 "
            "해당 폴더를 생성하세요."
        )

    if not BUZZ_VOLUME_DIR.is_dir():
        raise NotADirectoryError(
            "Buzz Volume 입력 경로가 폴더가 아닙니다.\n"
            f"확인 경로: {BUZZ_VOLUME_DIR}"
        )

    for directory in (
        BUZZ_VOLUME_COMPLETED_DIR,
        FAILED_RESULTS_DIR,
        RESPONSE_SAMPLE_DIR,
    ):
        directory.mkdir(
            parents=True,
            exist_ok=True,
        )


def build_excel_paths(
    input_date: str,
) -> tuple[Path, Path]:
    """
    YYYY-MM-DD 기준 날짜로 예상 파일명을 계산한다.

    입력:
        output/Buzz_Volume/
        {YYMMDD}_SLCC_SOV_Local Campaign Tracking_{월}월_v01.xlsx

    출력:
        output/Buzz_Volume/completed/
        {입력 파일 stem}_mentions_updated.xlsx

    사용자는 파일명이나 경로를 실행 인자로 입력할 필요가 없고,
    기준 날짜에 맞는 최종 통합·정제 Excel을 Buzz_Volume 폴더에
    넣기만 하면 된다.
    """

    input_date = input_date.strip()

    try:
        parsed_date = datetime.strptime(
            input_date,
            "%Y-%m-%d",
        )
    except ValueError as exc:
        raise ValueError(
            "날짜는 YYYY-MM-DD 형식으로 입력해야 합니다. "
            "예시) 2026-07-14"
        ) from exc

    ensure_buzz_volume_directories()

    date_text = parsed_date.strftime("%y%m%d")
    month = parsed_date.month

    input_filename = (
        f"{date_text}"
        "_SLCC_SOV_Local Campaign Tracking_"
        f"{month}월_v01.xlsx"
    )

    input_excel_path = (
        BUZZ_VOLUME_DIR
        / input_filename
    )

    output_excel_path = (
        BUZZ_VOLUME_COMPLETED_DIR
        / (
            f"{Path(input_filename).stem}"
            "_mentions_updated.xlsx"
        )
    )

    if not input_excel_path.exists():
        raise FileNotFoundError(
            "Buzz Volume 업데이트에 사용할 최종 통합·정제 "
            "Excel 파일을 찾을 수 없습니다.\n"
            f"기준 날짜: {input_date}\n"
            f"필요한 파일명: {input_filename}\n"
            f"파일을 넣을 폴더: {BUZZ_VOLUME_DIR}\n"
            f"확인 경로: {input_excel_path}"
        )

    if not input_excel_path.is_file():
        raise FileNotFoundError(
            "Buzz Volume 입력 Excel 경로가 파일이 아닙니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if input_excel_path.resolve() == output_excel_path.resolve():
        raise ValueError(
            "Buzz Volume 입력 파일과 결과 파일 경로가 같습니다."
        )

    return (
        input_excel_path,
        output_excel_path,
    )


# =============================================================================
# Base payload 로드
# =============================================================================

def load_base_payload(
    payload_path: Path,
) -> dict[str, Any]:
    if not payload_path.exists():
        raise FileNotFoundError(
            f"Base payload 파일을 찾을 수 없습니다: {payload_path}"
        )

    if not payload_path.is_file():
        raise FileNotFoundError(
            f"Base payload 경로가 파일이 아닙니다: {payload_path}"
        )

    try:
        with payload_path.open(
            mode="r",
            encoding="utf-8",
        ) as file:
            payload = json.load(file)

    except json.JSONDecodeError as exc:
        raise ValueError(
            f"Base payload가 올바른 JSON 형식이 아닙니다: {payload_path}"
        ) from exc

    if not isinstance(payload, dict):
        raise TypeError(
            "Base payload의 최상위 구조는 JSON object여야 합니다."
        )

    return payload


# =============================================================================
# 날짜 data cut 설정
# =============================================================================

SEOUL_TIMEZONE = ZoneInfo("Asia/Seoul")

# Daily 데이터 컷의 고정 시작일
DAILY_START_DATE = date(2026, 7, 1)

# 마지막 초 전체를 포함하도록 999 milliseconds까지 설정
DAILY_END_TIME = time(
    hour=17,
    minute=0,
    second=59,
    microsecond=999_000,
)

WEEKLY_START_TIME = time(
    hour=0,
    minute=0,
    second=0,
    microsecond=0,
)

WEEKLY_END_TIME = time(
    hour=23,
    minute=59,
    second=59,
    microsecond=999_000,
)


class DataCutType(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"


@dataclass(frozen=True)
class DateTimeRange:
    data_cut_type: DataCutType
    reference_date: date
    start_datetime: datetime
    end_datetime: datetime


def parse_reference_date(
    value: str,
) -> date:
    value = value.strip()

    try:
        return datetime.strptime(
            value,
            "%Y-%m-%d",
        ).date()

    except ValueError as exc:
        raise ValueError(
            "기준 날짜는 YYYY-MM-DD 형식이어야 합니다. "
            f"입력값: {value!r}"
        ) from exc


def build_daily_range(
    reference_date: date,
) -> DateTimeRange:
    if reference_date < DAILY_START_DATE:
        raise ValueError(
            "Daily 기준 날짜는 고정 시작일보다 "
            "이전일 수 없습니다. "
            f"고정 시작일: {DAILY_START_DATE}, "
            f"입력 날짜: {reference_date}"
        )

    start_datetime = datetime.combine(
        DAILY_START_DATE,
        time.min,
        tzinfo=SEOUL_TIMEZONE,
    )

    end_datetime = datetime.combine(
        reference_date,
        DAILY_END_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    return DateTimeRange(
        data_cut_type=DataCutType.DAILY,
        reference_date=reference_date,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
    )


def build_weekly_range(
    reference_date: date,
) -> DateTimeRange:
    # Python weekday(): 월요일=0, 일요일=6
    if reference_date.weekday() != 0:
        raise ValueError(
            "Weekly 기준 날짜는 반드시 월요일이어야 합니다. "
            f"입력 날짜: {reference_date}, "
            f"요일 번호: {reference_date.weekday()}"
        )

    previous_monday = reference_date - timedelta(days=7)
    previous_sunday = reference_date - timedelta(days=1)

    start_datetime = datetime.combine(
        previous_monday,
        WEEKLY_START_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    end_datetime = datetime.combine(
        previous_sunday,
        WEEKLY_END_TIME,
        tzinfo=SEOUL_TIMEZONE,
    )

    return DateTimeRange(
        data_cut_type=DataCutType.WEEKLY,
        reference_date=reference_date,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
    )


def build_datetime_range(
    data_cut_type: str,
    reference_date: date,
) -> DateTimeRange:
    normalized_type = data_cut_type.strip().lower()

    try:
        selected_type = DataCutType(normalized_type)

    except ValueError as exc:
        raise ValueError(
            "data_cut_type은 'daily' 또는 "
            f"'weekly'여야 합니다. 입력값: {data_cut_type!r}"
        ) from exc

    if selected_type == DataCutType.DAILY:
        return build_daily_range(
            reference_date=reference_date,
        )

    return build_weekly_range(
        reference_date=reference_date,
    )


UNIX_EPOCH = datetime(
    1970,
    1,
    1,
    tzinfo=timezone.utc,
)


def datetime_to_epoch_ms(
    value: datetime,
) -> int:
    if value.tzinfo is None:
        raise ValueError(
            "timezone 정보가 없는 datetime입니다."
        )

    utc_value = value.astimezone(timezone.utc)
    delta = utc_value - UNIX_EPOCH

    return (
        delta.days * 86_400_000
        + delta.seconds * 1_000
        + delta.microseconds // 1_000
    )


def datetime_range_to_epoch_ms(
    date_range: DateTimeRange,
) -> tuple[int, int]:
    start_time_ms = datetime_to_epoch_ms(
        date_range.start_datetime
    )

    end_time_ms = datetime_to_epoch_ms(
        date_range.end_datetime
    )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "변환된 startTime이 endTime보다 작지 않습니다."
        )

    return start_time_ms, end_time_ms


def build_data_cut(
    data_cut_type: str,
    reference_date_text: str,
) -> tuple[DateTimeRange, int, int]:
    """
    main()에서 받은 data_cut_type과 기준 날짜를 이용해
    날짜 범위 및 epoch milliseconds를 생성한다.
    """
    reference_date = parse_reference_date(
        reference_date_text
    )

    date_range = build_datetime_range(
        data_cut_type=data_cut_type,
        reference_date=reference_date,
    )

    start_time_ms, end_time_ms = (
        datetime_range_to_epoch_ms(date_range)
    )

    print()
    print(
        "데이터 컷 유형:",
        date_range.data_cut_type.value,
    )
    print(
        "시작 일시:",
        date_range.start_datetime,
    )
    print(
        "종료 일시:",
        date_range.end_datetime,
    )
    print("startTime:", start_time_ms)
    print("endTime:", end_time_ms)

    return date_range, start_time_ms, end_time_ms


# =============================================================================
# 날짜 data cut이 반영된 payload template 생성
# =============================================================================

def build_date_payload_template(
    base_payload: dict[str, Any],
    start_time_ms: int,
    end_time_ms: int,
) -> dict[str, Any]:
    if not isinstance(base_payload, dict):
        raise TypeError(
            "base_payload는 dict여야 합니다."
        )

    if not isinstance(start_time_ms, int):
        raise TypeError(
            "start_time_ms는 int여야 합니다."
        )

    if not isinstance(end_time_ms, int):
        raise TypeError(
            "end_time_ms는 int여야 합니다."
        )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "startTime은 endTime보다 작아야 합니다."
        )

    run_payload_template = copy.deepcopy(
        base_payload
    )

    run_payload_template["startTime"] = start_time_ms
    run_payload_template["endTime"] = end_time_ms

    return run_payload_template


# =============================================================================
# Excel에서 # / Query 추출 및 Query payload 생성
# =============================================================================

@dataclass(frozen=True)
class CampaignQueryTask:
    row_id: int
    query: str

    @property
    def task_key(self) -> tuple[int, str]:
        return self.row_id, self.query


@dataclass(frozen=True)
class CampaignQueryResult:
    row_id: int
    query: str
    success: bool
    mention_count: int | None
    error_message: str | None
    attempt_count: int
    status_code: int | None = None
    failed_response_path: str | None = None


def normalize_row_id(
    value: Any,
) -> int:
    if value is None or pd.isna(value):
        raise ValueError(
            "# 값이 비어 있습니다."
        )

    if isinstance(value, bool):
        raise ValueError(
            "# 값은 boolean일 수 없습니다."
        )

    try:
        numeric_value = float(value)

    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"# 값을 숫자로 변환할 수 없습니다: {value!r}"
        ) from exc

    if not numeric_value.is_integer():
        raise ValueError(
            "# 값은 정수여야 합니다. "
            f"입력값: {value!r}"
        )

    row_id = int(numeric_value)

    if row_id <= 0:
        raise ValueError(
            "# 값은 1 이상의 정수여야 합니다. "
            f"입력값: {value!r}"
        )

    return row_id


EXCEL_ROW_ID_FORMULA_PATTERN = re.compile(
    r"^=\s*ROW\(\s*\)\s*-\s*ROW\(\s*"
    r"(?:(?:'[^']+'|[^!()]+)!)?"
    r"\$?[A-Z]{1,3}\$?(\d+)\s*\)\s*$",
    flags=re.IGNORECASE,
)


def resolve_excel_row_id(
    value: Any,
    excel_row_number: int,
) -> int:
    """
    Excel '#' 셀의 값을 실제 정수 ID로 변환한다.

    지원 형식:
    - 일반 숫자: 1, 2, 3 ...
    - 행 번호 수식: =ROW()-ROW($B$4)

    예를 들어 Excel 11행에서 수식이 =ROW()-ROW($B$4)이면
    row_id는 11 - 4 = 7로 계산한다.
    """
    if isinstance(value, str):
        formula = value.strip()

        if formula.startswith("="):
            formula_match = EXCEL_ROW_ID_FORMULA_PATTERN.fullmatch(
                formula
            )

            if formula_match is None:
                raise ValueError(
                    "지원하지 않는 # 수식입니다. "
                    "지원 형식 예시: =ROW()-ROW($B$4). "
                    f"입력값: {value!r}"
                )

            anchor_row_number = int(
                formula_match.group(1)
            )
            calculated_row_id = (
                excel_row_number - anchor_row_number
            )

            return normalize_row_id(
                calculated_row_id
            )

    return normalize_row_id(value)


def normalize_query(
    value: Any,
) -> str:
    if value is None or pd.isna(value):
        raise ValueError(
            "Query 값이 비어 있습니다."
        )

    query = str(value).strip()

    if not query:
        raise ValueError(
            "Query 값이 빈 문자열입니다."
        )

    # Query 내부 문법은 변경하지 않고 양끝 공백만 제거한다.
    return query


def parse_campaign_date_value(
    value: Any,
    *,
    date_range: DateTimeRange,
) -> date:
    """
    Weekly 필터용 Campaign Date를 실제 날짜로 변환한다.

    Excel 화면에는 `08월 02일`처럼 월/일만 표시되더라도,
    셀의 실제 저장값이 datetime/date 또는 `YYYY-MM-DD` 형식이면
    실제 연도까지 포함된 날짜를 그대로 사용한다.

    지원 입력 예시:
    - openpyxl datetime/date 값
    - pandas Timestamp
    - 2026-08-02
    - 2026-08-02 00:00:00
    - 2026/08/02
    - 2026년 8월 2일
    - 8월 2일 / 8/2
    """
    if value is None or pd.isna(value):
        raise ValueError(
            "Campaign Date 값이 비어 있습니다."
        )

    # Excel 날짜 셀은 표시 형식과 무관하게 openpyxl에서
    # datetime 또는 date로 읽히므로 실제 저장 날짜를 우선 사용한다.
    if isinstance(value, pd.Timestamp):
        return value.date()

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    if not text:
        raise ValueError(
            "Campaign Date 값이 빈 문자열입니다."
        )

    # YYYY-MM-DD 및 YYYY-MM-DD HH:MM:SS 같은 ISO 형식을
    # 먼저 처리한다. Excel에서 날짜가 문자열로 저장된 경우도 지원한다.
    try:
        return datetime.fromisoformat(
            text.replace("Z", "+00:00")
        ).date()
    except ValueError:
        pass

    full_date_formats = (
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%m/%d/%Y",
        "%Y/%m/%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    )

    for date_format in full_date_formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            ).date()
        except ValueError:
            pass

    korean_full_match = re.fullmatch(
        r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*"
        r"(\d{1,2})\s*일",
        text,
    )

    if korean_full_match:
        return date(
            int(korean_full_match.group(1)),
            int(korean_full_match.group(2)),
            int(korean_full_match.group(3)),
        )

    # 과거 파일에서 실제 셀 값 자체가 월/일 문자열인 경우를 위한 fallback.
    month_day_match = re.fullmatch(
        r"(\d{1,2})\s*(?:월|/)\s*"
        r"(\d{1,2})\s*일?",
        text,
    )

    if month_day_match:
        month = int(month_day_match.group(1))
        day = int(month_day_match.group(2))
        start_date = date_range.start_datetime.date()
        end_date = date_range.end_datetime.date()

        candidate_dates: list[date] = []
        for year in range(
            start_date.year - 1,
            end_date.year + 2,
        ):
            try:
                candidate_dates.append(
                    date(year, month, day)
                )
            except ValueError:
                continue

        in_range_candidates = [
            candidate
            for candidate in candidate_dates
            if start_date <= candidate <= end_date
        ]

        if len(in_range_candidates) == 1:
            return in_range_candidates[0]

        if candidate_dates:
            def distance_from_target_range(
                candidate: date,
            ) -> int:
                if candidate < start_date:
                    return (start_date - candidate).days
                if candidate > end_date:
                    return (candidate - end_date).days
                return 0

            return min(
                candidate_dates,
                key=distance_from_target_range,
            )

    raise ValueError(
        "Campaign Date를 날짜로 해석할 수 없습니다. "
        "지원 예시: 2026-08-02, "
        "2026-08-02 00:00:00, 2026/08/02, "
        "8월 2일, 8/2. "
        f"입력값: {value!r}"
    )


def normalize_header_name(
    value: Any,
) -> str:
    """
    Excel Header 비교용 문자열을 정규화한다.

    - 줄바꿈, 탭, 연속 공백을 일반 공백 1개로 통일
    - 앞뒤 공백 제거
    - 대소문자 차이 제거

    예시:
    "Buzz\nVolume" -> "buzz volume"
    "  Buzz   Volume  " -> "buzz volume"
    """
    if value is None:
        return ""

    normalized_value = (
        str(value)
        .replace("\u00A0", " ")
    )

    return " ".join(
        normalized_value.split()
    ).casefold()


def validate_campaign_query_tasks(
    tasks: list[CampaignQueryTask],
) -> None:
    if not tasks:
        raise ValueError(
            "처리할 캠페인 Query가 없습니다."
        )

    row_id_counts = Counter(
        task.row_id
        for task in tasks
    )

    duplicated_row_ids = [
        row_id
        for row_id, count in row_id_counts.items()
        if count > 1
    ]

    if duplicated_row_ids:
        raise ValueError(
            "# 컬럼에 중복 값이 있습니다: "
            f"{duplicated_row_ids[:20]}"
        )

    task_key_counts = Counter(
        task.task_key
        for task in tasks
    )

    duplicated_task_keys = [
        task_key
        for task_key, count in task_key_counts.items()
        if count > 1
    ]

    if duplicated_task_keys:
        raise ValueError(
            "(#, Query) 고유키가 중복되었습니다: "
            f"{duplicated_task_keys[:20]}"
        )


def load_campaign_query(
    excel_path: Path,
    sheet_name: str,
    *,
    date_range: DateTimeRange,
) -> list[CampaignQueryTask]:
    """
    Excel 전체를 한 번 읽어 Query가 있는 행을 작업 목록으로 만든다.

    '#' 셀이 일반 숫자이거나 =ROW()-ROW($B$4) 형태의 수식인 경우를
    동일하게 처리하기 위해 openpyxl의 실제 셀 값을 사용한다.
    같은 Query가 여러 행에 있어도 각 행을 별도 API 작업으로 유지한다.

    Weekly 실행에서는 Campaign Date가 직전 월~일 범위에 포함되는
    행만 API 대상에 포함한다. Daily 실행은 기존처럼 Query가 있는
    전체 행을 처리한다.
    """
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel 파일을 찾을 수 없습니다: {excel_path}"
        )

    if not excel_path.is_file():
        raise FileNotFoundError(
            f"Excel 경로가 파일이 아닙니다: {excel_path}"
        )

    keep_vba = excel_path.suffix.lower() == ".xlsm"

    workbook = load_workbook(
        filename=excel_path,
        data_only=False,
        read_only=True,
        keep_vba=keep_vba,
        keep_links=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "대상 시트를 찾을 수 없습니다.\n"
                f"대상 시트: {sheet_name}\n"
                f"현재 시트 목록: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]

        required_header_names = [
            ROW_ID_COLUMN_NAME,
            "Query",
        ]

        if date_range.data_cut_type == DataCutType.WEEKLY:
            required_header_names.append(
                CAMPAIGN_DATE_COLUMN_NAME
            )

        required_headers = {
            header_name: []
            for header_name in required_header_names
        }
        normalized_header_lookup = {
            normalize_header_name(header_name): header_name
            for header_name in required_headers
        }

        header_row_cells = next(
            worksheet.iter_rows(
                min_row=HEADER_ROW,
                max_row=HEADER_ROW,
            )
        )

        for cell in header_row_cells:
            normalized_cell_header = normalize_header_name(
                cell.value
            )
            canonical_header_name = (
                normalized_header_lookup.get(
                    normalized_cell_header
                )
            )

            if canonical_header_name is not None:
                required_headers[canonical_header_name].append(
                    cell.column
                )

        for header_name, column_numbers in required_headers.items():
            if len(column_numbers) != 1:
                raise ValueError(
                    f"Header 행에서 {header_name!r} 컬럼을 "
                    "정확히 1개 찾아야 합니다.\n"
                    f"Header 행: {HEADER_ROW}\n"
                    f"발견 개수: {len(column_numbers)}"
                )

        row_id_column_number = required_headers[
            ROW_ID_COLUMN_NAME
        ][0]
        query_column_number = required_headers["Query"][0]
        campaign_date_column_number = None

        if date_range.data_cut_type == DataCutType.WEEKLY:
            campaign_date_column_number = required_headers[
                CAMPAIGN_DATE_COLUMN_NAME
            ][0]

        tasks: list[CampaignQueryTask] = []
        validation_errors: list[str] = []
        skipped_empty_query_rows: list[int] = []
        skipped_out_of_range_rows: list[int] = []

        required_column_numbers = [
            row_id_column_number,
            query_column_number,
        ]
        if campaign_date_column_number is not None:
            required_column_numbers.append(
                campaign_date_column_number
            )

        maximum_required_column = max(
            required_column_numbers
        )

        data_rows = worksheet.iter_rows(
            min_row=HEADER_ROW + 1,
            max_row=worksheet.max_row,
            max_col=maximum_required_column,
        )

        weekly_start_date = date_range.start_datetime.date()
        weekly_end_date = date_range.end_datetime.date()

        for excel_row_number, row_cells in enumerate(
            data_rows,
            start=HEADER_ROW + 1,
        ):
            raw_query = row_cells[
                query_column_number - 1
            ].value

            if (
                raw_query is None
                or (
                    isinstance(raw_query, str)
                    and not raw_query.strip()
                )
            ):
                skipped_empty_query_rows.append(
                    excel_row_number
                )
                continue

            if campaign_date_column_number is not None:
                raw_campaign_date = row_cells[
                    campaign_date_column_number - 1
                ].value

                try:
                    campaign_date = parse_campaign_date_value(
                        raw_campaign_date,
                        date_range=date_range,
                    )
                except ValueError as exc:
                    validation_errors.append(
                        f"Excel 행 {excel_row_number}: {exc}"
                    )
                    continue

                if not (
                    weekly_start_date
                    <= campaign_date
                    <= weekly_end_date
                ):
                    skipped_out_of_range_rows.append(
                        excel_row_number
                    )
                    continue

            raw_row_id = row_cells[
                row_id_column_number - 1
            ].value

            try:
                row_id = resolve_excel_row_id(
                    value=raw_row_id,
                    excel_row_number=excel_row_number,
                )
                query = normalize_query(raw_query)

                tasks.append(
                    CampaignQueryTask(
                        row_id=row_id,
                        query=query,
                    )
                )

            except ValueError as exc:
                validation_errors.append(
                    f"Excel 행 {excel_row_number}: {exc}"
                )

        if validation_errors:
            error_preview = "\n".join(
                validation_errors[:20]
            )
            remaining_error_count = max(
                0,
                len(validation_errors) - 20,
            )
            extra_message = (
                f"\n외 {remaining_error_count}개 오류"
                if remaining_error_count
                else ""
            )

            raise ValueError(
                "Excel Query 입력값 검증에 실패했습니다.\n"
                f"{error_preview}{extra_message}"
            )

        if skipped_empty_query_rows:
            print()
            print(
                "[SKIP] Query가 비어 있어 건너뛴 Excel 행 수: "
                f"{len(skipped_empty_query_rows)}"
            )
            print(
                "건너뛴 Excel 행: "
                f"{skipped_empty_query_rows[:20]}"
            )

            if len(skipped_empty_query_rows) > 20:
                print(
                    "추가로 건너뛴 행 수: "
                    f"{len(skipped_empty_query_rows) - 20}"
                )

        if skipped_out_of_range_rows:
            print()
            print(
                "[WEEKLY] Weekly 대상 기간 밖이라 건너뛴 행 수: "
                f"{len(skipped_out_of_range_rows)}"
            )
            print(
                "Weekly 대상 기간: "
                f"{weekly_start_date} ~ {weekly_end_date}"
            )
            print(
                "기간 밖 Excel 행: "
                f"{skipped_out_of_range_rows[:20]}"
            )

            if len(skipped_out_of_range_rows) > 20:
                print(
                    "추가로 기간 밖인 행 수: "
                    f"{len(skipped_out_of_range_rows) - 20}"
                )

        # Skip 결과를 먼저 출력한 뒤 최종 작업 목록을 검증한다.
        # 따라서 tasks가 0개여도 Query 공란 때문인지, Weekly 기간
        # 필터 때문인지 실행 로그에서 즉시 확인할 수 있다.
        if not tasks:
            error_lines = [
                "처리할 캠페인 Query가 없습니다.",
                f"Query 공란으로 제외된 행 수: {len(skipped_empty_query_rows)}",
            ]

            if date_range.data_cut_type == DataCutType.WEEKLY:
                error_lines.extend([
                    (
                        "Weekly 기간 밖으로 제외된 행 수: "
                        f"{len(skipped_out_of_range_rows)}"
                    ),
                    (
                        "Weekly 대상 기간: "
                        f"{weekly_start_date} ~ {weekly_end_date}"
                    ),
                ])

            raise ValueError(
                "\n".join(error_lines)
            )

        validate_campaign_query_tasks(tasks)

        return tasks

    finally:
        workbook.close()


def build_query_payload(
    run_payload_template: dict[str, Any],
    query: Any,
) -> dict[str, Any]:
    query_text = normalize_query(query)

    payload = copy.deepcopy(
        run_payload_template
    )

    filters = payload.get("filters")

    if not isinstance(filters, list):
        raise ValueError(
            "payload의 filters가 리스트가 아닙니다."
        )

    query_filters = [
        filter_item
        for filter_item in filters
        if isinstance(filter_item, dict)
        and filter_item.get("dimensionName") == "QUERY"
    ]

    if len(query_filters) != 1:
        raise ValueError(
            "QUERY 필터가 정확히 1개여야 합니다. "
            f"현재 발견된 개수: {len(query_filters)}"
        )

    query_filter = query_filters[0]

    if query_filter.get("filterType") != "IN":
        raise ValueError(
            "QUERY 필터의 filterType이 'IN'이 아닙니다."
        )

    query_filter["values"] = [query_text]

    return payload

def fetch_sprinklr_data(
    base_url: str,
    endpoint: str,
    api_key: str | None,
    access_token: str | None,
    payload: dict[str, Any],
    *,
    request_label: str,
) -> SprinklrFetchResult:
    """
    Sprinklr Reporting API를 호출한다.

    자동 재시도 대상:
    - Timeout 및 네트워크 연결 오류
    - 429
    - 500, 502, 503, 504
    - 403 응답 중 본문이 "Developer Over Rate"인 경우

    일반적인 400, 401, 403 권한 오류는 즉시 실패 처리한다.
    Rate limit 응답은 개별 스레드만 쉬지 않고 전역 limiter에
    cooldown을 적용하여 모든 worker가 함께 속도를 낮춘다.
    """
    if not api_key:
        raise ValueError(
            "API_KEY가 없습니다. "
            "SPRINKLR_API_KEY 환경변수를 확인하세요."
        )

    if not access_token:
        raise ValueError(
            "ACCESS_TOKEN이 없습니다. "
            "SPRINKLR_ACCESS_TOKEN 환경변수를 확인하세요."
        )

    if not isinstance(payload, dict):
        raise TypeError(
            "payload는 dict여야 합니다."
        )

    url = (
        f"{base_url.rstrip('/')}/"
        f"{endpoint.lstrip('/')}"
    )

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Key": api_key,
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    session = get_thread_session()
    max_attempts = MAX_RETRIES + 1

    for attempt_number in range(1, max_attempts + 1):
        if _STOP_EVENT.is_set():
            raise SprinklrRequestError(
                "사용자 중단 요청으로 API 작업을 종료했습니다.",
                attempt_count=max(0, attempt_number - 1),
            )

        try:
            RATE_LIMITER.acquire()

            response = session.post(
                url,
                headers=headers,
                json=payload,
                timeout=(
                    REQUEST_CONNECT_TIMEOUT_SECONDS,
                    REQUEST_READ_TIMEOUT_SECONDS,
                ),
            )

            response_text = response.text[:3000]
            rate_limit_error = is_rate_limit_response(response)
            retryable_http_error = (
                rate_limit_error
                or response.status_code in RETRYABLE_STATUS_CODES
            )

            if retryable_http_error:
                if rate_limit_error:
                    retry_delay = calculate_rate_limit_delay(
                        attempt_number=attempt_number,
                        retry_after_header=response.headers.get(
                            "Retry-After"
                        ),
                    )
                    effective_cooldown = RATE_LIMITER.defer(
                        retry_delay
                    )
                    error_name = (
                        "Developer Over Rate"
                        if is_developer_over_rate_response(response)
                        else "Rate Limit"
                    )

                    safe_print(
                        f"[COOLDOWN] 공통 rate-limit cooldown: "
                        f"{request_label}, {error_name}, "
                        f"HTTP {response.status_code}, "
                        f"약 {effective_cooldown:.1f}초, "
                        f"시도 {attempt_number}/{max_attempts}"
                    )
                else:
                    retry_delay = calculate_retry_delay(
                        attempt_number=attempt_number,
                        retry_after_header=response.headers.get(
                            "Retry-After"
                        ),
                    )
                    safe_print(
                        f"[RETRY] 재시도 예정: {request_label}, "
                        f"HTTP {response.status_code}, "
                        f"{retry_delay:.1f}초 후 "
                        f"재시도 ({attempt_number}/{max_attempts})"
                    )

                if attempt_number < max_attempts:
                    if rate_limit_error:
                        # 다음 루프의 acquire()에서 모든 worker가
                        # 공통 cooldown이 끝날 때까지 기다린다.
                        continue

                    if _STOP_EVENT.wait(timeout=retry_delay):
                        raise SprinklrRequestError(
                            "사용자 중단 요청으로 재시도 대기를 "
                            "종료했습니다.",
                            attempt_count=attempt_number,
                            status_code=response.status_code,
                            response_text=response_text,
                        )
                    continue

                error_description = (
                    "Sprinklr API rate limit 오류가"
                    if rate_limit_error
                    else "Sprinklr API 재시도 가능 HTTP 오류가"
                )
                raise SprinklrRequestError(
                    f"{error_description} 최대 시도 횟수까지 "
                    "계속 발생했습니다.\n"
                    f"status_code: {response.status_code}\n"
                    f"url: {url}\n"
                    f"response: {response_text}",
                    attempt_count=attempt_number,
                    status_code=response.status_code,
                    response_text=response_text,
                )

            try:
                response.raise_for_status()
            except requests.exceptions.HTTPError as exc:
                raise SprinklrRequestError(
                    "Sprinklr API가 재시도하지 않는 HTTP 오류를 "
                    "반환했습니다.\n"
                    f"status_code: {response.status_code}\n"
                    f"url: {url}\n"
                    f"response: {response_text}",
                    attempt_count=attempt_number,
                    status_code=response.status_code,
                    response_text=response_text,
                ) from exc

            try:
                response_json = response.json()
            except requests.exceptions.JSONDecodeError as exc:
                if attempt_number < max_attempts:
                    retry_delay = calculate_retry_delay(
                        attempt_number=attempt_number,
                        retry_after_header=None,
                    )
                    safe_print(
                        f"[RETRY] JSON 응답 재시도 예정: {request_label}, "
                        f"{retry_delay:.1f}초 후 "
                        f"재시도 ({attempt_number}/{max_attempts})"
                    )
                    if _STOP_EVENT.wait(timeout=retry_delay):
                        raise SprinklrRequestError(
                            "사용자 중단 요청으로 JSON 응답 "
                            "재시도 대기를 종료했습니다.",
                            attempt_count=attempt_number,
                            status_code=response.status_code,
                            response_text=response_text,
                        )
                    continue

                raise SprinklrRequestError(
                    "Sprinklr API 응답이 최대 시도 횟수까지 "
                    "JSON 형식이 아니었습니다.\n"
                    f"status_code: {response.status_code}\n"
                    f"response: {response_text}",
                    attempt_count=attempt_number,
                    status_code=response.status_code,
                    response_text=response_text,
                ) from exc

            if not isinstance(response_json, dict):
                raise SprinklrRequestError(
                    "Sprinklr API 응답의 최상위 구조는 "
                    "JSON object여야 합니다. "
                    f"실제 타입: {type(response_json).__name__}",
                    attempt_count=attempt_number,
                    status_code=response.status_code,
                    response_text=response_text,
                )

            return SprinklrFetchResult(
                response_json=response_json,
                attempt_count=attempt_number,
            )

        except requests.exceptions.Timeout as exc:
            if attempt_number < max_attempts:
                retry_delay = calculate_retry_delay(
                    attempt_number=attempt_number,
                    retry_after_header=None,
                )
                safe_print(
                    f"[RETRY] Timeout 재시도 예정: {request_label}, "
                    f"{retry_delay:.1f}초 후 "
                    f"재시도 ({attempt_number}/{max_attempts})"
                )
                if _STOP_EVENT.wait(timeout=retry_delay):
                    raise SprinklrRequestError(
                        "사용자 중단 요청으로 Timeout 재시도 "
                        "대기를 종료했습니다.",
                        attempt_count=attempt_number,
                    )
                continue

            raise SprinklrRequestError(
                "Sprinklr API 요청 시간이 최대 시도 횟수까지 "
                "초과되었습니다.\n"
                f"URL: {url}",
                attempt_count=attempt_number,
            ) from exc

        except requests.exceptions.ConnectionError as exc:
            if attempt_number < max_attempts:
                retry_delay = calculate_retry_delay(
                    attempt_number=attempt_number,
                    retry_after_header=None,
                )
                safe_print(
                    f"[RETRY] 연결 오류 재시도 예정: {request_label}, "
                    f"{retry_delay:.1f}초 후 "
                    f"재시도 ({attempt_number}/{max_attempts})"
                )
                if _STOP_EVENT.wait(timeout=retry_delay):
                    raise SprinklrRequestError(
                        "사용자 중단 요청으로 연결 오류 재시도 "
                        "대기를 종료했습니다.",
                        attempt_count=attempt_number,
                    )
                continue

            raise SprinklrRequestError(
                "Sprinklr API 서버 연결 오류가 최대 시도 횟수까지 "
                "발생했습니다.\n"
                f"URL: {url}",
                attempt_count=attempt_number,
            ) from exc

        except requests.exceptions.RequestException as exc:
            if attempt_number < max_attempts:
                retry_delay = calculate_retry_delay(
                    attempt_number=attempt_number,
                    retry_after_header=None,
                )
                safe_print(
                    f"[RETRY] 요청 오류 재시도 예정: {request_label}, "
                    f"{retry_delay:.1f}초 후 "
                    f"재시도 ({attempt_number}/{max_attempts})"
                )
                if _STOP_EVENT.wait(timeout=retry_delay):
                    raise SprinklrRequestError(
                        "사용자 중단 요청으로 요청 오류 재시도 "
                        "대기를 종료했습니다.",
                        attempt_count=attempt_number,
                    )
                continue

            raise SprinklrRequestError(
                "Sprinklr API 요청 오류가 최대 시도 횟수까지 "
                "발생했습니다.\n"
                f"URL: {url}",
                attempt_count=attempt_number,
            ) from exc

    raise RuntimeError(
        "Sprinklr API 재시도 루프가 예상하지 못한 상태로 종료되었습니다."
    )


def save_response_sample(
    response_json: dict[str, Any],
    widget_name: str,
    output_dir: str | Path = RESPONSE_SAMPLE_DIR,
) -> Path:
    safe_widget_name = (
        widget_name
        .replace(" ", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
    )

    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)

    file_path = path / f"{safe_widget_name}_response.json"

    with file_path.open(
        mode="w",
        encoding="utf-8",
    ) as file:
        json.dump(
            response_json,
            file,
            ensure_ascii=False,
            indent=2,
        )

    return file_path


def extract_mention_count(
    response_json: dict[str, Any],
) -> int:
    """
    다음 응답 구조에서 Mention 값을 추출한다.

    {
        "data": {
            "rows": [
                [265]
            ]
        }
    }
    """
    if not isinstance(response_json, dict):
        raise TypeError(
            "API 응답은 dict여야 합니다."
        )

    data = response_json.get("data")

    if not isinstance(data, dict):
        raise ValueError(
            "API 응답에 올바른 data object가 없습니다."
        )

    rows = data.get("rows")

    if not isinstance(rows, list):
        raise ValueError(
            "API 응답의 data.rows가 list가 아닙니다."
        )

    if not rows:
        raise ValueError(
            "API 응답의 data.rows가 비어 있습니다."
        )

    first_row = rows[0]

    if not isinstance(first_row, list):
        raise ValueError(
            "API 응답의 첫 번째 row가 list가 아닙니다."
        )

    if not first_row:
        raise ValueError(
            "API 응답의 첫 번째 row가 비어 있습니다."
        )

    raw_mention_count = first_row[0]

    if isinstance(raw_mention_count, bool):
        raise ValueError(
            "Mention 값이 boolean입니다."
        )

    try:
        numeric_mention_count = float(
            raw_mention_count
        )

    except (TypeError, ValueError) as exc:
        raise ValueError(
            "Mention 값을 숫자로 변환할 수 없습니다. "
            f"값: {raw_mention_count!r}"
        ) from exc

    if not numeric_mention_count.is_integer():
        raise ValueError(
            "Mention 값은 정수여야 합니다. "
            f"값: {raw_mention_count!r}"
        )

    mention_count = int(
        numeric_mention_count
    )

    if mention_count < 0:
        raise ValueError(
            "Mention 값은 0 이상이어야 합니다. "
            f"값: {mention_count}"
        )

    return mention_count




def split_into_batches(
    tasks: list[CampaignQueryTask],
    batch_size: int,
) -> list[list[CampaignQueryTask]]:
    if batch_size <= 0:
        raise ValueError(
            "batch_size는 1 이상이어야 합니다."
        )

    return [
        tasks[start_index:start_index + batch_size]
        for start_index in range(0, len(tasks), batch_size)
    ]


def process_campaign_task(
    campaign_task: CampaignQueryTask,
    run_payload_template: dict[str, Any],
) -> CampaignQueryResult:
    """
    캠페인 하나의 Payload 생성, API 호출, Mention 파싱을 담당한다.

    이 함수는 Excel을 읽거나 수정하지 않는다.
    """
    try:
        request_payload = build_query_payload(
            run_payload_template=run_payload_template,
            query=campaign_task.query,
        )

        fetch_result = fetch_sprinklr_data(
            base_url=SPRINKLR_BASE_URL,
            endpoint=ENDPOINT,
            api_key=API_KEY,
            access_token=ACCESS_TOKEN,
            payload=request_payload,
            request_label=f"#={campaign_task.row_id}",
        )

        try:
            mention_count = extract_mention_count(
                response_json=fetch_result.response_json,
            )

        except (TypeError, ValueError) as exc:
            failed_response_path: str | None = None

            if SAVE_FAILED_RESPONSES:
                saved_path = save_response_sample(
                    response_json=fetch_result.response_json,
                    widget_name=(
                        f"failed_Buzz_Volume_row_"
                        f"{campaign_task.row_id}"
                    ),
                    output_dir=RESPONSE_SAMPLE_DIR,
                )
                failed_response_path = str(saved_path)

            return CampaignQueryResult(
                row_id=campaign_task.row_id,
                query=campaign_task.query,
                success=False,
                mention_count=None,
                error_message=(
                    "Sprinklr 응답에서 Mention 값을 "
                    f"추출하지 못했습니다: {exc}"
                ),
                attempt_count=fetch_result.attempt_count,
                status_code=200,
                failed_response_path=failed_response_path,
            )

        return CampaignQueryResult(
            row_id=campaign_task.row_id,
            query=campaign_task.query,
            success=True,
            mention_count=mention_count,
            error_message=None,
            attempt_count=fetch_result.attempt_count,
            status_code=200,
        )

    except SprinklrRequestError as exc:
        return CampaignQueryResult(
            row_id=campaign_task.row_id,
            query=campaign_task.query,
            success=False,
            mention_count=None,
            error_message=str(exc),
            attempt_count=exc.attempt_count,
            status_code=exc.status_code,
        )

    except Exception as exc:
        return CampaignQueryResult(
            row_id=campaign_task.row_id,
            query=campaign_task.query,
            success=False,
            mention_count=None,
            error_message=(
                f"예상하지 못한 캠페인 처리 오류: "
                f"{type(exc).__name__}: {exc}"
            ),
            attempt_count=0,
        )


def process_campaign_batch(
    batch_tasks: list[CampaignQueryTask],
    run_payload_template: dict[str, Any],
    *,
    batch_label: str,
) -> list[CampaignQueryResult]:
    """한 배치의 작업을 최대 MAX_WORKERS개씩 병렬 실행한다."""
    if not batch_tasks:
        return []

    results: list[CampaignQueryResult] = []
    executor = ThreadPoolExecutor(
        max_workers=MAX_WORKERS,
        thread_name_prefix="sprinklr-api",
    )
    future_to_task = {}

    try:
        future_to_task = {
            executor.submit(
                process_campaign_task,
                campaign_task,
                run_payload_template,
            ): campaign_task
            for campaign_task in batch_tasks
        }

        completed_count = 0

        for future in as_completed(future_to_task):
            campaign_task = future_to_task[future]
            completed_count += 1

            try:
                result = future.result()
            except Exception as exc:
                result = CampaignQueryResult(
                    row_id=campaign_task.row_id,
                    query=campaign_task.query,
                    success=False,
                    mention_count=None,
                    error_message=(
                        "작업 스레드에서 처리되지 않은 오류: "
                        f"{type(exc).__name__}: {exc}"
                    ),
                    attempt_count=0,
                )

            results.append(result)

            if result.success:
                safe_print(
                    f"[{batch_label}] "
                    f"[{completed_count}/{len(batch_tasks)}] "
                    f"완료: #={result.row_id}, "
                    f"Mentions={result.mention_count}, "
                    f"시도={result.attempt_count}회"
                )
            else:
                safe_print(
                    f"[{batch_label}] "
                    f"[{completed_count}/{len(batch_tasks)}] "
                    f"실패: #={result.row_id}, "
                    f"시도={result.attempt_count}회, "
                    f"오류={result.error_message}"
                )

    except KeyboardInterrupt:
        _STOP_EVENT.set()
        safe_print(
            f"\n[STOP] {batch_label} 사용자 중단 감지. "
            "시작하지 않은 작업을 취소합니다."
        )

        for future in future_to_task:
            future.cancel()

        executor.shutdown(
            wait=False,
            cancel_futures=True,
        )
        raise

    else:
        executor.shutdown(
            wait=True,
            cancel_futures=False,
        )

    finally:
        close_registered_sessions()

    return results


def merge_campaign_results(
    batch_results: list[CampaignQueryResult],
    mentions_by_row_id: dict[int, int | str],
    failed_results_by_row_id: dict[int, CampaignQueryResult],
) -> None:
    """배치 결과를 성공 dictionary와 실패 dictionary에 반영한다."""
    for result in batch_results:
        if result.success:
            if result.mention_count is None:
                raise RuntimeError(
                    "성공 결과에 mention_count가 없습니다. "
                    f"#={result.row_id}"
                )

            if result.row_id in mentions_by_row_id:
                raise RuntimeError(
                    "성공 Mention 결과 저장 중 중복된 # 값이 "
                    "발견되었습니다. "
                    f"#={result.row_id}"
                )

            mentions_by_row_id[result.row_id] = (
                result.mention_count
            )
            failed_results_by_row_id.pop(
                result.row_id,
                None,
            )

        else:
            if result.row_id not in mentions_by_row_id:
                failed_results_by_row_id[result.row_id] = result


def save_failed_results_report(
    failed_results: list[CampaignQueryResult],
    report_path: Path,
) -> Path:
    report_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    serialized_results = [
        asdict(result)
        for result in sorted(
            failed_results,
            key=lambda item: item.row_id,
        )
    ]

    temporary_report_path = report_path.with_name(
        f".{report_path.stem}_temporary{report_path.suffix}"
    )

    try:
        with temporary_report_path.open(
            mode="w",
            encoding="utf-8",
        ) as file:
            json.dump(
                serialized_results,
                file,
                ensure_ascii=False,
                indent=2,
            )

        temporary_report_path.replace(report_path)

    finally:
        if temporary_report_path.exists():
            temporary_report_path.unlink()

    return report_path


def normalize_blank_query_buzz_volume_cell(
    buzz_volume_cell,
    *,
    excel_row_number: int,
) -> str:
    """
    Query가 비어 있어 API 호출에서 제외되는 행의 기존 Buzz Volume을
    정수 표시 형식으로 정규화한다.

    - 빈 셀: 값 유지, number_format="0"
    - int/정수형 float: int로 변환, number_format="0"
    - 숫자 문자열: int로 변환, number_format="0"
    - Excel 수식: 수식 유지, number_format="0"
    - API_Failed/기타 텍스트: 값 유지, General
    - 소수 값: 오류
    """

    existing_value = buzz_volume_cell.value

    if existing_value is None:
        buzz_volume_cell.number_format = "0"
        return "blank"

    if isinstance(existing_value, bool):
        buzz_volume_cell.number_format = "General"
        return "preserved_text"

    if isinstance(existing_value, (int, float)):
        numeric_value = float(existing_value)

        if not numeric_value.is_integer():
            raise ValueError(
                "Query 공란 행의 기존 Buzz Volume이 정수가 아닙니다.\n"
                f"Excel 행: {excel_row_number}\n"
                f"값: {existing_value!r}"
            )

        buzz_volume_cell.value = int(numeric_value)
        buzz_volume_cell.number_format = "0"
        return "converted_integer"

    if isinstance(existing_value, str):
        stripped_value = existing_value.strip()

        if not stripped_value:
            buzz_volume_cell.number_format = "0"
            return "blank"

        if stripped_value.startswith("="):
            buzz_volume_cell.number_format = "0"
            return "formula"

        if stripped_value == API_FAILED_VALUE:
            buzz_volume_cell.number_format = "General"
            return "preserved_text"

        try:
            numeric_value = float(stripped_value)
        except ValueError:
            buzz_volume_cell.number_format = "General"
            return "preserved_text"

        if not numeric_value.is_integer():
            raise ValueError(
                "Query 공란 행의 기존 Buzz Volume 숫자 문자열이 "
                "정수가 아닙니다.\n"
                f"Excel 행: {excel_row_number}\n"
                f"값: {existing_value!r}"
            )

        buzz_volume_cell.value = int(numeric_value)
        buzz_volume_cell.number_format = "0"
        return "converted_integer"

    buzz_volume_cell.number_format = "0"
    return "other"


def update_mentions_in_excel(
    input_excel_path: Path,
    output_excel_path: Path,
    sheet_name: str,
    mentions_by_row_id: dict[int, int | str],
    *,
    print_summary: bool = True,
) -> list[int]:
    """
    Excel의 '#' 값을 기준으로 Buzz Volume 결과를 매핑한다.
    성공 값은 정수 Mention, 최종 실패 값은 API_Failed로 저장한다.
    """
    if not input_excel_path.exists():
        raise FileNotFoundError(
            "입력 Excel 파일을 찾을 수 없습니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if not input_excel_path.is_file():
        raise FileNotFoundError(
            "입력 Excel 경로가 파일이 아닙니다.\n"
            f"확인 경로: {input_excel_path}"
        )

    if input_excel_path.resolve() == output_excel_path.resolve():
        raise ValueError(
            "입력 Excel과 출력 Excel 경로가 같습니다. "
            "원본 보호를 위해 다른 경로를 사용해야 합니다."
        )

    if not mentions_by_row_id:
        raise ValueError(
            "Excel에 입력할 Mention 결과가 없습니다."
        )

    normalized_mentions: dict[int, int | str] = {}
    zero_mention_row_ids: list[int] = []

    for raw_row_id, raw_mention_count in mentions_by_row_id.items():
        row_id = normalize_row_id(raw_row_id)

        if (
            isinstance(raw_mention_count, str)
            and raw_mention_count.strip() == API_FAILED_VALUE
        ):
            normalized_mentions[row_id] = API_FAILED_VALUE
            continue

        if isinstance(raw_mention_count, bool):
            raise ValueError(
                "Mention 값은 boolean일 수 없습니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            )

        try:
            numeric_mention_count = float(raw_mention_count)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                "Mention 값은 0 이상의 정수 또는 "
                f"{API_FAILED_VALUE!r}여야 합니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            ) from exc

        if not numeric_mention_count.is_integer():
            raise ValueError(
                "Mention 값은 정수여야 합니다. "
                f"#={row_id}, 값={raw_mention_count!r}"
            )

        mention_count = int(numeric_mention_count)

        if mention_count < 0:
            raise ValueError(
                "Mention 값은 0 이상이어야 합니다. "
                f"#={row_id}, 값={mention_count}"
            )

        if mention_count == 0:
            zero_mention_row_ids.append(row_id)

        normalized_mentions[row_id] = mention_count

    keep_vba = input_excel_path.suffix.lower() == ".xlsm"

    workbook = load_workbook(
        filename=input_excel_path,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "대상 시트를 찾을 수 없습니다.\n"
                f"대상 시트: {sheet_name}\n"
                f"현재 시트 목록: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]

        header_cells: dict[str, list[int]] = {
            ROW_ID_COLUMN_NAME: [],
            "Query": [],
            BUZZ_VOLUME_COLUMN_NAME: [],
        }

        # 실제 Excel Header는 "Buzz\nVolume"처럼 셀 내부 줄바꿈이
        # 포함될 수 있으므로, 비교할 때만 공백을 정규화한다.
        # Excel에 표시되는 원래 Header 값과 서식은 변경하지 않는다.
        normalized_header_lookup = {
            normalize_header_name(header_name): header_name
            for header_name in header_cells
        }

        for cell in worksheet[HEADER_ROW]:
            normalized_cell_header = normalize_header_name(
                cell.value
            )

            canonical_header_name = (
                normalized_header_lookup.get(
                    normalized_cell_header
                )
            )

            if canonical_header_name is not None:
                header_cells[canonical_header_name].append(
                    cell.column
                )

        if len(header_cells[ROW_ID_COLUMN_NAME]) != 1:
            raise ValueError(
                f"Header 행에서 {ROW_ID_COLUMN_NAME!r} 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells[ROW_ID_COLUMN_NAME])}"
            )

        if len(header_cells["Query"]) != 1:
            raise ValueError(
                "Header 행에서 'Query' 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells['Query'])}"
            )

        if len(header_cells[BUZZ_VOLUME_COLUMN_NAME]) != 1:
            raise ValueError(
                f"Header 행에서 {BUZZ_VOLUME_COLUMN_NAME!r} 컬럼을 "
                "정확히 1개 찾아야 합니다.\n"
                f"Header 행: {HEADER_ROW}\n"
                "발견 개수: "
                f"{len(header_cells[BUZZ_VOLUME_COLUMN_NAME])}"
            )

        row_id_column_number = (
            header_cells[ROW_ID_COLUMN_NAME][0]
        )
        query_column_number = (
            header_cells["Query"][0]
        )
        buzz_volume_column_number = (
            header_cells[BUZZ_VOLUME_COLUMN_NAME][0]
        )

        excel_row_by_id: dict[int, int] = {}
        invalid_row_id_errors: list[str] = []

        # Query 공란 행은 API 호출 및 Mention 매핑 대상에서는 제외하지만,
        # 기존 Buzz Volume 값과 표시 형식은 정수형으로 정규화한다.
        blank_query_format_counts = Counter()

        for excel_row_number in range(
            HEADER_ROW + 1,
            worksheet.max_row + 1,
        ):
            raw_query = worksheet.cell(
                row=excel_row_number,
                column=query_column_number,
            ).value

            query_is_blank = (
                raw_query is None
                or (
                    isinstance(raw_query, str)
                    and not raw_query.strip()
                )
            )

            if query_is_blank:
                blank_query_buzz_volume_cell = worksheet.cell(
                    row=excel_row_number,
                    column=buzz_volume_column_number,
                )

                normalization_result = (
                    normalize_blank_query_buzz_volume_cell(
                        blank_query_buzz_volume_cell,
                        excel_row_number=excel_row_number,
                    )
                )
                blank_query_format_counts[
                    normalization_result
                ] += 1

                # 기존과 동일하게 API 결과 매핑 대상에서는 제외한다.
                continue

            raw_row_id = worksheet.cell(
                row=excel_row_number,
                column=row_id_column_number,
            ).value

            try:
                row_id = resolve_excel_row_id(
                    value=raw_row_id,
                    excel_row_number=excel_row_number,
                )
            except ValueError as exc:
                invalid_row_id_errors.append(
                    f"Excel 행 {excel_row_number}: {exc}"
                )
                continue

            if row_id in excel_row_by_id:
                raise ValueError(
                    "Query가 입력된 Excel 행의 # 값이 "
                    "중복되었습니다.\n"
                    f"중복 # 값: {row_id}\n"
                    f"첫 번째 행: {excel_row_by_id[row_id]}\n"
                    f"두 번째 행: {excel_row_number}"
                )

            excel_row_by_id[row_id] = excel_row_number

        if invalid_row_id_errors:
            error_preview = "\n".join(invalid_row_id_errors[:20])
            remaining_error_count = max(
                0,
                len(invalid_row_id_errors) - 20,
            )
            extra_message = (
                f"\n외 {remaining_error_count}개 오류"
                if remaining_error_count
                else ""
            )

            raise ValueError(
                "Excel의 # 컬럼 검증에 실패했습니다.\n"
                f"{error_preview}{extra_message}"
            )

        missing_row_ids = sorted(
            set(normalized_mentions)
            - set(excel_row_by_id)
        )

        if missing_row_ids:
            raise ValueError(
                "Mention 결과의 # 값을 Excel에서 찾을 수 없습니다.\n"
                f"미매칭 # 값: {missing_row_ids}"
            )

        updated_count = 0
        newly_filled_count = 0
        overwritten_count = 0
        mention_total = 0
        api_failed_count = 0

        for row_id, mention_value in normalized_mentions.items():
            excel_row_number = excel_row_by_id[row_id]

            buzz_volume_cell = worksheet.cell(
                row=excel_row_number,
                column=buzz_volume_column_number,
            )

            existing_buzz_volume = buzz_volume_cell.value

            if (
                existing_buzz_volume is None
                or (
                    isinstance(existing_buzz_volume, str)
                    and not existing_buzz_volume.strip()
                )
            ):
                newly_filled_count += 1
            else:
                overwritten_count += 1

            # 기존 셀이 비어 있으면 새 값을 입력하고,
            # 기존 숫자나 수식이 있으면 새 API 결과로 덮어쓴다.
            # 셀의 기존 서식은 openpyxl이 그대로 유지한다.
            buzz_volume_cell.value = mention_value

            updated_count += 1
            if mention_value == API_FAILED_VALUE:
                buzz_volume_cell.number_format = "General"
                api_failed_count += 1
            else:
                buzz_volume_cell.number_format = "0"
                mention_total += mention_value

        if updated_count != len(normalized_mentions):
            raise RuntimeError(
                "Mention 업데이트 건수가 예상과 다릅니다.\n"
                f"예상 건수: {len(normalized_mentions)}\n"
                f"실제 건수: {updated_count}"
            )

        output_excel_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        temporary_output_path = output_excel_path.with_name(
            f".{output_excel_path.stem}"
            f"_temporary"
            f"{output_excel_path.suffix}"
        )

        try:
            workbook.save(temporary_output_path)
            temporary_output_path.replace(output_excel_path)
        finally:
            if temporary_output_path.exists():
                temporary_output_path.unlink()

        if print_summary:
            print()
            print("[OK] Buzz Volume Excel 적재 완료")
            print(f"전체 업데이트 행 수: {updated_count}")
            print(f"빈 셀 신규 입력: {newly_filled_count}")
            print(f"기존 값 덮어쓰기: {overwritten_count}")
            print(f"정상 Mention 합계: {mention_total:,}")
            print(f"API_Failed 입력: {api_failed_count}")
            print(
                "Buzz Volume=0 행 수: "
                f"{len(zero_mention_row_ids)}"
            )
            print(
                "Buzz Volume=0 # 목록: "
                f"{sorted(zero_mention_row_ids)}"
            )
            print(
                "Query 공란 행 정수형 처리: "
                f"{sum(blank_query_format_counts.values())}"
            )
            print(
                "  - int 변환: "
                f"{blank_query_format_counts['converted_integer']}"
            )
            print(
                "  - 수식 유지 + 정수 표시: "
                f"{blank_query_format_counts['formula']}"
            )
            print(
                "  - 빈 셀 정수 표시: "
                f"{blank_query_format_counts['blank']}"
            )
            print(f"결과 파일: {output_excel_path}")

    finally:
        workbook.close()

    return sorted(zero_mention_row_ids)


# =============================================================================
# Main
# =============================================================================


def save_checkpoint(
    *,
    input_excel_path: Path,
    output_excel_path: Path,
    sheet_name: str,
    mentions_by_row_id: dict[int, int],
    checkpoint_label: str,
) -> None:
    if not mentions_by_row_id:
        safe_print(
            f"[CHECKPOINT] 체크포인트 건너뜀: {checkpoint_label} "
            "- 저장할 성공 결과가 없습니다."
        )
        return

    update_mentions_in_excel(
        input_excel_path=input_excel_path,
        output_excel_path=output_excel_path,
        sheet_name=sheet_name,
        mentions_by_row_id=mentions_by_row_id,
        print_summary=False,
    )

    safe_print(
        f"[CHECKPOINT] 체크포인트 저장 완료: {checkpoint_label}, "
        f"누적 성공 {len(mentions_by_row_id)}개, "
        f"파일={output_excel_path}"
    )


def main() -> None:
    _STOP_EVENT.clear()

    data_cut_type = input(
        "데이터 컷 유형을 입력하세요 "
        "(daily / weekly): "
    ).strip()

    reference_date_text = input(
        "기준 날짜를 입력하세요 "
        "(YYYY-MM-DD): "
    ).strip()

    date_range, start_time_ms, end_time_ms = build_data_cut(
        data_cut_type=data_cut_type,
        reference_date_text=reference_date_text,
    )

    input_excel_path, output_excel_path = build_excel_paths(
        reference_date_text
    )

    print()
    print("Buzz Volume 파일 경로")
    print(f"- 입력 Excel: {input_excel_path}")
    print(f"- 완료 결과 Excel: {output_excel_path}")
    print(f"- 실패 결과 폴더: {FAILED_RESULTS_DIR}")
    print(f"- 실패 응답 샘플 폴더: {RESPONSE_SAMPLE_DIR}")

    if date_range.data_cut_type == DataCutType.WEEKLY:
        target_sheet_name = "로컬 캠페인 리스트_QHB8 (Weekly)"
    else:
        target_sheet_name = "로컬 캠페인 리스트_QHB8"

    base_payload_path = PAYLOAD_DIR / "buzz_volume_base_payload.json"

    base_payload = load_base_payload(base_payload_path)

    run_payload_template = build_date_payload_template(
        base_payload=base_payload,
        start_time_ms=start_time_ms,
        end_time_ms=end_time_ms,
    )

    campaign_tasks = load_campaign_query(
        excel_path=input_excel_path,
        sheet_name=target_sheet_name,
        date_range=date_range,
    )

    campaign_task_by_row_id = {
        task.row_id: task
        for task in campaign_tasks
    }

    if len(campaign_task_by_row_id) != len(campaign_tasks):
        raise RuntimeError(
            "캠페인 작업 dictionary 생성 중 # 중복이 발견되었습니다."
        )

    mentions_by_row_id: dict[int, int] = {}
    failed_results_by_row_id: dict[
        int,
        CampaignQueryResult,
    ] = {}

    batches = split_into_batches(
        tasks=campaign_tasks,
        batch_size=BATCH_SIZE,
    )

    print()
    print(f"대상 시트: {target_sheet_name}")
    print(f"전체 API 호출 대상: {len(campaign_tasks)}개")
    print(f"배치 크기: {BATCH_SIZE}개")
    print(f"전체 배치 수: {len(batches)}개")
    print(f"동시 작업 수: {MAX_WORKERS}개")
    print(
        "초당 요청 제한: "
        f"{MAX_REQUESTS_PER_SECOND}회/"
        f"{RATE_LIMIT_PERIOD_SECONDS:g}초"
    )
    print(
        "개별 요청 최대 시도 횟수: "
        f"{MAX_RETRIES + 1}회"
    )
    print(
        "동일 Query도 행별로 각각 호출하며, "
        "기존 Buzz Volume은 새 결과로 덮어씁니다."
    )

    for batch_number, batch_tasks in enumerate(
        batches,
        start=1,
    ):
        batch_label = (
            f"Batch {batch_number}/{len(batches)}"
        )

        print()
        print(
            f"[START] {batch_label} 시작 - "
            f"캠페인 {len(batch_tasks)}개"
        )

        batch_results = process_campaign_batch(
            batch_tasks=batch_tasks,
            run_payload_template=run_payload_template,
            batch_label=batch_label,
        )

        merge_campaign_results(
            batch_results=batch_results,
            mentions_by_row_id=mentions_by_row_id,
            failed_results_by_row_id=(
                failed_results_by_row_id
            ),
        )

        batch_success_count = sum(
            result.success
            for result in batch_results
        )
        batch_failure_count = (
            len(batch_results) - batch_success_count
        )

        print(
            f"[OK] {batch_label} 처리 완료 - "
            f"성공 {batch_success_count}개, "
            f"실패 {batch_failure_count}개"
        )

        if SAVE_CHECKPOINT_AFTER_EACH_BATCH:
            save_checkpoint(
                input_excel_path=input_excel_path,
                output_excel_path=output_excel_path,
                sheet_name=target_sheet_name,
                mentions_by_row_id=mentions_by_row_id,
                checkpoint_label=batch_label,
            )

    # 최초 전체 배치 이후에도 실패한 행만 모아 추가 재처리한다.
    for retry_round in range(
        1,
        FINAL_FAILED_RETRY_ROUNDS + 1,
    ):
        if not failed_results_by_row_id:
            break

        retry_tasks = [
            campaign_task_by_row_id[row_id]
            for row_id in sorted(
                failed_results_by_row_id
            )
        ]
        retry_batches = split_into_batches(
            tasks=retry_tasks,
            batch_size=BATCH_SIZE,
        )

        print()
        print(
            f"[RETRY] 최종 실패 행 재처리 {retry_round}/"
            f"{FINAL_FAILED_RETRY_ROUNDS} 시작 - "
            f"대상 {len(retry_tasks)}개"
        )

        # 재처리 성공 결과를 merge할 수 있도록 기존 실패 기록을
        # 유지한 상태에서 각 결과로 갱신한다.
        for retry_batch_number, retry_batch_tasks in enumerate(
            retry_batches,
            start=1,
        ):
            retry_label = (
                f"Retry {retry_round} "
                f"Batch {retry_batch_number}/"
                f"{len(retry_batches)}"
            )

            retry_results = process_campaign_batch(
                batch_tasks=retry_batch_tasks,
                run_payload_template=run_payload_template,
                batch_label=retry_label,
            )

            merge_campaign_results(
                batch_results=retry_results,
                mentions_by_row_id=mentions_by_row_id,
                failed_results_by_row_id=(
                    failed_results_by_row_id
                ),
            )

            if SAVE_CHECKPOINT_AFTER_EACH_BATCH:
                save_checkpoint(
                    input_excel_path=input_excel_path,
                    output_excel_path=output_excel_path,
                    sheet_name=target_sheet_name,
                    mentions_by_row_id=mentions_by_row_id,
                    checkpoint_label=retry_label,
                )

        print(
            f"[RETRY] 재처리 {retry_round} 완료 - "
            f"남은 실패 {len(failed_results_by_row_id)}개"
        )

    expected_row_ids = {
        task.row_id
        for task in campaign_tasks
    }
    actual_row_ids = set(mentions_by_row_id)
    failed_row_ids = set(failed_results_by_row_id)

    unexpected_result_ids = sorted(
        actual_row_ids - expected_row_ids
    )
    overlapping_ids = sorted(
        actual_row_ids & failed_row_ids
    )

    if unexpected_result_ids:
        raise RuntimeError(
            "예상하지 않은 # 값의 API 결과가 생성되었습니다.\n"
            f"예상하지 않은 # 값: {unexpected_result_ids}"
        )

    if overlapping_ids:
        raise RuntimeError(
            "같은 # 값이 성공과 실패 결과에 동시에 존재합니다.\n"
            f"중복 상태 # 값: {overlapping_ids}"
        )

    accounted_row_ids = actual_row_ids | failed_row_ids
    unaccounted_row_ids = sorted(
        expected_row_ids - accounted_row_ids
    )

    if unaccounted_row_ids:
        raise RuntimeError(
            "성공 또는 실패 결과가 생성되지 않은 캠페인이 있습니다.\n"
            f"미처리 # 값: {unaccounted_row_ids}"
        )

    # 성공 결과는 정수로, 최종 실패 결과는 API_Failed로 기록한다.
    # 배치 처리 중에는 실패 셀을 건드리지 않고 모든 재시도 종료 후에만
    # API_Failed로 덮어쓴다.
    final_buzz_values_by_row_id: dict[int, int | str] = dict(
        mentions_by_row_id
    )
    final_buzz_values_by_row_id.update({
        row_id: API_FAILED_VALUE
        for row_id in failed_results_by_row_id
    })

    zero_mention_row_ids: list[int] = []

    if final_buzz_values_by_row_id:
        zero_mention_row_ids = update_mentions_in_excel(
            input_excel_path=input_excel_path,
            output_excel_path=output_excel_path,
            sheet_name=target_sheet_name,
            mentions_by_row_id=final_buzz_values_by_row_id,
            print_summary=True,
        )

    failed_report_path = (
        FAILED_RESULTS_DIR
        / (
            f"{output_excel_path.stem}"
            "_failed_results.json"
        )
    )

    if failed_results_by_row_id:
        saved_failed_report_path = save_failed_results_report(
            failed_results=list(
                failed_results_by_row_id.values()
            ),
            report_path=failed_report_path,
        )

        print()
        print("[WARN] 일부 캠페인 API 처리가 최종 실패했습니다.")
        print(
            f"성공 결과: {len(mentions_by_row_id)}개"
        )
        print(
            f"최종 실패 결과: "
            f"{len(failed_results_by_row_id)}개"
        )
        print(
            "실패 # 값: "
            f"{sorted(failed_results_by_row_id)[:50]}"
        )
        print(
            f"실패 상세 파일: {saved_failed_report_path}"
        )
        if output_excel_path.exists():
            print(
                f"부분 결과 Excel: {output_excel_path}"
            )
            print(
                "최종 실패 행의 Buzz Volume에는 "
                f"{API_FAILED_VALUE}가 입력되었습니다."
            )
        else:
            print(
                "성공 결과가 없어 부분 결과 Excel은 "
                "생성되지 않았습니다."
            )

        raise RuntimeError(
            "일부 캠페인의 Buzz Volume을 추출하지 못했습니다. "
            "성공 결과와 API_Failed 표시는 Excel에 저장했으며, "
            "실패 상세 JSON을 확인하세요."
        )

    if actual_row_ids != expected_row_ids:
        missing_result_ids = sorted(
            expected_row_ids - actual_row_ids
        )
        raise RuntimeError(
            "최종 API 성공 결과와 Excel Query 대상의 # 값이 "
            "일치하지 않습니다.\n"
            f"누락 결과 # 값: {missing_result_ids}"
        )

    if failed_report_path.exists():
        failed_report_path.unlink()

    print()
    print("[OK] 전체 작업 완료")
    print(
        f"데이터 컷: {date_range.start_datetime} "
        f"~ {date_range.end_datetime}"
    )
    print(
        f"전체 API 호출 성공: {len(mentions_by_row_id)}개"
    )
    print(
        f"Buzz Volume=0 행 수: {len(zero_mention_row_ids)}개"
    )
    print(
        f"Buzz Volume=0 # 목록: {zero_mention_row_ids}"
    )
    print(
        f"결과 Excel: {output_excel_path}"
    )


if __name__ == "__main__":
    main()
