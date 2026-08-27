from __future__ import annotations

import argparse
import json
from copy import copy
import mimetypes
import os
import re
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================================================
# 1. Configuration
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
PROMPTS_DIR = BASE_DIR / "prompts"
CONFIG_DIR = BASE_DIR / "config"

ENV_INPUT_DATE = "LOCAL_CAMPAIGN_INPUT_DATE"
ENV_RUN_NUMBER = "LOCAL_CAMPAIGN_RUN_NUMBER"
ENV_OUTPUT_DIR = "LOCAL_CAMPAIGN_OUTPUT_DIR"

DEFAULT_COUNTRY_SUBSIDIARY_MAPPING_FILENAME = (
    "country_subsidiary_mapping.json"
)
DEFAULT_COUNTRY_SUBSIDIARY_MAPPING_PATH = (
    CONFIG_DIR / DEFAULT_COUNTRY_SUBSIDIARY_MAPPING_FILENAME
)
EXPECTED_COUNTRY_MAPPING_COUNT = int(
    os.getenv("EXPECTED_COUNTRY_MAPPING_COUNT", "260")
)
UNASSIGNED_SUBSIDIARY_VALUE = "UNASSIGNED"

# 계정 식별자 기반 확정 규칙
SAMSUNG_ACCOUNT_TOKENS = ("samsung", "삼성")
SAMSUNG_GULF_IDENTIFIER = "samsunggulf"
SAMSUNG_GULF_PUBLISHER_COUNTRY = "Dubai"
SAMSUNG_GULF_EXPECTED_SUBSIDIARY = "SGE"

SAMSUNG_KOREA_IDENTIFIER = "samsungkorea"
SAMSUNG_KOREA_PUBLISHER_COUNTRY = "Korea"
SAMSUNG_KOREA_EXPECTED_SUBSIDIARY = "KOREA"

# 일반 Samsung 계정에서 국가까지 확보된 경우 적용할 최소 신뢰도
SAMSUNG_ACCOUNT_MIN_CONFIDENCE = int(
    os.getenv("SAMSUNG_ACCOUNT_MIN_CONFIDENCE", "90")
)

# media_extractor 결과 파일의 입력 시트
INPUT_SHEET_NAME = "llm_input"

# Gemini 결과를 입력할 formatted Excel의 대상 시트
OUTPUT_SHEET_NAME = "로컬 캠페인 리스트_QHB8"

# 별도로 저장하는 LLM 실행 로그 Excel의 시트
LLM_LOG_SHEET_NAME = "llm_result"

DEFAULT_USER_PROMPT_FILENAME = "user_prompt.txt"
DEFAULT_SYSTEM_PROMPT_FILENAME = "system_prompt.txt"
DEFAULT_RESPONSE_SCHEMA_FILENAME = "response_schema.json"

DEFAULT_TARGET_FILENAME_TEMPLATE = (
    "{input_date}_SLCC_SOV_Local Campaign Tracking_"
    "{input_month}월_v01_formatted.xlsx"
)

# target sheet에서 URL 컬럼을 자동 탐색할 때 사용하는 후보명
TARGET_URL_COLUMN_CANDIDATES = (
    "URL",
    "Url",
    "url",
    "Permalink",
    "Post URL",
    "Original Post URL",
    "게시물 URL",
    "링크",
)

# 같은 URL이 여러 행에 있을 때 추가 검증에 사용하는 ID 컬럼 후보명
TARGET_CAMPAIGN_ID_COLUMN_CANDIDATES = (
    "Campaign ID",
    "campaign_id",
    "Campaign Id",
    "snMsgID",
    "snMsgId",
    "Post ID",
)

TARGET_OUTPUT_COLUMNS = (
    "Campaign Name",
    "Product",
    "CXP Product Feature",
    "Description",
)

# 게시자 유형/국가/법인 판별을 위해 Gemini가 함께 반환하는 내부 결과 컬럼
PUBLISHER_OUTPUT_COLUMNS = (
    "Publisher Type",
    "Publisher Country",
    "Publisher Classification Reason",
    "Publisher Classification Confidence",
    "Requires Manual Review",
)

ALL_RESPONSE_OUTPUT_COLUMNS = (
    *TARGET_OUTPUT_COLUMNS,
    *PUBLISHER_OUTPUT_COLUMNS,
)

TARGET_SUBSIDIARY_COLUMN = (
    "Subsidiary (Country) / Influencer (Subsidiary)"
)
TARGET_INFLUENCER_COLUMN = "Influencer"

TARGET_WRITEBACK_COLUMNS = (
    *TARGET_OUTPUT_COLUMNS,
    TARGET_SUBSIDIARY_COLUMN,
    TARGET_INFLUENCER_COLUMN,
)

ALLOWED_PUBLISHER_TYPES = {
    "OWNED",
    "INFLUENCER",
    "UNKNOWN",
}

# 이 값보다 낮은 게시자 판별 결과는 최종 Excel에 자동 기록하지 않는다.
MIN_AUTO_PUBLISHER_CONFIDENCE = int(
    os.getenv("MIN_AUTO_PUBLISHER_CONFIDENCE", "70")
)

# v9 프롬프트/스키마: 인플루언서 정보와 Giveaway 리워드 줄을 포함할 수 있다.
MAX_DESCRIPTION_LENGTH = int(
    os.getenv("MAX_DESCRIPTION_LENGTH", "180")
)

# header row를 직접 지정하지 않으면 1~이 값까지 탐색
TARGET_HEADER_SCAN_MAX_ROWS = 15

# False이면 formatted Excel에 기존 값이 있는 셀은 보존한다.
OVERWRITE_EXISTING_TARGET_VALUES = False

GOOGLE_CLOUD_PROJECT = os.getenv(
    "GOOGLE_CLOUD_PROJECT",
    "slcc-buzz-agent-dev",
)
GOOGLE_CLOUD_LOCATION = os.getenv(
    "GOOGLE_CLOUD_LOCATION",
    "global",
)
GEMINI_MODEL = os.getenv(
    "GEMINI_MODEL",
    "gemini-2.5-flash",
)

TEMPERATURE = float(os.getenv("GEMINI_TEMPERATURE", "0.1"))
MAX_OUTPUT_TOKENS = int(os.getenv("GEMINI_MAX_OUTPUT_TOKENS", "4096"))
MAX_RETRIES = int(os.getenv("GEMINI_MAX_RETRIES", "3"))

# 로컬 파일을 inline bytes로 보낼 때의 프로그램 자체 안전 제한이다.
MAX_TOTAL_INLINE_MB = int(os.getenv("MAX_TOTAL_INLINE_MB", "450"))
MAX_TOTAL_INLINE_BYTES = MAX_TOTAL_INLINE_MB * 1024 * 1024

SHOW_PROGRESS = True
SHOW_SKIPPED_ROWS = True
SHOW_API_ERRORS = True

SUPPORTED_LOCAL_MEDIA_MIME_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
    ".mp4": "video/mp4",
    ".mov": "video/quicktime",
    ".webm": "video/webm",
    ".m4v": "video/x-m4v",
}

DEFAULT_RESPONSE_JSON_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "Campaign Name": {
            "type": "string",
            "description": (
                "게시물의 텍스트와 모든 미디어를 종합해 작성한 영어 명사형 "
                "캠페인 제목. 공식 캠페인명이 명확하면 우선 사용하고, 그렇지 "
                "않으면 제품·기능·장면·인물·참여 방식을 바탕으로 3~8단어의 "
                "간결한 설명형 제목을 작성한다."
            ),
        },
        "Product": {
            "type": "string",
            "description": (
                "게시물에서 홍보하거나 주요하게 다루는 삼성 제품의 공식 영문명. "
                "QHB8·QH8·Q8·H8·B8 등 내부 코드네임으로 변환하지 않는다. "
                "여러 제품은 쉼표가 아닌 줄바꿈 문자로 구분한다."
            ),
        },
        "CXP Product Feature": {
            "type": "string",
            "description": (
                "제공된 CXP 허용 리스트에서 선택한 핵심 기능. 최대 3개까지 "
                "중요도 순으로 작성하고, 여러 값은 줄바꿈 문자로 구분한다."
            ),
        },
        "Description": {
            "type": "string",
            "maxLength": MAX_DESCRIPTION_LENGTH,
            "description": (
                "한국어 보고서형 설명. 일반 게시물은 한 문장으로 작성하고, "
                "경품 게시물은 본문 다음 줄에 '→ 구체적인 경품 설명' 형식을 "
                "허용한다. 기본 120자 이내를 목표로 하되 최대 180자까지 허용한다."
            ),
        },
        "Publisher Type": {
            "type": "string",
            "enum": ["OWNED", "INFLUENCER", "UNKNOWN"],
            "description": (
                "게시물을 업로드한 계정의 유형. profile_name, profile_url 또는 "
                "sender_screen_name에 대소문자 구분 없이 samsung 또는 삼성이 "
                "포함되면 반드시 OWNED를 반환한다. 게시물 본문이나 콘텐츠 속 "
                "Samsung 언급만으로는 OWNED로 판정하지 않는다. Samsung 공식 "
                "계정의 미디어에 배우·가수·크리에이터·인플루언서가 등장해도 "
                "게시자 유형은 OWNED로 유지한다."
            ),
        },
        "Publisher Country": {
            "type": "string",
            "description": (
                "게시 계정의 국가 또는 확정 계정 매핑 지역. SamsungGulf 계정이면 "
                "정확히 Dubai를 반환하고, Samsung Korea 계정이면 정확히 Korea를 "
                "반환한다. South Korea는 반환하지 않는다. 그 외에는 코드가 제공한 "
                "Country 허용 목록 중 정확한 영문 문자열 하나를 반환하며, 근거가 "
                "부족하면 빈 문자열을 반환한다. Subsidiary Code는 생성하지 않는다."
            ),
        },
        "Publisher Classification Reason": {
            "type": "string",
            "description": (
                "계정명, Screen Name, Profile URL, Website, Bio, Location, "
                "Verification 등 판정에 사용한 핵심 근거를 짧게 작성한다."
            ),
        },
        "Publisher Classification Confidence": {
            "type": "integer",
            "minimum": 0,
            "maximum": 100,
            "description": "게시자 유형과 국가 판정 신뢰도(0~100).",
        },
        "Requires Manual Review": {
            "type": "boolean",
            "description": (
                "게시자 유형 또는 국가를 확정하지 못했거나, 근거가 상충하거나, "
                "핵심 제품·인플루언서·경품 정보가 불명확하면 true를 반환한다."
            ),
        },
    },
    "required": list(ALL_RESPONSE_OUTPUT_COLUMNS),
    "additionalProperties": False,
}


# =========================================================
# 2. Data Models
# =========================================================

@dataclass(frozen=True, slots=True)
class LLMInputSet:
    dataframe_index: int

    campaign_id: str
    source_sheet_name: str
    raw_row_number: int

    platform: str | None
    conversation_stream: str | None
    profile_name: str | None
    profile_url: str | None

    sender_profile_available: bool | None
    sender_screen_name: str | None
    sender_follower_count: str | None
    sender_location: str | None
    sender_detailed_location: str | None
    sender_bio: str | None
    sender_website: str | None
    sender_verified: bool | None
    sender_verified_type: str | None
    sender_profile_tags: str | None

    original_post_url: str | None

    post_media_type: str | None
    media_types: tuple[str, ...]
    media_paths: tuple[Path, ...]

    extraction_status: str | None
    llm_input_mode: str | None
    llm_input_value: str | None
    llm_ready: bool
    user_action_required: bool
    user_action: str | None

    missing_media_paths: tuple[Path, ...] = ()


@dataclass(frozen=True, slots=True)
class PromptBundle:
    user_prompt: str
    system_prompt: str | None
    user_prompt_path: Path
    system_prompt_path: Path | None
    response_schema: dict[str, Any]
    response_schema_path: Path | None


@dataclass(slots=True)
class LLMCallResult:
    dataframe_index: int
    api_status: str
    api_error_message: str | None = None
    raw_llm_response: str | None = None
    parsed_output: dict[str, Any] | None = None
    processed_at_utc: str | None = None
    model_name: str | None = None
    prompt_token_count: int | None = None
    candidates_token_count: int | None = None
    total_token_count: int | None = None

    # 계정명/프로필 URL 기반 deterministic override 추적 정보
    publisher_override_applied: bool = False
    publisher_override_rule: str | None = None
    publisher_override_source_field: str | None = None
    publisher_override_source_value: str | None = None
    publisher_override_original_type: str | None = None
    publisher_override_original_country: str | None = None


@dataclass(frozen=True, slots=True)
class PublisherAccountOverride:
    applied: bool
    rule: str | None = None
    source_field: str | None = None
    source_value: str | None = None
    normalized_identifier: str | None = None
    forced_publisher_type: str | None = None
    forced_publisher_country: str | None = None
    reason: str | None = None


@dataclass(frozen=True, slots=True)
class TargetSheetLayout:
    header_row_number: int
    header_column_map: dict[str, int]
    url_column_name: str
    url_column_number: int
    campaign_id_column_name: str | None
    campaign_id_column_number: int | None


@dataclass(frozen=True, slots=True)
class CountryMappingResult:
    publisher_country: str | None
    mapped_subsidiary: str | None
    status: str
    message: str


@dataclass(frozen=True, slots=True)
class PublisherWritebackDecision:
    publisher_type: str
    publisher_country: str | None
    mapped_subsidiary: str | None
    country_mapping_status: str
    country_mapping_message: str
    subsidiary_display_value: str | None
    influencer_value: str | None
    formatted_url_value: str | None
    formatted_url_target: str | None
    eligible: bool
    skip_reason: str | None
    code_requires_manual_review: bool


# =========================================================
# 3. Basic Normalization Helpers
# =========================================================

URL_PATTERN = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)


def optional_text(value: Any) -> str | None:
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    text = str(value).strip()
    return text or None


def normalize_campaign_id(value: Any) -> str:
    """campaign_id를 비교 가능한 문자열로 변환한다."""

    if value is None:
        raise ValueError("campaign_id가 비어 있습니다.")

    try:
        if pd.isna(value):
            raise ValueError("campaign_id가 비어 있습니다.")
    except (TypeError, ValueError):
        pass

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    if not text:
        raise ValueError("campaign_id가 비어 있습니다.")

    return text


def normalize_integer(value: Any, field_name: str) -> int:
    if value is None:
        raise ValueError(f"{field_name}이 비어 있습니다.")

    try:
        if pd.isna(value):
            raise ValueError(f"{field_name}이 비어 있습니다.")
    except (TypeError, ValueError):
        pass

    try:
        numeric_value = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{field_name}을 정수로 변환할 수 없습니다: {value!r}"
        ) from exc

    if not numeric_value.is_integer():
        raise ValueError(
            f"{field_name}에 정수가 아닌 값이 들어 있습니다: {value!r}"
        )

    return int(numeric_value)


def normalize_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    normalized = str(value).strip().lower()

    if normalized in {"true", "1", "yes", "y", "t"}:
        return True

    if normalized in {"false", "0", "no", "n", "f", ""}:
        return False

    return default


def normalize_optional_bool(value: Any) -> bool | None:
    """Excel의 TRUE/FALSE 문자열을 bool로 변환하고 빈 값은 None으로 유지한다."""

    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False

    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "t"}:
        return True
    if normalized in {"false", "0", "no", "n", "f"}:
        return False

    return None


def split_multiline_values(value: Any) -> tuple[str, ...]:
    text = optional_text(value)
    if text is None:
        return ()

    values: list[str] = []
    seen: set[str] = set()

    for line in text.splitlines():
        normalized = line.strip()
        if not normalized or normalized in seen:
            continue

        seen.add(normalized)
        values.append(normalized)

    return tuple(values)


def is_unknown_media_input(
    input_set: LLMInputSet,
) -> bool:
    """
    media_extractor 결과에서 게시물 또는 개별 미디어 타입이 UNKNOWN인지 확인한다.

    UNKNOWN은 "분석 불가"를 의미하지 않는다.
    미디어 추출이 완전하지 않더라도 게시물 텍스트, permalink,
    계정 메타데이터와 확보된 유효 미디어를 이용해 1차 LLM 분석을 수행한다.
    """

    post_media_type = (
        input_set.post_media_type or ""
    ).strip().lower()

    media_types = {
        str(media_type).strip().lower()
        for media_type in input_set.media_types
        if str(media_type).strip()
    }

    return (
        post_media_type == "unknown"
        or "unknown" in media_types
    )


def normalize_cxp_feature_value(value: str) -> str:
    """
    CXP Product Feature 값을 Excel 셀 내부 줄바꿈 형식으로 정규화한다.

    지원 입력 예:
        Camera_Horizontal Lock, Camera_Super Steady
        Camera_Horizontal Lock\nCamera_Super Steady
        - Camera_Horizontal Lock\n- Camera_Super Steady

    최종 저장 예:
        Camera_Horizontal Lock
        Camera_Super Steady
    """

    if not value.strip():
        return ""

    # 쉼표, 세미콜론, 실제 줄바꿈을 모두 항목 구분자로 처리
    raw_items = re.split(r"[\r\n,;]+", value)

    normalized_items: list[str] = []
    seen: set[str] = set()

    for raw_item in raw_items:
        item = raw_item.strip()

        # bullet 또는 번호 목록 기호 제거
        item = re.sub(
            r"^(?:[-*•]|\d+[.)])\s*",
            "",
            item,
        ).strip()

        if not item or item in seen:
            continue

        seen.add(item)
        normalized_items.append(item)

    # 프롬프트 규칙에 따라 최대 3개까지만 저장
    return "\n".join(normalized_items[:3])



def normalize_product_value(value: str) -> str:
    """복수 Product를 Excel 셀 내부 줄바꿈 형식으로 정규화한다.

    Gemini가 프롬프트를 어겨 쉼표/세미콜론/bullet을 사용하더라도
    최종 Excel에는 제품별 한 줄 형식으로 저장한다.
    """

    if not value.strip():
        return ""

    normalized_source = value.replace("\\n", "\n")
    raw_items = re.split(r"[\r\n,;]+", normalized_source)

    normalized_items: list[str] = []
    seen: set[str] = set()

    for raw_item in raw_items:
        item = re.sub(
            r"^(?:[-*•]|\d+[.)])\s*",
            "",
            raw_item.strip(),
        ).strip()

        # 제품명 내부의 불필요한 연속 공백만 정리하고 표기는 바꾸지 않는다.
        item = " ".join(item.split())
        if not item or item in seen:
            continue

        seen.add(item)
        normalized_items.append(item)

    return "\n".join(normalized_items)


def normalize_description_value(value: str) -> str:
    """Description의 본문과 Giveaway 리워드 줄을 안전하게 정규화한다.

    허용 형태:
        일반 게시물: 본문 한 문장
        Giveaway: 본문\n→ 구체적인 경품 설명

    그 밖의 임의 줄바꿈은 공백으로 합친다.
    """

    if not value.strip():
        return ""

    normalized_source = (
        value.replace("\\n", "\n")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
    )

    # 첫 번째 화살표를 기준으로 본문과 리워드를 분리한다.
    arrow_parts = re.split(r"\s*→\s*", normalized_source, maxsplit=1)
    main_text = " ".join(arrow_parts[0].split())

    if len(arrow_parts) == 1:
        return main_text[:MAX_DESCRIPTION_LENGTH].rstrip()

    reward_text = " ".join(arrow_parts[1].split())
    if not reward_text:
        return main_text[:MAX_DESCRIPTION_LENGTH].rstrip()

    prefix = f"{main_text}\n→ "
    if len(prefix) >= MAX_DESCRIPTION_LENGTH:
        return main_text[:MAX_DESCRIPTION_LENGTH].rstrip()

    remaining_length = MAX_DESCRIPTION_LENGTH - len(prefix)
    reward_text = reward_text[:remaining_length].rstrip()
    return f"{prefix}{reward_text}".rstrip()



def normalize_account_identifier(value: Any) -> str:
    """계정명·프로필 URL을 계정 규칙 비교용 문자열로 정규화한다.

    대소문자를 무시하고 공백, 마침표, 밑줄, 하이픈, URL 기호를 제거한다.
    한글은 보존하므로 '삼성'도 직접 감지할 수 있다.

    예:
        Samsung Gulf -> samsunggulf
        samsung_gulf -> samsunggulf
        https://instagram.com/SamsungGulf -> httpsinstagramcomsamsunggulf
    """

    text = optional_text(value)
    if text is None:
        return ""

    return re.sub(r"[\W_]+", "", text.casefold(), flags=re.UNICODE)


def detect_publisher_account_override(
    input_set: LLMInputSet,
) -> PublisherAccountOverride:
    """게시 계정 관련 필드에서 Samsung 계정 확정 규칙을 찾는다.

    검사 범위는 게시 계정 필드로 한정한다.
    게시물 본문, 캡션, 해시태그, 미디어 텍스트는 검사하지 않는다.

    우선순위:
    1. SamsungGulf -> OWNED / Dubai
    2. SamsungKorea -> OWNED / Korea
    3. 일반 Samsung 또는 삼성 포함 -> OWNED
    """

    account_fields = (
        ("profile_url", input_set.profile_url),
        ("profile_name", input_set.profile_name),
        ("sender_screen_name", input_set.sender_screen_name),
    )

    normalized_fields = [
        (
            field_name,
            optional_text(raw_value),
            normalize_account_identifier(raw_value),
        )
        for field_name, raw_value in account_fields
    ]

    # 가장 구체적인 SamsungGulf 규칙부터 검사한다.
    for field_name, raw_value, normalized_value in normalized_fields:
        if (
            raw_value is not None
            and SAMSUNG_GULF_IDENTIFIER in normalized_value
        ):
            return PublisherAccountOverride(
                applied=True,
                rule="samsung_gulf_fixed",
                source_field=field_name,
                source_value=raw_value,
                normalized_identifier=normalized_value,
                forced_publisher_type="OWNED",
                forced_publisher_country=(
                    SAMSUNG_GULF_PUBLISHER_COUNTRY
                ),
                reason=(
                    f"{field_name}에서 SamsungGulf 계정을 확인하여 "
                    "OWNED / Dubai 고정 규칙을 적용함"
                ),
            )

    # Samsung Korea 계정은 일반 Samsung 규칙보다 먼저 검사한다.
    for field_name, raw_value, normalized_value in normalized_fields:
        if (
            raw_value is not None
            and SAMSUNG_KOREA_IDENTIFIER in normalized_value
        ):
            return PublisherAccountOverride(
                applied=True,
                rule="samsung_korea_fixed",
                source_field=field_name,
                source_value=raw_value,
                normalized_identifier=normalized_value,
                forced_publisher_type="OWNED",
                forced_publisher_country=(
                    SAMSUNG_KOREA_PUBLISHER_COUNTRY
                ),
                reason=(
                    f"{field_name}에서 Samsung Korea 계정을 확인하여 "
                    "OWNED / Korea 고정 규칙을 적용함"
                ),
            )

    # 일반 Samsung/삼성 계정 규칙
    for field_name, raw_value, normalized_value in normalized_fields:
        if raw_value is None:
            continue

        if any(
            token in normalized_value
            for token in SAMSUNG_ACCOUNT_TOKENS
        ):
            return PublisherAccountOverride(
                applied=True,
                rule="samsung_account_owned",
                source_field=field_name,
                source_value=raw_value,
                normalized_identifier=normalized_value,
                forced_publisher_type="OWNED",
                forced_publisher_country=None,
                reason=(
                    f"{field_name}에 Samsung/삼성 계정 식별자가 있어 "
                    "Publisher Type을 OWNED로 확정함"
                ),
            )

    # Samsung/삼성 식별자가 없는 모든 외부 게시 계정은
    # 계정 성격(개인, 매체, 통신사, 유통사, 기업 등)과 관계없이
    # INFLUENCER로 고정한다.
    representative_field_name = None
    representative_raw_value = None
    representative_normalized_value = None

    for field_name, raw_value, normalized_value in normalized_fields:
        if raw_value is not None:
            representative_field_name = field_name
            representative_raw_value = raw_value
            representative_normalized_value = normalized_value
            break

    return PublisherAccountOverride(
        applied=True,
        rule="non_samsung_influencer",
        source_field=representative_field_name,
        source_value=representative_raw_value,
        normalized_identifier=representative_normalized_value,
        forced_publisher_type="INFLUENCER",
        forced_publisher_country=None,
        reason=(
            "게시 계정 관련 필드에서 Samsung/삼성 식별자가 확인되지 않아 "
            "비Samsung 계정 규칙에 따라 Publisher Type을 "
            "INFLUENCER로 확정함"
        ),
    )


def apply_publisher_account_override(
    parsed_output: dict[str, Any],
    input_set: LLMInputSet,
) -> tuple[dict[str, Any], PublisherAccountOverride]:
    """Gemini 결과에 계정 식별자 기반 확정 규칙을 적용한다."""

    override = detect_publisher_account_override(input_set)
    if not override.applied:
        return dict(parsed_output), override

    overridden = dict(parsed_output)

    original_type = optional_text(overridden.get("Publisher Type"))
    original_country = optional_text(
        overridden.get("Publisher Country")
    )
    existing_reason = optional_text(
        overridden.get("Publisher Classification Reason")
    )

    overridden["Publisher Type"] = (
        override.forced_publisher_type or "OWNED"
    )

    if override.forced_publisher_country:
        # SamsungGulf / SamsungKorea: Country까지 고정
        overridden["Publisher Country"] = (
            override.forced_publisher_country
        )
        overridden["Publisher Classification Confidence"] = 100
        overridden["Requires Manual Review"] = False
    elif overridden["Publisher Type"] == "OWNED":
        # 일반 Samsung 계정: 유형은 확정하되 국가는 AI 결과를 유지한다.
        if original_country:
            try:
                current_confidence = int(
                    overridden.get(
                        "Publisher Classification Confidence",
                        0,
                    )
                )
            except (TypeError, ValueError):
                current_confidence = 0

            overridden["Publisher Classification Confidence"] = max(
                current_confidence,
                SAMSUNG_ACCOUNT_MIN_CONFIDENCE,
            )
            # 전체 결과의 다른 불확실성은 보존한다. 다만 게시자 write-back
            # 단계에서는 deterministic override를 별도로 인식한다.
        else:
            overridden["Publisher Country"] = ""
            overridden["Requires Manual Review"] = True
    else:
        # 비Samsung 계정은 Publisher Type만 INFLUENCER로 확정한다.
        # Country, Confidence, Manual Review는 Gemini의 다른 판단 결과를
        # 그대로 유지하여 국가/인물/경품 등의 불확실성을 숨기지 않는다.
        overridden["Publisher Country"] = original_country or ""

    reason_parts = [override.reason or "Samsung 계정 규칙 적용"]
    if original_type or original_country:
        reason_parts.append(
            "Gemini 원판단="
            f"type:{original_type or 'blank'}, "
            f"country:{original_country or 'blank'}"
        )
    if existing_reason:
        reason_parts.append(f"기존 근거:{existing_reason}")

    overridden["Publisher Classification Reason"] = "; ".join(
        reason_parts
    )

    return overridden, override


def validate_special_account_mappings(
    country_to_subsidiary: dict[str, str],
) -> None:
    """고정 Samsung 계정 Country가 기대 Subsidiary로 매핑되는지 확인한다."""

    required_mappings = (
        (
            "SamsungGulf",
            SAMSUNG_GULF_PUBLISHER_COUNTRY,
            SAMSUNG_GULF_EXPECTED_SUBSIDIARY,
        ),
        (
            "SamsungKorea",
            SAMSUNG_KOREA_PUBLISHER_COUNTRY,
            SAMSUNG_KOREA_EXPECTED_SUBSIDIARY,
        ),
    )

    for account_name, publisher_country, expected_subsidiary in required_mappings:
        actual_subsidiary = country_to_subsidiary.get(
            publisher_country
        )

        if actual_subsidiary != expected_subsidiary:
            raise ValueError(
                f"{account_name} 고정 매핑이 올바르지 않습니다. "
                f"expected={publisher_country} -> "
                f"{expected_subsidiary}, "
                f"actual={actual_subsidiary!r}"
            )

def json_safe_excel_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, (str, int, float, bool)):
        return value

    return json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
    )


def is_blank_excel_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def clean_url_candidate(value: str) -> str:
    return value.strip().rstrip(".,;:)]}>'\"")


def extract_url_strings(value: Any) -> tuple[str, ...]:
    text = optional_text(value)
    if text is None:
        return ()

    matches = [clean_url_candidate(item) for item in URL_PATTERN.findall(text)]
    if matches:
        return tuple(dict.fromkeys(item for item in matches if item))

    if text.lower().startswith(("http://", "https://")):
        return (clean_url_candidate(text),)

    return ()


def split_post_and_comment_urls(
    value: Any,
) -> tuple[str | None, str | None]:
    """
    URL 셀/문자열에서 원 게시물 URL과 소비자 댓글 URL을 분리한다.

    raw_to_processed.py의 URL 셀 규칙:
        POST_URL
        COMMENT_URL

    최종 llm_analysis write-back 이후 규칙:
        [당사 게시글] 또는 [인플루언서 게시글]
        POST_URL

        [소비자 반응]
        COMMENT_URL

    두 형식 모두 URL 등장 순서는 동일하므로:
    - 첫 번째 URL = 원 게시물
    - 두 번째 URL = 소비자 댓글
    로 해석한다.

    댓글 URL이 없으면 두 번째 반환값은 None이다.
    """
    urls = extract_url_strings(value)

    if not urls:
        return None, None

    post_url = urls[0]
    comment_url = (
        urls[1]
        if len(urls) >= 2
        else None
    )

    return post_url, comment_url


def normalize_post_url(value: Any) -> str | None:
    """플랫폼별 게시물 URL을 비교 가능한 canonical URL로 변환한다."""

    text = optional_text(value)
    if text is None:
        return None

    if not text.lower().startswith(("http://", "https://")):
        text = "https://" + text

    parsed = urlparse(text)
    host = parsed.netloc.lower().split("@")[ -1].split(":")[0]
    path = re.sub(r"/{2,}", "/", parsed.path or "/")
    query = parse_qs(parsed.query)

    host_aliases = {
        "twitter.com": "x.com",
        "www.twitter.com": "x.com",
        "mobile.twitter.com": "x.com",
        "www.x.com": "x.com",
        "mobile.x.com": "x.com",
        "www.instagram.com": "instagram.com",
        "m.instagram.com": "instagram.com",
        "www.youtube.com": "youtube.com",
        "m.youtube.com": "youtube.com",
        "music.youtube.com": "youtube.com",
        "www.facebook.com": "facebook.com",
        "m.facebook.com": "facebook.com",
        "web.facebook.com": "facebook.com",
    }
    host = host_aliases.get(host, host)

    # X / Twitter는 username과 무관하게 status ID로 통일한다.
    if host == "x.com":
        status_match = re.search(r"/status/(\d+)", path, re.IGNORECASE)
        if status_match:
            return f"https://x.com/i/status/{status_match.group(1)}"

    # Instagram post/reel/tv는 shortcode로 통일한다.
    if host == "instagram.com":
        instagram_match = re.search(
            r"/(p|reel|reels|tv)/([^/?#]+)",
            path,
            re.IGNORECASE,
        )
        if instagram_match:
            media_kind = instagram_match.group(1).lower()
            if media_kind == "reels":
                media_kind = "reel"
            shortcode = instagram_match.group(2)
            return f"https://instagram.com/{media_kind}/{shortcode}"

    # YouTube URL을 video ID 중심으로 통일한다.
    if host == "youtu.be":
        video_id = path.strip("/").split("/")[0]
        if video_id:
            return f"https://youtube.com/watch?v={video_id}"

    if host == "youtube.com":
        video_id: str | None = None
        if path.rstrip("/") == "/watch":
            video_id = (query.get("v") or [None])[0]
        else:
            youtube_match = re.search(
                r"/(?:shorts|embed|live)/([^/?#]+)",
                path,
                re.IGNORECASE,
            )
            if youtube_match:
                video_id = youtube_match.group(1)

        if video_id:
            return f"https://youtube.com/watch?v={video_id}"

    # Facebook의 일부 게시물 URL은 식별자가 query에만 존재한다.
    if host == "facebook.com":
        clean_path = path.rstrip("/") or "/"
        preserved_query: dict[str, str] = {}

        if clean_path.endswith("/permalink.php"):
            for key in ("story_fbid", "id"):
                value_list = query.get(key)
                if value_list and value_list[0]:
                    preserved_query[key] = value_list[0]
        elif clean_path.endswith("/photo.php"):
            value_list = query.get("fbid")
            if value_list and value_list[0]:
                preserved_query["fbid"] = value_list[0]
        elif clean_path in {"/watch", "/watch/"}:
            value_list = query.get("v")
            if value_list and value_list[0]:
                preserved_query["v"] = value_list[0]

        return urlunparse(
            (
                "https",
                host,
                clean_path,
                "",
                urlencode(preserved_query),
                "",
            )
        )

    clean_path = path.rstrip("/") or "/"
    return urlunparse(("https", host, clean_path, "", "", ""))


def extract_normalized_post_urls(value: Any) -> tuple[str, ...]:
    normalized_urls: list[str] = []
    seen: set[str] = set()

    for raw_url in extract_url_strings(value):
        normalized_url = normalize_post_url(raw_url)
        if normalized_url and normalized_url not in seen:
            seen.add(normalized_url)
            normalized_urls.append(normalized_url)

    return tuple(normalized_urls)


def load_country_subsidiary_mapping(
    mapping_path: str | Path,
) -> dict[str, str]:
    """확정 Excel에서 생성한 Country → Subsidiary JSON을 검증해 읽는다.

    강제 규칙:
    - Country는 JSON에 있는 문자열을 그대로 key로 사용한다.
    - 별칭, 대소문자 보정, 국가명 통합을 하지 않는다.
    - Subsidiary는 모두 대문자여야 한다.
    - UNASSIGNED도 매핑에 포함하되 자동 write-back에는 사용하지 않는다.
    """

    mapping_path = Path(mapping_path)
    if not mapping_path.is_file():
        raise FileNotFoundError(
            "Country–Subsidiary 매핑 JSON을 찾을 수 없습니다: "
            f"{mapping_path}"
        )

    try:
        payload = json.loads(
            mapping_path.read_text(encoding="utf-8-sig")
        )
    except json.JSONDecodeError as exc:
        raise ValueError(
            "Country–Subsidiary 매핑 JSON 문법이 올바르지 않습니다: "
            f"{mapping_path}"
        ) from exc

    if not isinstance(payload, dict):
        raise ValueError("Country–Subsidiary JSON 최상위 값은 object여야 합니다.")

    raw_mapping = payload.get("country_to_subsidiary")
    if not isinstance(raw_mapping, dict):
        raise ValueError(
            "Country–Subsidiary JSON에 country_to_subsidiary object가 없습니다."
        )

    country_to_subsidiary: dict[str, str] = {}

    for raw_country, raw_subsidiary in raw_mapping.items():
        if not isinstance(raw_country, str):
            raise ValueError(
                "Country key는 문자열이어야 합니다: "
                f"{raw_country!r}"
            )
        if not isinstance(raw_subsidiary, str):
            raise ValueError(
                "Subsidiary 값은 문자열이어야 합니다: "
                f"{raw_country!r} -> {raw_subsidiary!r}"
            )

        country = raw_country.strip()
        subsidiary = raw_subsidiary.strip()

        if not country:
            raise ValueError("Country–Subsidiary 매핑에 빈 Country key가 있습니다.")
        if country != raw_country:
            raise ValueError(
                "Country key 앞뒤에 공백이 있습니다. Excel 표기를 그대로 "
                f"정리한 JSON을 사용하세요: {raw_country!r}"
            )
        if not subsidiary:
            raise ValueError(
                f"Subsidiary 값이 비어 있습니다: {country}"
            )
        if subsidiary != subsidiary.upper():
            raise ValueError(
                "Subsidiary Code는 모두 대문자여야 합니다: "
                f"{country} -> {subsidiary}"
            )
        if country in country_to_subsidiary:
            raise ValueError(
                f"동일한 Country key가 중복되었습니다: {country}"
            )

        country_to_subsidiary[country] = subsidiary

    metadata = payload.get("metadata")
    if metadata is not None and not isinstance(metadata, dict):
        raise ValueError("Country–Subsidiary JSON metadata는 object여야 합니다.")

    metadata_total = None
    if isinstance(metadata, dict):
        metadata_total = metadata.get("total_country_count")

    if metadata_total is not None:
        try:
            metadata_total_int = int(metadata_total)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                "metadata.total_country_count를 정수로 해석할 수 없습니다: "
                f"{metadata_total!r}"
            ) from exc

        if metadata_total_int != len(country_to_subsidiary):
            raise ValueError(
                "metadata.total_country_count와 실제 매핑 개수가 다릅니다: "
                f"metadata={metadata_total_int}, actual={len(country_to_subsidiary)}"
            )

    if (
        EXPECTED_COUNTRY_MAPPING_COUNT > 0
        and len(country_to_subsidiary) != EXPECTED_COUNTRY_MAPPING_COUNT
    ):
        raise ValueError(
            "Country–Subsidiary 매핑 개수가 예상값과 다릅니다: "
            f"actual={len(country_to_subsidiary)}, "
            f"expected={EXPECTED_COUNTRY_MAPPING_COUNT}"
        )

    return country_to_subsidiary


def summarize_country_subsidiary_mapping(
    country_to_subsidiary: dict[str, str],
) -> dict[str, int]:
    unassigned_count = sum(
        1
        for subsidiary in country_to_subsidiary.values()
        if subsidiary == UNASSIGNED_SUBSIDIARY_VALUE
    )
    return {
        "total": len(country_to_subsidiary),
        "assigned": len(country_to_subsidiary) - unassigned_count,
        "unassigned": unassigned_count,
    }


def resolve_country_subsidiary(
    publisher_country: Any,
    country_to_subsidiary: dict[str, str],
) -> CountryMappingResult:
    """AI가 반환한 Country를 별칭 처리 없이 정확히 조회한다."""

    country = optional_text(publisher_country)
    if country is None:
        return CountryMappingResult(
            publisher_country=None,
            mapped_subsidiary=None,
            status="country_blank",
            message="Publisher Country가 비어 있습니다.",
        )

    if country not in country_to_subsidiary:
        return CountryMappingResult(
            publisher_country=country,
            mapped_subsidiary=None,
            status="country_not_found",
            message=(
                "Publisher Country가 확정 매핑 JSON의 Country key와 "
                f"정확히 일치하지 않습니다: {country}"
            ),
        )

    subsidiary = country_to_subsidiary[country]
    if subsidiary == UNASSIGNED_SUBSIDIARY_VALUE:
        return CountryMappingResult(
            publisher_country=country,
            mapped_subsidiary=subsidiary,
            status="unassigned",
            message=(
                "해당 Country의 Subsidiary가 UNASSIGNED이므로 "
                "자동 기록하지 않습니다."
            ),
        )

    return CountryMappingResult(
        publisher_country=country,
        mapped_subsidiary=subsidiary,
        status="mapped",
        message=f"정확 일치 매핑 성공: {country} -> {subsidiary}",
    )


def build_allowed_country_instruction(
    country_to_subsidiary: dict[str, str],
) -> str:
    """AI가 Excel의 Country 문자열만 반환하도록 허용 목록을 만든다."""

    country_lines = "\n".join(
        f"- {country}"
        for country in country_to_subsidiary
    )

    return (
        "Publisher Country는 아래 Country 이름 중 정확히 하나를 "
        "대소문자와 문장부호까지 그대로 반환하거나, 근거가 부족하면 "
        "빈 문자열을 반환한다. 약어, 번역, 별칭, 자체 정규화는 금지한다.\n"
        "[ALLOWED PUBLISHER COUNTRY VALUES]\n"
        f"{country_lines}"
    )

# =========================================================
# 4. Unified llm_input Excel -> LLMInputSet
# =========================================================

REQUIRED_INPUT_COLUMNS = {
    "campaign_id",
    "source_sheet",
    "raw_row_number",
    "platform",
    "permalink",
    "status",
    "llm_input_mode",
    "llm_ready",
    "user_action_required",
}

OPTIONAL_SENDER_PROFILE_COLUMNS = (
    "sender_profile_available",
    "sender_screen_name",
    "sender_follower_count",
    "sender_location",
    "sender_detailed_location",
    "sender_bio",
    "sender_website",
    "sender_verified",
    "sender_verified_type",
    "sender_profile_tags",
)


def load_llm_input_dataframe(
    result_excel_path: str | Path,
) -> pd.DataFrame:
    result_excel_path = Path(result_excel_path)

    if not result_excel_path.is_file():
        raise FileNotFoundError(
            f"미디어 결과 Excel 파일을 찾을 수 없습니다: {result_excel_path}"
        )

    try:
        dataframe = pd.read_excel(
            result_excel_path,
            sheet_name=INPUT_SHEET_NAME,
            dtype={
                "campaign_id": "string",
                "source_sheet": "string",
            },
        )
    except ValueError as exc:
        raise ValueError(
            f"'{INPUT_SHEET_NAME}' 시트를 읽을 수 없습니다. "
            "최신 통합 media_extractor 결과 파일인지 확인하세요."
        ) from exc

    dataframe.columns = dataframe.columns.astype(str).str.strip()

    missing_columns = sorted(
        REQUIRED_INPUT_COLUMNS - set(dataframe.columns)
    )
    if missing_columns:
        raise ValueError(
            f"'{INPUT_SHEET_NAME}' 시트에 필수 컬럼이 없습니다: "
            f"{missing_columns}"
        )

    # media_extractor에서 과거 컬럼명 sender_follower를 사용한 경우에도
    # 최신 표준명 sender_follower_count로 호환한다.
    if (
        "sender_follower_count" not in dataframe.columns
        and "sender_follower" in dataframe.columns
    ):
        dataframe["sender_follower_count"] = dataframe["sender_follower"]

    # 과거 media_extractor 결과도 읽을 수 있도록 Sender Profile 컬럼은
    # 선택 컬럼으로 취급하되, 최신 결과에서는 그대로 LLM까지 전달한다.
    for column_name in OPTIONAL_SENDER_PROFILE_COLUMNS:
        if column_name not in dataframe.columns:
            dataframe[column_name] = pd.NA

    # 데이터가 0건이어도 오류로 중단하지 않는다. 이후 단계에서 헤더만 있는
    # LLM 로그와 원본을 복사한 최종 formatted Excel을 생성할 수 있다.
    return dataframe.copy()


def build_llm_input_sets_from_dataframe(
    dataframe: pd.DataFrame,
) -> list[LLMInputSet]:
    """통합 llm_input 시트의 post 1행을 LLMInputSet 1개로 변환한다."""

    llm_input_sets: list[LLMInputSet] = []

    for dataframe_index, row in dataframe.iterrows():
        campaign_id = normalize_campaign_id(row.get("campaign_id"))
        source_sheet_name = optional_text(row.get("source_sheet"))

        if source_sheet_name is None:
            raise ValueError(
                f"source_sheet가 비어 있습니다: Excel row={dataframe_index + 2}"
            )

        raw_row_number = normalize_integer(
            row.get("raw_row_number"),
            "raw_row_number",
        )

        media_path_strings = split_multiline_values(
            row.get("local_media_paths")
        )
        media_paths = tuple(Path(path_text) for path_text in media_path_strings)
        missing_media_paths = tuple(
            path for path in media_paths if not path.is_file()
        )

        # raw_to_processed.py의 URL 셀은
        #   POST_URL
        #   COMMENT_URL
        # 형식일 수 있다.
        # LLM 분석에는 반드시 원 게시물 URL만 전달한다.
        original_post_url, _consumer_comment_url = (
            split_post_and_comment_urls(
                row.get("permalink")
            )
        )

        llm_input_sets.append(
            LLMInputSet(
                dataframe_index=int(dataframe_index),
                campaign_id=campaign_id,
                source_sheet_name=source_sheet_name,
                raw_row_number=raw_row_number,
                platform=optional_text(row.get("platform")),
                conversation_stream=optional_text(
                    row.get("conversation_stream")
                ),
                profile_name=optional_text(row.get("user_name")),
                profile_url=optional_text(row.get("profile_url")),
                sender_profile_available=normalize_optional_bool(
                    row.get("sender_profile_available")
                ),
                sender_screen_name=optional_text(
                    row.get("sender_screen_name")
                ),
                sender_follower_count=optional_text(
                    row.get("sender_follower_count")
                ),
                sender_location=optional_text(row.get("sender_location")),
                sender_detailed_location=optional_text(
                    row.get("sender_detailed_location")
                ),
                sender_bio=optional_text(row.get("sender_bio")),
                sender_website=optional_text(row.get("sender_website")),
                sender_verified=normalize_optional_bool(
                    row.get("sender_verified")
                ),
                sender_verified_type=optional_text(
                    row.get("sender_verified_type")
                ),
                sender_profile_tags=optional_text(
                    row.get("sender_profile_tags")
                ),
                original_post_url=original_post_url,
                post_media_type=optional_text(row.get("post_media_type")),
                media_types=split_multiline_values(row.get("media_types")),
                media_paths=media_paths,
                extraction_status=optional_text(row.get("status")),
                llm_input_mode=optional_text(row.get("llm_input_mode")),
                llm_input_value=optional_text(row.get("llm_input_value")),
                llm_ready=normalize_bool(
                    row.get("llm_ready"),
                    default=False,
                ),
                user_action_required=normalize_bool(
                    row.get("user_action_required"),
                    default=False,
                ),
                user_action=optional_text(row.get("user_action")),
                missing_media_paths=missing_media_paths,
            )
        )

    return llm_input_sets


def build_llm_input_sets(
    result_excel_path: str | Path,
) -> list[LLMInputSet]:
    dataframe = load_llm_input_dataframe(result_excel_path)
    return build_llm_input_sets_from_dataframe(dataframe)


# =========================================================
# 5. Prompt / Response Schema Loading
# =========================================================


def read_text_file(path: Path) -> str:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(f"텍스트 파일을 찾을 수 없습니다: {path}")

    text = path.read_text(encoding="utf-8-sig").strip()
    if not text:
        raise ValueError(f"텍스트 파일이 비어 있습니다: {path}")

    return text


def resolve_user_prompt_path(
    prompts_dir: Path,
    explicit_prompt_path: Path | None,
) -> Path:
    if explicit_prompt_path is not None:
        return Path(explicit_prompt_path)

    default_path = prompts_dir / DEFAULT_USER_PROMPT_FILENAME
    if default_path.is_file():
        return default_path

    candidate_paths = sorted(
        path
        for path in prompts_dir.glob("*.txt")
        if path.name != DEFAULT_SYSTEM_PROMPT_FILENAME
    )

    if len(candidate_paths) == 1:
        return candidate_paths[0]

    if not candidate_paths:
        raise FileNotFoundError(
            f"prompts 폴더에 사용자 프롬프트 txt 파일이 없습니다: {prompts_dir}"
        )

    raise ValueError(
        "사용자 프롬프트 txt 파일이 여러 개입니다. "
        "--prompt-file로 사용할 파일을 명시하세요: "
        f"{[path.name for path in candidate_paths]}"
    )


def validate_response_schema(
    schema: dict[str, Any],
) -> None:
    """v9의 정확한 9개 응답 필드와 타입을 검증한다."""

    if str(schema.get("type", "")).lower() != "object":
        raise ValueError("response schema의 최상위 type은 object여야 합니다.")

    properties = schema.get("properties")
    if not isinstance(properties, dict) or not properties:
        raise ValueError(
            "response schema에는 하나 이상의 properties가 있어야 합니다."
        )

    expected_columns = set(ALL_RESPONSE_OUTPUT_COLUMNS)
    actual_columns = set(properties)
    if actual_columns != expected_columns:
        raise ValueError(
            "response schema properties가 v9 출력 컬럼과 일치하지 않습니다. "
            f"missing={sorted(expected_columns - actual_columns)}, "
            f"extra={sorted(actual_columns - expected_columns)}"
        )

    expected_types = {
        "Campaign Name": "string",
        "Product": "string",
        "CXP Product Feature": "string",
        "Description": "string",
        "Publisher Type": "string",
        "Publisher Country": "string",
        "Publisher Classification Reason": "string",
        "Publisher Classification Confidence": "integer",
        "Requires Manual Review": "boolean",
    }

    for column_name, expected_type in expected_types.items():
        column_schema = properties.get(column_name)
        if not isinstance(column_schema, dict):
            raise ValueError(
                "response schema property 정의가 object가 아닙니다: "
                f"{column_name}"
            )
        if column_schema.get("type") != expected_type:
            raise ValueError(
                "response schema property type이 올바르지 않습니다: "
                f"{column_name}, expected={expected_type}, "
                f"actual={column_schema.get('type')}"
            )

    publisher_type_schema = properties["Publisher Type"]
    publisher_enum = set(publisher_type_schema.get("enum") or [])
    if publisher_enum != ALLOWED_PUBLISHER_TYPES:
        raise ValueError(
            "Publisher Type enum은 OWNED, INFLUENCER, UNKNOWN이어야 합니다. "
            f"actual={sorted(publisher_enum)}"
        )

    description_schema = properties["Description"]
    description_max_length = description_schema.get("maxLength")
    if description_max_length != MAX_DESCRIPTION_LENGTH:
        raise ValueError(
            "Description maxLength가 Python 설정과 일치하지 않습니다: "
            f"schema={description_max_length}, "
            f"python={MAX_DESCRIPTION_LENGTH}"
        )

    required = set(schema.get("required") or [])
    if required != expected_columns:
        raise ValueError(
            "response schema required가 v9 출력 컬럼과 일치하지 않습니다. "
            f"missing={sorted(expected_columns - required)}, "
            f"extra={sorted(required - expected_columns)}"
        )

    if schema.get("additionalProperties") is not False:
        raise ValueError(
            "response schema additionalProperties는 false여야 합니다."
        )


def load_response_schema(
    schema_path: Path | None,
    prompts_dir: Path = PROMPTS_DIR,
) -> tuple[dict[str, Any], Path | None]:
    if schema_path is None:
        default_path = prompts_dir / DEFAULT_RESPONSE_SCHEMA_FILENAME
        schema_path = default_path if default_path.is_file() else None

    if schema_path is None:
        schema = json.loads(json.dumps(DEFAULT_RESPONSE_JSON_SCHEMA))
        validate_response_schema(schema)
        return schema, None

    schema_path = Path(schema_path)
    if not schema_path.is_file():
        raise FileNotFoundError(
            f"response schema 파일을 찾을 수 없습니다: {schema_path}"
        )

    try:
        schema = json.loads(schema_path.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"response schema JSON 문법이 올바르지 않습니다: {schema_path}"
        ) from exc

    if not isinstance(schema, dict):
        raise ValueError("response schema 최상위 값은 JSON object여야 합니다.")

    validate_response_schema(schema)
    return schema, schema_path


def load_prompt_bundle(
    prompts_dir: Path = PROMPTS_DIR,
    user_prompt_path: Path | None = None,
    system_prompt_path: Path | None = None,
    response_schema_path: Path | None = None,
) -> PromptBundle:
    prompts_dir = Path(prompts_dir)
    if not prompts_dir.exists():
        raise FileNotFoundError(
            f"prompts 폴더를 찾을 수 없습니다: {prompts_dir}"
        )

    resolved_user_prompt_path = resolve_user_prompt_path(
        prompts_dir=prompts_dir,
        explicit_prompt_path=user_prompt_path,
    )
    user_prompt = read_text_file(resolved_user_prompt_path)

    if system_prompt_path is None:
        candidate = prompts_dir / DEFAULT_SYSTEM_PROMPT_FILENAME
        system_prompt_path = candidate if candidate.is_file() else None
    elif system_prompt_path is not None:
        system_prompt_path = Path(system_prompt_path)

    system_prompt = (
        read_text_file(system_prompt_path)
        if system_prompt_path is not None
        else None
    )

    response_schema, resolved_schema_path = load_response_schema(
        response_schema_path,
        prompts_dir=prompts_dir,
    )

    return PromptBundle(
        user_prompt=user_prompt,
        system_prompt=system_prompt,
        user_prompt_path=resolved_user_prompt_path,
        system_prompt_path=system_prompt_path,
        response_schema=response_schema,
        response_schema_path=resolved_schema_path,
    )


def response_output_columns(
    response_schema: dict[str, Any],
) -> list[str]:
    properties = response_schema.get("properties", {})
    return list(properties.keys())


# =========================================================
# 6. Gemini Request Construction
# =========================================================


def build_post_context_text(
    input_set: LLMInputSet,
    user_prompt: str,
    country_to_subsidiary: dict[str, str],
) -> str:
    metadata_lines = [
        f"campaign_id: {input_set.campaign_id}",
        f"source_sheet_name: {input_set.source_sheet_name}",
        f"raw_row_number: {input_set.raw_row_number}",
        f"platform: {input_set.platform or ''}",
        f"profile_name: {input_set.profile_name or ''}",
        f"profile_url: {input_set.profile_url or ''}",
        f"sender_profile_available: {input_set.sender_profile_available}",
        f"sender_screen_name: {input_set.sender_screen_name or ''}",
        (
            "sender_follower_count: "
            f"{input_set.sender_follower_count or ''}"
        ),
        (
            "sender_follower_count_definition: 현재 게시 플랫폼에서 "
            "게시물을 업로드한 게시 계정의 팔로워 수이며, 콘텐츠에 "
            "등장하는 다른 인물의 팔로워 수로 자동 사용하지 않는다."
        ),
        f"sender_location: {input_set.sender_location or ''}",
        (
            "sender_detailed_location: "
            f"{input_set.sender_detailed_location or ''}"
        ),
        f"sender_bio: {input_set.sender_bio or ''}",
        f"sender_website: {input_set.sender_website or ''}",
        f"sender_verified: {input_set.sender_verified}",
        f"sender_verified_type: {input_set.sender_verified_type or ''}",
        f"sender_profile_tags: {input_set.sender_profile_tags or ''}",
        f"original_post_url: {input_set.original_post_url or ''}",
        f"conversation_stream: {input_set.conversation_stream or ''}",
        f"post_media_type: {input_set.post_media_type or ''}",
        f"media_types: {', '.join(input_set.media_types)}",
        f"media_count: {len(input_set.media_paths)}",
    ]

    publisher_instruction = (
        "[PUBLISHER CLASSIFICATION RULES]\n"
        "1. 먼저 profile_name, profile_url, sender_screen_name을 검사한다. "
        "세 필드 중 하나라도 대소문자 구분 없이 samsung 또는 삼성을 포함하면 "
        "Publisher Type을 반드시 OWNED로 반환한다. 공백·마침표·밑줄·하이픈 "
        "차이는 무시한다.\n"
        "2. SamsungGulf, Samsung Gulf, samsung_gulf, samsung.gulf, "
        "samsung-gulf가 확인되면 Publisher Type=OWNED, "
        "Publisher Country=Dubai, Confidence=100, "
        "Requires Manual Review=false를 반환한다.\n"
        "3. SamsungKorea, Samsung Korea, samsung_korea, samsung.korea, "
        "samsung-korea 또는 'Samsung Korea 삼성전자'가 확인되면 "
        "Publisher Type=OWNED, Publisher Country=Korea, Confidence=100, "
        "Requires Manual Review=false를 반환한다. South Korea는 반환하지 않는다.\n"
        "4. 위 고정 계정이 아닌 일반 Samsung 계정은 OWNED로 확정하되, "
        "국가는 Location, Website, Bio, 계정명 등 강한 근거로 판단한다. "
        "국가를 확정할 수 없으면 Publisher Country는 빈 문자열이고 "
        "Requires Manual Review=true이다.\n"
        "5. 게시물 본문, 캡션, 해시태그 또는 미디어 속 Samsung 언급만으로는 "
        "OWNED로 판정하지 않는다. 게시 계정 관련 필드만 강제 규칙에 사용한다.\n"
        "6. Samsung 공식 계정의 미디어에 배우·가수·크리에이터·인플루언서가 "
        "등장하고 신원·직업·출연 근거의 내부 판단 신뢰도가 99 이상이면 "
        "Description에 해당 인물 정보를 포함할 수 있다. 그러나 게시 계정이 "
        "Samsung 공식 계정이면 Publisher Type은 계속 OWNED이다. 후속 코드는 "
        "URL에 [당사 게시글]을 붙이고 Influencer=No로 기록한다.\n"
        "7. sender_follower_count는 현재 게시 플랫폼에서 게시물을 업로드한 "
        "게시 계정의 팔로워 수이다. OWNED 계정에서는 콘텐츠 속 등장 인물의 "
        "팔로워 수로 사용하지 않는다. INFLUENCER 계정이며 게시 계정과 언급 "
        "인물이 동일할 때만 현재 게시 플랫폼의 팔로워 수로 사용할 수 있다.\n"
        "8. 위 강제 규칙에 해당하지 않으면 Publisher Type은 OWNED, "
        "INFLUENCER, UNKNOWN 중 하나를 기존 증거 기준으로 반환한다.\n"
        "9. Sprinklr Profile Tags는 참고 정보일 뿐이며 태그 하나만으로 "
        "분류하지 않는다.\n"
        "10. Publisher Country는 코드가 제공하는 허용 Country 목록 중 정확히 "
        "하나를 그대로 반환하거나, 근거가 부족하면 빈 문자열을 반환한다.\n"
        "11. Subsidiary Code, SGE, KOREA, SGE (Dubai), KOREA (Korea)는 "
        "응답에 생성하지 않는다. 실제 Subsidiary는 코드가 확정 JSON으로 매핑한다.\n"
        "12. 유형 또는 국가가 불확실하거나 증거가 충돌하면 "
        "Requires Manual Review=true로 설정한다.\n"
        "13. Publisher Classification Confidence는 0~100 정수이다.\n"
        + build_allowed_country_instruction(country_to_subsidiary)
        + "\n"
    )

    input_mode = (input_set.llm_input_mode or "").lower()

    if input_mode == "post_text_only":
        analysis_source_instruction = (
            "This post has been confirmed to contain no native image or video. "
            "Analyze the post using the provided conversation_stream, "
            "original_post_url, and account/profile metadata only. "
        )
    elif is_unknown_media_input(input_set):
        analysis_source_instruction = (
            "The media type for this post is unresolved or partially unknown. "
            "Perform the analysis using the provided conversation_stream, "
            "original_post_url, account/profile metadata, and any valid media "
            "that is attached. Do not infer facts from media that is unavailable. "
        )
    else:
        analysis_source_instruction = (
            "Analyze all attached media as one unified post-level input. "
        )

    return (
        "[ANALYSIS INSTRUCTION]\n"
        f"{user_prompt}\n\n"
        + publisher_instruction
        + "\n[POST METADATA]\n"
        + "\n".join(metadata_lines)
        + "\n\n"
        + analysis_source_instruction
        + "Return exactly one structured response for the post. "
        "Do not invent profile facts, country, or subsidiary information."
    )

def detect_mime_type(path: Path) -> str:
    suffix = path.suffix.lower()

    if suffix in SUPPORTED_LOCAL_MEDIA_MIME_TYPES:
        return SUPPORTED_LOCAL_MEDIA_MIME_TYPES[suffix]

    guessed_mime_type, _ = mimetypes.guess_type(path.name)
    if guessed_mime_type and (
        guessed_mime_type.startswith("image/")
        or guessed_mime_type.startswith("video/")
    ):
        return guessed_mime_type

    raise ValueError(f"지원하지 않는 로컬 미디어 확장자입니다: {path}")


def validate_inline_media_size(
    media_paths: tuple[Path, ...],
) -> int:
    total_bytes = sum(path.stat().st_size for path in media_paths)

    if total_bytes > MAX_TOTAL_INLINE_BYTES:
        raise ValueError(
            "한 게시물의 로컬 미디어 총 크기가 프로그램 안전 제한을 "
            f"초과했습니다: {total_bytes / (1024 * 1024):.1f} MB > "
            f"{MAX_TOTAL_INLINE_MB} MB. 큰 파일은 GCS URI 방식으로 "
            "전환하거나 MAX_TOTAL_INLINE_MB를 조정하세요."
        )

    return total_bytes


def build_gemini_contents(
    input_set: LLMInputSet,
    user_prompt: str,
    types_module: Any,
    country_to_subsidiary: dict[str, str],
) -> list[Any]:
    contents: list[Any] = [
        build_post_context_text(
            input_set=input_set,
            user_prompt=user_prompt,
            country_to_subsidiary=country_to_subsidiary,
        )
    ]

    input_mode = (input_set.llm_input_mode or "").lower()
    platform = (input_set.platform or "").lower()

    if input_mode == "youtube_url" or platform == "youtube":
        youtube_url, _ = split_post_and_comment_urls(
            input_set.llm_input_value
        )
        youtube_url = (
            youtube_url
            or input_set.original_post_url
        )

        if not youtube_url:
            raise ValueError("YouTube URL이 없습니다.")

        contents.append(
            types_module.Part.from_uri(
                file_uri=youtube_url,
                mime_type="video/mp4",
            )
        )
        return contents

    if input_mode == "post_text_only":
        return contents

    if is_unknown_media_input(input_set):
        # UNKNOWN은 미디어 추출 실패 여부와 무관하게 1차 LLM 분석을 수행한다.
        # 실제로 존재하고 지원 가능한 로컬 미디어만 추가하고,
        # 사용할 수 있는 미디어가 없으면 텍스트/URL/계정 메타데이터만 전달한다.
        existing_media_paths = tuple(
            media_path
            for media_path in input_set.media_paths
            if media_path.is_file()
        )

        if not existing_media_paths:
            return contents

        try:
            validate_inline_media_size(
                existing_media_paths
            )
        except ValueError:
            # UNKNOWN 게시물의 1차 분석 자체를 막지 않기 위해
            # 첨부 제한을 초과하면 텍스트 기반 분석으로 fallback한다.
            return contents

        for media_path in existing_media_paths:
            try:
                mime_type = detect_mime_type(
                    media_path
                )
                media_bytes = media_path.read_bytes()
            except (OSError, ValueError):
                # 개별 미디어 파일 문제 때문에 게시물 전체 1차 분석을
                # 중단하지 않는다.
                continue

            contents.append(
                types_module.Part.from_bytes(
                    data=media_bytes,
                    mime_type=mime_type,
                )
            )

        return contents

    if not input_set.media_paths:
        raise ValueError("LLM에 전달할 local_media_paths가 없습니다.")

    validate_inline_media_size(input_set.media_paths)

    for media_path in input_set.media_paths:
        mime_type = detect_mime_type(media_path)
        media_bytes = media_path.read_bytes()

        contents.append(
            types_module.Part.from_bytes(
                data=media_bytes,
                mime_type=mime_type,
            )
        )

    return contents


# =========================================================
# 7. Gemini Client / API Call
# =========================================================


def create_genai_client() -> tuple[Any, Any]:
    try:
        from google import genai
        from google.genai import types
    except ImportError as exc:
        raise ImportError(
            "google-genai 패키지가 설치되어 있지 않습니다. "
            "다음 명령으로 설치하세요: uv pip install google-genai"
        ) from exc

    client_kwargs = {
        "project": GOOGLE_CLOUD_PROJECT,
        "location": GOOGLE_CLOUD_LOCATION,
        "http_options": types.HttpOptions(api_version="v1"),
    }

    try:
        client = genai.Client(
            enterprise=True,
            **client_kwargs,
        )
    except (TypeError, ValueError):
        client = genai.Client(
            vertexai=True,
            **client_kwargs,
        )

    return client, types


def is_retryable_api_error(exc: Exception) -> bool:
    text = f"{type(exc).__name__}: {exc}".lower()
    retry_markers = {
        "429",
        "500",
        "502",
        "503",
        "504",
        "resource exhausted",
        "deadline exceeded",
        "temporarily unavailable",
        "service unavailable",
        "timeout",
        # Structured output이 일시적으로 잘리거나 형식이 깨진 경우 재호출한다.
        "json으로 파싱하지 못했습니다",
        "structured response",
        "gemini 응답 key가 response schema와 일치하지 않습니다",
        "unterminated string",
        "expecting property name",
        "expecting value",
    }
    return any(marker in text for marker in retry_markers)


def validate_parsed_output(
    parsed_output: dict[str, Any],
    response_schema: dict[str, Any],
) -> dict[str, Any]:
    output_columns = response_output_columns(response_schema)

    expected = set(output_columns)
    actual = set(parsed_output)

    if actual != expected:
        raise ValueError(
            "Gemini 응답 key가 response schema와 일치하지 않습니다. "
            f"expected={sorted(expected)}, actual={sorted(actual)}"
        )

    validated: dict[str, Any] = {}

    string_columns = {
        *TARGET_OUTPUT_COLUMNS,
        "Publisher Country",
        "Publisher Classification Reason",
    }

    for column_name in output_columns:
        value = parsed_output.get(column_name)

        if column_name in string_columns:
            if value is None:
                value = ""
            if not isinstance(value, str):
                raise ValueError(
                    "Gemini 응답 값이 문자열이 아닙니다: "
                    f"{column_name}={value!r}"
                )

            if column_name == "Product":
                normalized_value = normalize_product_value(value)
            elif column_name == "CXP Product Feature":
                normalized_value = normalize_cxp_feature_value(value)
            elif column_name == "Description":
                normalized_value = normalize_description_value(value)
            else:
                normalized_value = " ".join(value.split())

            validated[column_name] = normalized_value
            continue

        if column_name == "Publisher Type":
            if not isinstance(value, str):
                raise ValueError(
                    "Publisher Type은 문자열이어야 합니다: "
                    f"{value!r}"
                )
            publisher_type = value.strip().upper()
            if publisher_type not in ALLOWED_PUBLISHER_TYPES:
                raise ValueError(
                    "Publisher Type 값이 허용 목록에 없습니다: "
                    f"{publisher_type!r}"
                )
            validated[column_name] = publisher_type
            continue

        if column_name == "Publisher Classification Confidence":
            try:
                confidence = int(value)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    "Publisher Classification Confidence를 정수로 "
                    f"변환할 수 없습니다: {value!r}"
                ) from exc
            if not 0 <= confidence <= 100:
                raise ValueError(
                    "Publisher Classification Confidence는 0~100이어야 합니다: "
                    f"{confidence}"
                )
            validated[column_name] = confidence
            continue

        if column_name == "Requires Manual Review":
            if isinstance(value, bool):
                manual_review = value
            elif isinstance(value, str):
                normalized = value.strip().lower()
                if normalized in {"true", "1", "yes", "y"}:
                    manual_review = True
                elif normalized in {"false", "0", "no", "n"}:
                    manual_review = False
                else:
                    raise ValueError(
                        "Requires Manual Review 값을 bool로 해석할 수 없습니다: "
                        f"{value!r}"
                    )
            else:
                raise ValueError(
                    "Requires Manual Review는 boolean이어야 합니다: "
                    f"{value!r}"
                )
            validated[column_name] = manual_review
            continue

        validated[column_name] = value

    # 안전 보정: UNKNOWN 또는 국가 미확정은 반드시 수동 검토로 보낸다.
    if (
        validated.get("Publisher Type") == "UNKNOWN"
        or not validated.get("Publisher Country")
    ):
        validated["Requires Manual Review"] = True

    return validated


def parse_structured_response(
    response: Any,
    response_schema: dict[str, Any],
) -> dict[str, Any]:
    parsed = getattr(response, "parsed", None)
    parsed_output: dict[str, Any] | None = None

    if isinstance(parsed, dict):
        parsed_output = parsed
    elif parsed is not None:
        if hasattr(parsed, "model_dump"):
            dumped = parsed.model_dump()
            if isinstance(dumped, dict):
                parsed_output = dumped

        if parsed_output is None and hasattr(parsed, "dict"):
            dumped = parsed.dict()
            if isinstance(dumped, dict):
                parsed_output = dumped

    if parsed_output is None:
        response_text = optional_text(getattr(response, "text", None))
        if response_text is None:
            raise ValueError("Gemini 응답 text가 비어 있습니다.")

        try:
            loaded = json.loads(response_text)
        except json.JSONDecodeError as exc:
            head = response_text[:700]
            tail = response_text[-300:] if len(response_text) > 700 else ""
            raise ValueError(
                "Gemini 응답을 JSON으로 파싱하지 못했습니다. "
                f"length={len(response_text)}, "
                f"head={head!r}, tail={tail!r}"
            ) from exc

        if not isinstance(loaded, dict):
            raise ValueError(
                "Gemini structured response가 JSON object가 아닙니다."
            )
        parsed_output = loaded

    return validate_parsed_output(parsed_output, response_schema)


def extract_usage_metadata(
    response: Any,
) -> tuple[int | None, int | None, int | None]:
    usage = getattr(response, "usage_metadata", None)
    if usage is None:
        return None, None, None

    return (
        getattr(usage, "prompt_token_count", None),
        getattr(usage, "candidates_token_count", None),
        getattr(usage, "total_token_count", None),
    )


def call_gemini_for_input(
    client: Any,
    types_module: Any,
    input_set: LLMInputSet,
    prompt_bundle: PromptBundle,
    country_to_subsidiary: dict[str, str],
) -> LLMCallResult:
    processed_at_utc = datetime.now(timezone.utc).isoformat()

    try:
        contents = build_gemini_contents(
            input_set=input_set,
            user_prompt=prompt_bundle.user_prompt,
            types_module=types_module,
            country_to_subsidiary=country_to_subsidiary,
        )
    except Exception as exc:
        return LLMCallResult(
            dataframe_index=input_set.dataframe_index,
            api_status="input_build_failed",
            api_error_message=f"{type(exc).__name__}: {exc}",
            processed_at_utc=processed_at_utc,
            model_name=GEMINI_MODEL,
        )

    config = types_module.GenerateContentConfig(
        system_instruction=prompt_bundle.system_prompt,
        temperature=TEMPERATURE,
        max_output_tokens=MAX_OUTPUT_TOKENS,
        response_mime_type="application/json",
        response_json_schema=prompt_bundle.response_schema,
    )

    last_exception: Exception | None = None
    last_raw_response: str | None = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=contents,
                config=config,
            )

            last_raw_response = optional_text(
                getattr(response, "text", None)
            )
            parsed_output = parse_structured_response(
                response,
                response_schema=prompt_bundle.response_schema,
            )
            original_publisher_type = optional_text(
                parsed_output.get("Publisher Type")
            )
            original_publisher_country = optional_text(
                parsed_output.get("Publisher Country")
            )
            parsed_output, publisher_override = (
                apply_publisher_account_override(
                    parsed_output=parsed_output,
                    input_set=input_set,
                )
            )
            prompt_tokens, candidate_tokens, total_tokens = (
                extract_usage_metadata(response)
            )

            return LLMCallResult(
                dataframe_index=input_set.dataframe_index,
                api_status="success",
                raw_llm_response=last_raw_response,
                parsed_output=parsed_output,
                processed_at_utc=datetime.now(timezone.utc).isoformat(),
                model_name=GEMINI_MODEL,
                prompt_token_count=prompt_tokens,
                candidates_token_count=candidate_tokens,
                total_token_count=total_tokens,
                publisher_override_applied=(
                    publisher_override.applied
                ),
                publisher_override_rule=publisher_override.rule,
                publisher_override_source_field=(
                    publisher_override.source_field
                ),
                publisher_override_source_value=(
                    publisher_override.source_value
                ),
                publisher_override_original_type=(
                    original_publisher_type
                ),
                publisher_override_original_country=(
                    original_publisher_country
                ),
            )

        except Exception as exc:
            last_exception = exc

            if attempt >= MAX_RETRIES or not is_retryable_api_error(exc):
                break

            sleep_seconds = min(2 ** (attempt - 1), 8)
            time.sleep(sleep_seconds)

    assert last_exception is not None

    return LLMCallResult(
        dataframe_index=input_set.dataframe_index,
        api_status="api_failed",
        api_error_message=(
            f"{type(last_exception).__name__}: {last_exception}"
        ),
        raw_llm_response=last_raw_response,
        processed_at_utc=datetime.now(timezone.utc).isoformat(),
        model_name=GEMINI_MODEL,
    )


# =========================================================
# 8. Eligibility / LLM Pipeline Execution
# =========================================================


def build_skip_result(input_set: LLMInputSet) -> LLMCallResult | None:
    mode = (input_set.llm_input_mode or "").lower()
    platform = (input_set.platform or "").lower()

    # UNKNOWN media는 media_extractor의 준비/수동조치 상태와 관계없이
    # 게시물 텍스트, permalink, 계정 메타데이터 및 확보된 미디어를 이용해
    # 최소 1회의 LLM 분석을 수행한다.
    if is_unknown_media_input(input_set):
        return None

    if input_set.user_action_required:
        return LLMCallResult(
            dataframe_index=input_set.dataframe_index,
            api_status="skipped_user_action_required",
            api_error_message=input_set.user_action,
        )

    if not input_set.llm_ready:
        return LLMCallResult(
            dataframe_index=input_set.dataframe_index,
            api_status="skipped_not_ready",
            api_error_message=(
                "media_extractor 결과에서 llm_ready=False입니다."
            ),
        )

    if input_set.missing_media_paths:
        return LLMCallResult(
            dataframe_index=input_set.dataframe_index,
            api_status="skipped_missing_media_files",
            api_error_message=(
                "다음 미디어 파일을 찾을 수 없습니다: "
                + " | ".join(
                    str(path) for path in input_set.missing_media_paths
                )
            ),
        )

    if mode == "youtube_url" or platform == "youtube":
        youtube_url, _ = split_post_and_comment_urls(
            input_set.llm_input_value
        )
        youtube_url = (
            youtube_url
            or input_set.original_post_url
        )

        if not youtube_url:
            return LLMCallResult(
                dataframe_index=input_set.dataframe_index,
                api_status="skipped_invalid_input",
                api_error_message="YouTube URL이 없습니다.",
            )
        return None

    if mode == "post_text_only":
        if not input_set.conversation_stream:
            return LLMCallResult(
                dataframe_index=input_set.dataframe_index,
                api_status="skipped_invalid_input",
                api_error_message=(
                    "post_text_only 입력인데 conversation_stream이 비어 있습니다."
                ),
            )
        return None

    if not input_set.media_paths:
        return LLMCallResult(
            dataframe_index=input_set.dataframe_index,
            api_status="skipped_invalid_input",
            api_error_message="local_media_paths가 비어 있습니다.",
        )

    return None


def initialize_result_dataframe(
    input_dataframe: pd.DataFrame,
    prompt_bundle: PromptBundle,
) -> pd.DataFrame:
    output_columns = response_output_columns(prompt_bundle.response_schema)

    management_columns = [
        "api_status",
        "api_error_message",
        "raw_llm_response",
        "processed_at_utc",
        "model_name",
        "prompt_token_count",
        "candidates_token_count",
        "total_token_count",
        "target_mapping_status",
        "target_mapping_message",
        "target_sheet_name",
        "target_row_number",
        "mapped_subsidiary",
        "country_mapping_status",
        "country_mapping_message",
        "subsidiary_display_value",
        "final_influencer_value",
        "final_formatted_url",
        "publisher_writeback_eligible",
        "publisher_writeback_skip_reason",
        "code_requires_manual_review",
        "publisher_override_applied",
        "publisher_override_rule",
        "publisher_override_source_field",
        "publisher_override_source_value",
        "publisher_override_original_type",
        "publisher_override_original_country",
    ]

    collisions = sorted(
        set(input_dataframe.columns)
        & (set(management_columns) | set(output_columns))
    )
    if collisions:
        raise ValueError(
            "입력 llm_input 시트에 LLM 결과/관리 컬럼이 이미 존재합니다. "
            f"중복 컬럼={collisions}. 원본 campaign_media_result 파일을 사용하세요."
        )

    result_dataframe = input_dataframe.copy()
    for column in management_columns + output_columns:
        result_dataframe[column] = pd.NA

    return result_dataframe


def write_call_result_to_dataframe(
    result_dataframe: pd.DataFrame,
    call_result: LLMCallResult,
    output_columns: Iterable[str],
) -> None:
    row_index = call_result.dataframe_index

    management_values = {
        "api_status": call_result.api_status,
        "api_error_message": call_result.api_error_message,
        "raw_llm_response": call_result.raw_llm_response,
        "processed_at_utc": call_result.processed_at_utc,
        "model_name": call_result.model_name,
        "prompt_token_count": call_result.prompt_token_count,
        "candidates_token_count": call_result.candidates_token_count,
        "total_token_count": call_result.total_token_count,
        "publisher_override_applied": (
            call_result.publisher_override_applied
        ),
        "publisher_override_rule": call_result.publisher_override_rule,
        "publisher_override_source_field": (
            call_result.publisher_override_source_field
        ),
        "publisher_override_source_value": (
            call_result.publisher_override_source_value
        ),
        "publisher_override_original_type": (
            call_result.publisher_override_original_type
        ),
        "publisher_override_original_country": (
            call_result.publisher_override_original_country
        ),
    }

    for column_name, value in management_values.items():
        result_dataframe.at[row_index, column_name] = value

    if call_result.parsed_output:
        for output_column in output_columns:
            result_dataframe.at[row_index, output_column] = (
                json_safe_excel_value(
                    call_result.parsed_output.get(output_column)
                )
            )


def run_llm_pipeline(
    result_excel_path: Path,
    prompt_bundle: PromptBundle,
    country_to_subsidiary: dict[str, str],
    max_rows: int | None = None,
    dry_run: bool = False,
) -> pd.DataFrame:
    input_dataframe = load_llm_input_dataframe(result_excel_path)
    input_sets = build_llm_input_sets_from_dataframe(input_dataframe)
    result_dataframe = initialize_result_dataframe(
        input_dataframe,
        prompt_bundle,
    )

    output_columns = response_output_columns(prompt_bundle.response_schema)

    if not input_sets:
        client = None
        types_module = None
        print(
            "llm_input 시트에 처리할 게시물이 없습니다. "
            "API 호출 없이 빈 LLM 로그를 생성합니다."
        )
    elif dry_run:
        client = None
        types_module = None
    else:
        client, types_module = create_genai_client()

    attempted_count = 0

    try:
        total_rows = len(input_sets)

        for current_position, input_set in enumerate(input_sets, start=1):
            skip_result = build_skip_result(input_set)

            if skip_result is not None:
                call_result = skip_result

                if SHOW_SKIPPED_ROWS:
                    print(
                        f"[{current_position}/{total_rows}] SKIP "
                        f"{input_set.campaign_id}: "
                        f"{call_result.api_status}"
                    )

            elif max_rows is not None and attempted_count >= max_rows:
                call_result = LLMCallResult(
                    dataframe_index=input_set.dataframe_index,
                    api_status="skipped_max_rows_limit",
                    api_error_message=(
                        f"max_rows={max_rows} 테스트 제한으로 호출하지 않음"
                    ),
                )

            elif dry_run:
                call_result = LLMCallResult(
                    dataframe_index=input_set.dataframe_index,
                    api_status="dry_run_ready",
                )

                if SHOW_PROGRESS:
                    print(
                        f"[{current_position}/{total_rows}] DRY-RUN READY "
                        f"{input_set.campaign_id} "
                        f"media={len(input_set.media_paths)}"
                    )

            else:
                attempted_count += 1

                if SHOW_PROGRESS:
                    print(
                        f"[{current_position}/{total_rows}] CALL "
                        f"{input_set.campaign_id} "
                        f"mode={input_set.llm_input_mode} "
                        f"media={len(input_set.media_paths)}"
                    )

                assert client is not None
                assert types_module is not None

                try:
                    call_result = call_gemini_for_input(
                        client=client,
                        types_module=types_module,
                        input_set=input_set,
                        prompt_bundle=prompt_bundle,
                        country_to_subsidiary=country_to_subsidiary,
                    )
                except Exception as exc:
                    call_result = LLMCallResult(
                        dataframe_index=input_set.dataframe_index,
                        api_status="unexpected_pipeline_error",
                        api_error_message=f"{type(exc).__name__}: {exc}",
                        processed_at_utc=(
                            datetime.now(timezone.utc).isoformat()
                        ),
                        model_name=GEMINI_MODEL,
                    )

                if (
                    SHOW_API_ERRORS
                    and call_result.api_status
                    in {
                        "api_failed",
                        "input_build_failed",
                        "unexpected_pipeline_error",
                    }
                ):
                    print(
                        "  처리 실패: "
                        f"{call_result.api_error_message}"
                    )

            write_call_result_to_dataframe(
                result_dataframe=result_dataframe,
                call_result=call_result,
                output_columns=output_columns,
            )

    finally:
        if client is not None:
            close_method = getattr(client, "close", None)
            if callable(close_method):
                close_method()

    return result_dataframe


# =========================================================
# 9. LLM Log Excel Output
# =========================================================


def save_llm_result_excel(
    output_excel_path: Path,
    result_dataframe: pd.DataFrame,
) -> None:
    output_excel_path = Path(output_excel_path)
    output_excel_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        result_dataframe.to_excel(
            writer,
            sheet_name=LLM_LOG_SHEET_NAME,
            index=False,
        )

        worksheet = writer.book[LLM_LOG_SHEET_NAME]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_map = {cell.value: cell.column for cell in worksheet[1]}

        wrap_column_names = {
            "conversation_stream",
            "permalink",
            "profile_url",
            "sender_location",
            "sender_detailed_location",
            "sender_bio",
            "sender_website",
            "sender_profile_tags",
            "media_types",
            "media_source_urls",
            "local_media_paths",
            "error_message",
            "user_action",
            "api_error_message",
            "raw_llm_response",
            "target_mapping_message",
            "country_mapping_message",
            "publisher_writeback_skip_reason",
            "final_formatted_url",
            "Product",
            "CXP Product Feature",
            "Description",
            "Publisher Classification Reason",
            "publisher_override_source_value",
        }

        for column_name in wrap_column_names:
            column_number = header_map.get(column_name)
            if column_number is None:
                continue

            for row_number in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )
                new_alignment = copy(cell.alignment)
                new_alignment.wrap_text = True
                new_alignment.vertical = "top"
                cell.alignment = new_alignment

        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")

            if header in wrap_column_names:
                width = 45
            elif header in {
                "campaign_id",
                "source_sheet",
                "llm_input_mode",
                "api_status",
                "target_mapping_status",
            }:
                width = 28
            else:
                width = 18

            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = width


# =========================================================
# 10. Formatted Excel URL Mapping / Write-back
# =========================================================


def get_header_column_map(
    worksheet: Worksheet,
    header_row_number: int,
) -> dict[str, int]:
    header_column_map: dict[str, int] = {}

    for cell in worksheet[header_row_number]:
        header_name = optional_text(cell.value)
        if header_name is None:
            continue

        if header_name in header_column_map:
            raise ValueError(
                f"대상 시트 헤더에 중복 컬럼이 있습니다: {header_name}"
            )

        header_column_map[header_name] = cell.column

    return header_column_map


def select_first_existing_column(
    header_column_map: dict[str, int],
    candidates: Iterable[str],
) -> str | None:
    for candidate in candidates:
        if candidate in header_column_map:
            return candidate

    lowered_lookup = {
        column_name.strip().lower(): column_name
        for column_name in header_column_map
    }
    for candidate in candidates:
        matched = lowered_lookup.get(candidate.strip().lower())
        if matched:
            return matched

    return None


def resolve_target_sheet_layout(
    worksheet: Worksheet,
    explicit_header_row: int | None = None,
    explicit_url_column: str | None = None,
    explicit_campaign_id_column: str | None = None,
) -> TargetSheetLayout:
    candidate_header_rows = (
        [explicit_header_row]
        if explicit_header_row is not None
        else list(
            range(
                1,
                min(worksheet.max_row, TARGET_HEADER_SCAN_MAX_ROWS) + 1,
            )
        )
    )

    diagnostics: list[str] = []

    for header_row_number in candidate_header_rows:
        header_column_map = get_header_column_map(
            worksheet,
            header_row_number,
        )

        missing_output_columns = [
            column_name
            for column_name in TARGET_WRITEBACK_COLUMNS
            if column_name not in header_column_map
        ]
        if missing_output_columns:
            diagnostics.append(
                f"row={header_row_number}: missing={missing_output_columns}"
            )
            continue

        if explicit_url_column:
            url_column_name = (
                explicit_url_column
                if explicit_url_column in header_column_map
                else None
            )
        else:
            url_column_name = select_first_existing_column(
                header_column_map,
                TARGET_URL_COLUMN_CANDIDATES,
            )

        if url_column_name is None:
            diagnostics.append(
                f"row={header_row_number}: URL column not found"
            )
            continue

        if explicit_campaign_id_column:
            campaign_id_column_name = (
                explicit_campaign_id_column
                if explicit_campaign_id_column in header_column_map
                else None
            )
            if campaign_id_column_name is None:
                raise ValueError(
                    "지정한 campaign ID 컬럼을 찾을 수 없습니다: "
                    f"{explicit_campaign_id_column}"
                )
        else:
            campaign_id_column_name = select_first_existing_column(
                header_column_map,
                TARGET_CAMPAIGN_ID_COLUMN_CANDIDATES,
            )

        return TargetSheetLayout(
            header_row_number=header_row_number,
            header_column_map=header_column_map,
            url_column_name=url_column_name,
            url_column_number=header_column_map[url_column_name],
            campaign_id_column_name=campaign_id_column_name,
            campaign_id_column_number=(
                header_column_map[campaign_id_column_name]
                if campaign_id_column_name
                else None
            ),
        )

    raise ValueError(
        "대상 formatted 시트에서 URL 및 결과 컬럼이 있는 헤더 행을 "
        "찾지 못했습니다. --target-header-row 또는 --target-url-column을 "
        f"지정하세요. 검사 결과={diagnostics}"
    )


def get_url_cell_values(cell: Any) -> tuple[str, ...]:
    values: list[str] = []

    hyperlink = getattr(cell, "hyperlink", None)
    hyperlink_target = getattr(hyperlink, "target", None)
    if hyperlink_target:
        values.append(str(hyperlink_target))

    if cell.value is not None:
        values.append(str(cell.value))

    return tuple(dict.fromkeys(values))


def build_target_url_row_lookup(
    worksheet: Worksheet,
    layout: TargetSheetLayout,
) -> dict[str, list[int]]:
    """
    대상 Excel URL 셀의 첫 번째 URL(원 게시물)만 행 매핑 키로 사용한다.

    댓글 URL까지 lookup key로 넣으면 X처럼 댓글 자체가 별도 status URL인
    플랫폼에서 다른 게시물 행과 충돌할 수 있으므로 소비자 반응 URL은
    매핑 키에서 제외한다.
    """
    lookup: dict[str, list[int]] = defaultdict(list)

    for row_number in range(
        layout.header_row_number + 1,
        worksheet.max_row + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=layout.url_column_number,
        )

        normalized_post_urls: list[str] = []

        for cell_value in get_url_cell_values(cell):
            post_url, _ = split_post_and_comment_urls(
                cell_value
            )

            normalized_url = normalize_post_url(
                post_url
            )

            if (
                normalized_url
                and normalized_url
                not in normalized_post_urls
            ):
                normalized_post_urls.append(
                    normalized_url
                )

        for normalized_url in normalized_post_urls:
            lookup[normalized_url].append(
                row_number
            )

    return dict(lookup)


def candidate_target_rows_for_result(
    result_row: pd.Series,
    url_row_lookup: dict[str, list[int]],
) -> tuple[list[int], tuple[str, ...]]:
    """
    LLM 결과 행에서도 첫 번째 URL(원 게시물)만 target row 매핑에 사용한다.
    """
    normalized_urls: list[str] = []

    for source_value in (
        result_row.get("permalink"),
        result_row.get("llm_input_value"),
    ):
        post_url, _ = split_post_and_comment_urls(
            source_value
        )

        normalized_url = normalize_post_url(
            post_url
        )

        if (
            normalized_url
            and normalized_url
            not in normalized_urls
        ):
            normalized_urls.append(
                normalized_url
            )

    unique_urls = tuple(
        normalized_urls
    )
    candidate_rows: set[int] = set()

    for normalized_url in unique_urls:
        candidate_rows.update(
            url_row_lookup.get(
                normalized_url,
                [],
            )
        )

    return sorted(candidate_rows), unique_urls


def disambiguate_target_row(
    worksheet: Worksheet,
    layout: TargetSheetLayout,
    candidate_rows: list[int],
    expected_campaign_id: str,
    raw_row_number: int | None,
) -> int | None:
    if len(candidate_rows) == 1:
        return candidate_rows[0]

    if not candidate_rows:
        return None

    if layout.campaign_id_column_number is not None:
        matching_rows: list[int] = []

        for row_number in candidate_rows:
            raw_campaign_id = worksheet.cell(
                row=row_number,
                column=layout.campaign_id_column_number,
            ).value

            try:
                actual_campaign_id = normalize_campaign_id(raw_campaign_id)
            except ValueError:
                continue

            if actual_campaign_id == expected_campaign_id:
                matching_rows.append(row_number)

        if len(matching_rows) == 1:
            return matching_rows[0]

        if matching_rows:
            candidate_rows = matching_rows

    # 최후 fallback: 기존 raw row가 후보 행 중 정확히 하나라면 사용한다.
    if raw_row_number is not None and raw_row_number in candidate_rows:
        return raw_row_number

    return None



def get_primary_post_url(
    result_row: pd.Series,
) -> str | None:
    """
    permalink/llm_input_value에서 첫 번째 URL인 원 게시물 URL만 반환한다.
    """
    for source_value in (
        result_row.get("permalink"),
        result_row.get("llm_input_value"),
    ):
        post_url, _ = split_post_and_comment_urls(
            source_value
        )

        if post_url:
            return post_url

    return None


def get_consumer_comment_url(
    result_row: pd.Series,
) -> str | None:
    """
    permalink/llm_input_value에 두 번째 URL이 존재하면 소비자 댓글 URL로 반환한다.
    """
    for source_value in (
        result_row.get("permalink"),
        result_row.get("llm_input_value"),
    ):
        _, comment_url = split_post_and_comment_urls(
            source_value
        )

        if comment_url:
            return comment_url

    return None


def format_publisher_url(
    original_url: str,
    publisher_type: str,
    consumer_comment_url: str | None = None,
) -> str:
    """
    최종 URL 셀 문자열을 deterministic하게 생성한다.

    댓글 있음:
        [당사 게시글] / [인플루언서 게시글]
        POST_URL

        [소비자 반응]
        COMMENT_URL

    댓글 없음:
        [당사 게시글] / [인플루언서 게시글]
        POST_URL
    """
    if publisher_type == "INFLUENCER":
        header = "[인플루언서 게시글]"
    elif publisher_type == "OWNED":
        header = "[당사 게시글]"
    else:
        raise ValueError(
            "URL 헤더는 OWNED 또는 INFLUENCER일 때만 생성할 수 있습니다."
        )

    consumer_response = (
        consumer_comment_url
        if consumer_comment_url
        else "N/A"
    )

    return (
        f"{header}\n"
        f"{original_url}\n\n"
        f"[소비자 반응]\n"
        f"{consumer_response}"
    )


def build_publisher_writeback_decision(
    result_row: pd.Series,
    country_to_subsidiary: dict[str, str],
) -> PublisherWritebackDecision:
    """AI는 유형·국가만 판단하고 Subsidiary는 JSON으로 강제 확정한다."""

    publisher_type = (
        optional_text(result_row.get("Publisher Type")) or ""
    ).upper()
    publisher_country = optional_text(
        result_row.get("Publisher Country")
    )
    requires_manual_review = normalize_optional_bool(
        result_row.get("Requires Manual Review")
    )

    confidence_value = result_row.get(
        "Publisher Classification Confidence"
    )
    try:
        confidence = int(confidence_value)
    except (TypeError, ValueError):
        confidence = 0

    mapping_result = resolve_country_subsidiary(
        publisher_country=publisher_country,
        country_to_subsidiary=country_to_subsidiary,
    )

    def blocked(reason: str) -> PublisherWritebackDecision:
        return PublisherWritebackDecision(
            publisher_type=publisher_type,
            publisher_country=mapping_result.publisher_country,
            mapped_subsidiary=mapping_result.mapped_subsidiary,
            country_mapping_status=mapping_result.status,
            country_mapping_message=mapping_result.message,
            subsidiary_display_value=None,
            influencer_value=None,
            formatted_url_value=None,
            formatted_url_target=get_primary_post_url(result_row),
            eligible=False,
            skip_reason=reason,
            code_requires_manual_review=True,
        )

    publisher_override_applied = normalize_bool(
        result_row.get("publisher_override_applied"),
        default=False,
    )

    # Publisher Type은 계정명/프로필 URL/Screen Name 기반 deterministic
    # 규칙으로 Samsung=OWNED, 비Samsung=INFLUENCER 중 하나로 확정된다.
    # Requires Manual Review나 낮은 Confidence가 제품/국가/인물/경품 등
    # 다른 항목에서 발생해도 확정된 Publisher Type 자체는 막지 않는다.
    if requires_manual_review is True and not publisher_override_applied:
        return blocked("Requires Manual Review=true")

    if (
        confidence < MIN_AUTO_PUBLISHER_CONFIDENCE
        and not publisher_override_applied
    ):
        return blocked(
            "게시자 판별 신뢰도가 자동 기록 기준보다 낮습니다: "
            f"{confidence} < {MIN_AUTO_PUBLISHER_CONFIDENCE}"
        )

    if publisher_type not in {"OWNED", "INFLUENCER"}:
        return blocked(
            f"Publisher Type={publisher_type or 'blank'}"
        )

    original_url = get_primary_post_url(
        result_row
    )
    consumer_comment_url = (
        get_consumer_comment_url(
            result_row
        )
    )

    formatted_url_value = (
        format_publisher_url(
            original_url=original_url,
            publisher_type=publisher_type,
            consumer_comment_url=(
                consumer_comment_url
            ),
        )
        if original_url
        else None
    )

    # 게시글 유형은 Samsung 여부만으로 확정하므로 Country 매핑 성공 여부와
    # 분리한다. Country를 못 찾더라도 Influencer Yes/No와 URL 헤더는 기록하고,
    # Subsidiary 표시값만 보류한다.
    if publisher_type == "OWNED":
        influencer_value = "No"
    else:
        influencer_value = "Yes"

    if mapping_result.status != "mapped":
        return PublisherWritebackDecision(
            publisher_type=publisher_type,
            publisher_country=mapping_result.publisher_country,
            mapped_subsidiary=mapping_result.mapped_subsidiary,
            country_mapping_status=mapping_result.status,
            country_mapping_message=mapping_result.message,
            subsidiary_display_value=None,
            influencer_value=influencer_value,
            formatted_url_value=formatted_url_value,
            formatted_url_target=original_url,
            eligible=True,
            skip_reason=mapping_result.message,
            code_requires_manual_review=True,
        )

    assert mapping_result.mapped_subsidiary is not None
    assert mapping_result.publisher_country is not None

    if publisher_type == "OWNED":
        subsidiary_display = (
            f"{mapping_result.mapped_subsidiary} "
            f"({mapping_result.publisher_country})"
        )
    else:
        subsidiary_display = (
            f"Influencer ({mapping_result.mapped_subsidiary})"
        )

    return PublisherWritebackDecision(
        publisher_type=publisher_type,
        publisher_country=mapping_result.publisher_country,
        mapped_subsidiary=mapping_result.mapped_subsidiary,
        country_mapping_status=mapping_result.status,
        country_mapping_message=mapping_result.message,
        subsidiary_display_value=subsidiary_display,
        influencer_value=influencer_value,
        formatted_url_value=formatted_url_value,
        formatted_url_target=original_url,
        eligible=True,
        skip_reason=None,
        code_requires_manual_review=False,
    )


def write_publisher_decision_to_dataframe(
    result_dataframe: pd.DataFrame,
    dataframe_index: Any,
    decision: PublisherWritebackDecision,
) -> None:
    values = {
        "mapped_subsidiary": decision.mapped_subsidiary,
        "country_mapping_status": decision.country_mapping_status,
        "country_mapping_message": decision.country_mapping_message,
        "subsidiary_display_value": decision.subsidiary_display_value,
        "final_influencer_value": decision.influencer_value,
        "final_formatted_url": decision.formatted_url_value,
        "publisher_writeback_eligible": decision.eligible,
        "publisher_writeback_skip_reason": decision.skip_reason,
        "code_requires_manual_review": decision.code_requires_manual_review,
    }

    for column_name, value in values.items():
        result_dataframe.at[dataframe_index, column_name] = value



def write_output_values_to_target_row(
    worksheet: Worksheet,
    layout: TargetSheetLayout,
    target_row_number: int,
    result_row: pd.Series,
    overwrite_existing: bool,
    publisher_decision: PublisherWritebackDecision,
) -> tuple[str, str, str | None]:
    written_columns: list[str] = []
    preserved_columns: list[str] = []
    skipped_columns: list[str] = []

    writeback_values: dict[str, Any] = {
        column_name: result_row.get(column_name)
        for column_name in TARGET_OUTPUT_COLUMNS
    }
    writeback_values.update(
        {
            TARGET_SUBSIDIARY_COLUMN: (
                publisher_decision.subsidiary_display_value
            ),
            TARGET_INFLUENCER_COLUMN: (
                publisher_decision.influencer_value
            ),
        }
    )

    for column_name in TARGET_WRITEBACK_COLUMNS:
        value = writeback_values.get(column_name)

        if value is None and column_name in {
            TARGET_SUBSIDIARY_COLUMN,
            TARGET_INFLUENCER_COLUMN,
        }:
            skipped_columns.append(column_name)
            continue

        column_number = layout.header_column_map[column_name]
        target_cell = worksheet.cell(
            row=target_row_number,
            column=column_number,
        )

        if (
            not overwrite_existing
            and not is_blank_excel_value(target_cell.value)
        ):
            preserved_columns.append(column_name)
            continue

        try:
            if pd.isna(value):
                value = ""
        except (TypeError, ValueError):
            pass

        target_cell.value = "" if value is None else str(value)

        if column_name in {
            "Product",
            "CXP Product Feature",
            "Description",
            TARGET_SUBSIDIARY_COLUMN,
        }:
            new_alignment = copy(target_cell.alignment)
            new_alignment.wrap_text = True
            new_alignment.vertical = "top"
            target_cell.alignment = new_alignment

        written_columns.append(column_name)

    # URL은 행 매핑에 먼저 사용한 다음, 확정된 게시자 유형에 따라
    # [당사 게시글] 또는 [인플루언서 게시글] 헤더 형식으로 변환한다.
    #
    # 중요:
    # raw_to_processed.py가 미리 적재한 두 번째 URL(소비자 댓글)을
    # 현재 target URL 셀에서도 다시 읽어 보존한다.
    actual_formatted_url_value: str | None = None

    if (
        publisher_decision.eligible
        and publisher_decision.formatted_url_target
    ):
        url_cell = worksheet.cell(
            row=target_row_number,
            column=layout.url_column_number,
        )

        existing_post_url, existing_comment_url = (
            split_post_and_comment_urls(
                url_cell.value
            )
        )

        result_comment_url = (
            get_consumer_comment_url(
                result_row
            )
        )

        consumer_comment_url = (
            existing_comment_url
            or result_comment_url
        )

        original_url = (
            publisher_decision.formatted_url_target
            or existing_post_url
        )

        actual_formatted_url_value = (
            format_publisher_url(
                original_url=original_url,
                publisher_type=(
                    publisher_decision.publisher_type
                ),
                consumer_comment_url=(
                    consumer_comment_url
                ),
            )
        )

        url_cell.value = (
            actual_formatted_url_value
        )

        # Excel 셀 자체 hyperlink는 원 게시물로 유지한다.
        # 셀 내부의 소비자 댓글 URL은 텍스트로 함께 보존된다.
        url_cell.hyperlink = original_url

        new_alignment = copy(url_cell.alignment)
        new_alignment.wrap_text = True
        new_alignment.vertical = "top"
        url_cell.alignment = new_alignment
        written_columns.append(
            layout.url_column_name
        )

    message_parts: list[str] = []
    if written_columns:
        message_parts.append("작성=" + ", ".join(written_columns))
    if preserved_columns:
        message_parts.append(
            "기존값 보존=" + ", ".join(preserved_columns)
        )
    if skipped_columns:
        message_parts.append(
            "게시자 컬럼 자동기록 보류=" + ", ".join(skipped_columns)
        )
        if publisher_decision.skip_reason:
            message_parts.append(
                "사유=" + publisher_decision.skip_reason
            )

    if written_columns and preserved_columns:
        status = "partially_written_existing_preserved"
    elif written_columns and skipped_columns:
        status = "written_publisher_review_required"
    elif written_columns:
        status = "written"
    elif preserved_columns and skipped_columns:
        status = "existing_values_preserved_publisher_review_required"
    elif preserved_columns:
        status = "existing_values_preserved"
    else:
        status = "publisher_review_required_no_values_written"

    return (
        status,
        " | ".join(message_parts) or "작성된 값이 없습니다.",
        actual_formatted_url_value,
    )


def map_and_write_formatted_excel(
    target_excel_path: Path,
    output_excel_path: Path | None,
    result_dataframe: pd.DataFrame,
    country_to_subsidiary: dict[str, str],
    target_sheet_name: str = OUTPUT_SHEET_NAME,
    target_header_row: int | None = None,
    target_url_column: str | None = None,
    target_campaign_id_column: str | None = None,
    overwrite_existing: bool = OVERWRITE_EXISTING_TARGET_VALUES,
    dry_run: bool = False,
) -> pd.DataFrame:
    target_excel_path = Path(target_excel_path)
    if not target_excel_path.is_file():
        raise FileNotFoundError(
            f"결과 입력 대상 formatted Excel을 찾을 수 없습니다: {target_excel_path}"
        )

    workbook = load_workbook(target_excel_path)

    try:
        if target_sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"formatted Excel에서 대상 시트를 찾을 수 없습니다: "
                f"{target_sheet_name}. available={workbook.sheetnames}"
            )

        worksheet = workbook[target_sheet_name]
        layout = resolve_target_sheet_layout(
            worksheet=worksheet,
            explicit_header_row=target_header_row,
            explicit_url_column=target_url_column,
            explicit_campaign_id_column=target_campaign_id_column,
        )
        url_row_lookup = build_target_url_row_lookup(worksheet, layout)

        for dataframe_index, result_row in result_dataframe.iterrows():
            api_status = optional_text(result_row.get("api_status")) or ""

            # dry-run은 API 결과가 없어도 URL 매핑 가능성 자체를 검사한다.
            should_map = dry_run or api_status == "success"
            if not should_map:
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_status",
                ] = "not_mapped_api_not_success"
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_message",
                ] = f"api_status={api_status}"
                continue

            publisher_decision = build_publisher_writeback_decision(
                result_row=result_row,
                country_to_subsidiary=country_to_subsidiary,
            )
            write_publisher_decision_to_dataframe(
                result_dataframe=result_dataframe,
                dataframe_index=dataframe_index,
                decision=publisher_decision,
            )

            candidate_rows, normalized_urls = candidate_target_rows_for_result(
                result_row,
                url_row_lookup,
            )

            if not normalized_urls:
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_status",
                ] = "url_missing"
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_message",
                ] = "permalink/llm_input_value에서 게시물 URL을 찾지 못했습니다."
                continue

            if not candidate_rows:
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_status",
                ] = "url_not_found"
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_message",
                ] = " | ".join(normalized_urls)
                continue

            try:
                expected_campaign_id = normalize_campaign_id(
                    result_row.get("campaign_id")
                )
            except ValueError:
                expected_campaign_id = ""

            raw_row_number: int | None
            try:
                raw_row_number = normalize_integer(
                    result_row.get("raw_row_number"),
                    "raw_row_number",
                )
            except ValueError:
                raw_row_number = None

            target_row_number = disambiguate_target_row(
                worksheet=worksheet,
                layout=layout,
                candidate_rows=candidate_rows,
                expected_campaign_id=expected_campaign_id,
                raw_row_number=raw_row_number,
            )

            if target_row_number is None:
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_status",
                ] = "duplicate_url_ambiguous"
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_message",
                ] = (
                    f"candidate_rows={candidate_rows}, "
                    f"urls={list(normalized_urls)}"
                )
                continue

            result_dataframe.at[
                dataframe_index,
                "target_sheet_name",
            ] = target_sheet_name
            result_dataframe.at[
                dataframe_index,
                "target_row_number",
            ] = target_row_number

            if dry_run:
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_status",
                ] = "dry_run_url_matched"
                result_dataframe.at[
                    dataframe_index,
                    "target_mapping_message",
                ] = (
                    f"target_row={target_row_number}, "
                    f"url_column={layout.url_column_name}, "
                    f"header_row={layout.header_row_number}"
                )
                continue

            (
                write_status,
                write_message,
                actual_formatted_url_value,
            ) = write_output_values_to_target_row(
                worksheet=worksheet,
                layout=layout,
                target_row_number=target_row_number,
                result_row=result_row,
                overwrite_existing=overwrite_existing,
                publisher_decision=publisher_decision,
            )

            if actual_formatted_url_value is not None:
                result_dataframe.at[
                    dataframe_index,
                    "final_formatted_url",
                ] = actual_formatted_url_value

            result_dataframe.at[
                dataframe_index,
                "target_mapping_status",
            ] = write_status
            result_dataframe.at[
                dataframe_index,
                "target_mapping_message",
            ] = write_message

        if not dry_run:
            if output_excel_path is None:
                raise ValueError(
                    "dry_run=False일 때 output_excel_path가 필요합니다."
                )

            output_excel_path = Path(output_excel_path)
            output_excel_path.parent.mkdir(parents=True, exist_ok=True)

            if target_excel_path.resolve() == output_excel_path.resolve():
                raise ValueError(
                    "target Excel과 output Excel 경로가 같습니다. "
                    "원본 보호를 위해 다른 출력 경로를 지정하세요."
                )

            workbook.save(output_excel_path)

    finally:
        workbook.close()

    return result_dataframe


# =========================================================
# 11. CLI / Main
# =========================================================


def validate_input_date(input_date: str) -> str:
    normalized = input_date.strip()

    if not re.fullmatch(r"\d{6}", normalized):
        raise ValueError(
            f"input_date는 YYMMDD 6자리여야 합니다: {input_date!r}"
        )

    try:
        datetime.strptime(
            normalized,
            "%y%m%d",
        )
    except ValueError as exc:
        raise ValueError(
            "input_date는 실제 존재하는 YYMMDD 날짜여야 합니다. "
            f"입력값={input_date!r}, 예시='260724'"
        ) from exc

    return normalized


def infer_input_date_from_path(path: Path) -> str | None:
    match = re.match(r"(\d{6})", path.name)

    if match is None:
        return None

    try:
        return validate_input_date(
            match.group(1)
        )
    except ValueError:
        return None


def resolve_path_from_project(
    path: Path,
) -> Path:
    """
    상대경로는 현재 터미널 위치가 아니라 프로젝트 루트 기준으로 해석한다.
    """

    expanded_path = path.expanduser()

    if not expanded_path.is_absolute():
        expanded_path = BASE_DIR / expanded_path

    return expanded_path.resolve()


def extract_run_info_from_directory_name(
    directory_name: str,
) -> tuple[str, int]:
    """
    실행 폴더명에서 작업 날짜와 차수를 추출한다.

    예:
        260805      -> ("260805", 1)
        260805_1차  -> ("260805", 1)
        260805_2차  -> ("260805", 2)
    """

    match = re.fullmatch(
        r"(\d{6})(?:_(\d+)차)?",
        directory_name,
    )

    if match is None:
        raise ValueError(
            "실행 output 폴더명이 허용 형식과 일치하지 않습니다.\n"
            f"폴더명: {directory_name}\n"
            "허용 형식 예: 260805, 260805_2차"
        )

    input_date = validate_input_date(
        match.group(1)
    )

    run_number = (
        int(match.group(2))
        if match.group(2) is not None
        else 1
    )

    if run_number < 1:
        raise ValueError(
            "실행 차수는 1 이상이어야 합니다: "
            f"{directory_name}"
        )

    return input_date, run_number


def resolve_execution_output_dir(
    cli_output_dir: Path | None,
    input_date: str | None,
    required: bool,
) -> tuple[Path | None, str | None, int | None]:
    """
    이번 모듈이 사용할 기존 실행 output 폴더를 확정한다.

    우선순위:
        1. --output-dir
        2. run_pipeline.py가 전달한 LOCAL_CAMPAIGN_OUTPUT_DIR
        3. 기본 경로가 필요하지 않고 모든 파일 경로를 직접 지정한 경우 None

    이 함수는 새로운 차수 폴더를 생성하지 않는다.
    """

    environment_output_dir_text = os.getenv(
        ENV_OUTPUT_DIR
    )

    environment_output_dir = (
        resolve_path_from_project(
            Path(environment_output_dir_text)
        )
        if environment_output_dir_text
        else None
    )

    resolved_cli_output_dir = (
        resolve_path_from_project(
            cli_output_dir
        )
        if cli_output_dir is not None
        else None
    )

    if (
        resolved_cli_output_dir is not None
        and environment_output_dir is not None
        and resolved_cli_output_dir != environment_output_dir
    ):
        raise ValueError(
            "--output-dir과 run_pipeline.py가 전달한 "
            "output 경로가 서로 다릅니다.\n"
            f"--output-dir: {resolved_cli_output_dir}\n"
            f"환경변수 경로: {environment_output_dir}"
        )

    output_dir = (
        resolved_cli_output_dir
        or environment_output_dir
    )

    if output_dir is None:
        if required:
            raise RuntimeError(
                "실행 output 폴더가 지정되지 않았습니다.\n"
                "전체 실행은 run_pipeline.py를 통해 시작하세요.\n"
                "기존 차수에서 이 모듈만 단독 실행하는 경우:\n"
                "python llm_analysis_pipeline.py "
                '--output-dir "output\\260805_2차"'
            )

        return None, input_date, None

    if not output_dir.exists():
        raise FileNotFoundError(
            "지정된 실행 output 폴더를 찾을 수 없습니다.\n"
            f"경로: {output_dir}\n"
            "이 모듈은 차수 폴더를 생성하지 않습니다."
        )

    if not output_dir.is_dir():
        raise NotADirectoryError(
            "지정된 output 경로가 폴더가 아닙니다.\n"
            f"경로: {output_dir}"
        )

    directory_date, directory_run_number = (
        extract_run_info_from_directory_name(
            output_dir.name
        )
    )

    if input_date is not None and input_date != directory_date:
        raise ValueError(
            "입력 날짜와 실행 output 폴더의 날짜가 일치하지 않습니다.\n"
            f"입력 날짜: {input_date}\n"
            f"output 폴더: {output_dir}"
        )

    environment_input_date = os.getenv(
        ENV_INPUT_DATE
    )

    if (
        environment_input_date
        and environment_input_date != directory_date
    ):
        raise ValueError(
            "run_pipeline.py가 전달한 작업 날짜와 "
            "output 폴더의 날짜가 일치하지 않습니다.\n"
            f"전달 날짜: {environment_input_date}\n"
            f"output 폴더 날짜: {directory_date}"
        )

    environment_run_number_text = os.getenv(
        ENV_RUN_NUMBER
    )

    if environment_run_number_text:
        try:
            environment_run_number = int(
                environment_run_number_text
            )
        except ValueError as exc:
            raise ValueError(
                "LOCAL_CAMPAIGN_RUN_NUMBER가 정수가 아닙니다: "
                f"{environment_run_number_text}"
            ) from exc

        if environment_run_number != directory_run_number:
            raise ValueError(
                "run_pipeline.py가 전달한 실행 차수와 "
                "output 폴더의 차수가 일치하지 않습니다.\n"
                f"전달 차수: {environment_run_number}\n"
                f"폴더 차수: {directory_run_number}"
            )

    print(
        f"[INFO] 실행 차수: {directory_run_number}차"
    )
    print(
        f"[INFO] 실행 output 폴더: {output_dir}"
    )

    return (
        output_dir,
        directory_date,
        directory_run_number,
    )


def build_default_input_path(
    input_date: str,
    output_dir: Path,
) -> Path:
    """
    같은 실행 output 폴더에서 media_extractor 결과 Excel을 찾는다.
    """

    return (
        output_dir
        / f"{input_date}_campaign_media_result.xlsx"
    )


def build_default_target_excel_path(
    input_date: str,
    output_dir: Path,
) -> Path:
    """
    같은 실행 output 폴더에서 formatted Excel을 찾는다.
    """

    input_date_obj = datetime.strptime(
        input_date,
        "%y%m%d",
    )

    filename = DEFAULT_TARGET_FILENAME_TEMPLATE.format(
        input_date=input_date,
        input_month=input_date_obj.month,
    )

    return output_dir / filename


def build_default_llm_log_path(
    input_path: Path,
    output_dir: Path | None,
) -> Path:
    """
    LLM 실행 로그는 실행 output 폴더를 우선 사용한다.
    모든 경로를 직접 지정한 경우에는 입력 파일과 같은 폴더를 사용한다.
    """

    output_directory = (
        output_dir
        if output_dir is not None
        else input_path.parent
    )

    return (
        output_directory
        / f"{input_path.stem}_llm_result.xlsx"
    )


def build_default_output_path(
    target_excel_path: Path,
    output_dir: Path | None,
) -> Path:
    """
    Gemini 결과가 반영된 최종 formatted Excel을
    같은 실행 output 폴더에 저장한다.
    """

    output_directory = (
        output_dir
        if output_dir is not None
        else target_excel_path.parent
    )

    return (
        output_directory
        / (
            f"{target_excel_path.stem}_llm_completed"
            f"{target_excel_path.suffix}"
        )
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "media_extractor의 llm_input Excel을 읽어 Gemini에 post 단위로 "
            "전달하고, URL로 formatted Excel 행을 찾아 캠페인 분석 결과와 "
            "게시자 유형·국가·법인·Influencer 값을 입력합니다."
        )
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        help=(
            "기존 실행 output 폴더. "
            "예: output/260805_2차. "
            "기본 파일 경로를 사용하는 단독 실행 시 반드시 지정합니다."
        ),
    )
    parser.add_argument(
        "--input-excel",
        type=Path,
        help="입력 {date}_campaign_media_result.xlsx 경로",
    )
    parser.add_argument(
        "--target-excel",
        type=Path,
        help="Gemini 결과를 입력할 formatted Excel 경로",
    )
    parser.add_argument(
        "--input-date",
        type=str,
        help=(
            "YYMMDD. 미지정 시 --output-dir 폴더명, "
            "입력 파일명 또는 사용자 입력에서 결정"
        ),
    )
    parser.add_argument(
        "--llm-log-excel",
        type=Path,
        help="LLM 실행 로그 Excel 저장 경로",
    )
    parser.add_argument(
        "--output-excel",
        type=Path,
        help="Gemini 결과가 반영된 최종 formatted Excel 저장 경로",
    )
    parser.add_argument(
        "--overwrite-results",
        action="store_true",
        help=(
            "기존 LLM 로그/최종 formatted 결과가 있을 때 "
            "새 결과가 모두 완성된 후 기존 파일을 교체합니다."
        ),
    )
    parser.add_argument(
        "--target-sheet",
        type=str,
        default=OUTPUT_SHEET_NAME,
        help=f"formatted Excel 대상 시트명. 기본값={OUTPUT_SHEET_NAME}",
    )
    parser.add_argument(
        "--target-header-row",
        type=int,
        help="formatted 대상 시트의 헤더 Excel 행 번호. 미지정 시 자동 탐색",
    )
    parser.add_argument(
        "--target-url-column",
        type=str,
        help="formatted 대상 시트의 URL 컬럼명. 미지정 시 후보명에서 자동 탐색",
    )
    parser.add_argument(
        "--target-campaign-id-column",
        type=str,
        help="중복 URL 검증용 campaign ID 컬럼명",
    )
    parser.add_argument(
        "--overwrite-existing",
        action="store_true",
        help="formatted Excel 결과 컬럼에 기존 값이 있어도 덮어씀",
    )
    parser.add_argument(
        "--prompt-file",
        type=Path,
        help="사용자 프롬프트 txt 경로",
    )
    parser.add_argument(
        "--system-prompt-file",
        type=Path,
        help="선택 시스템 프롬프트 txt 경로",
    )
    parser.add_argument(
        "--response-schema-file",
        type=Path,
        help="출력 컬럼을 정의하는 JSON Schema 파일",
    )
    parser.add_argument(
        "--country-mapping-file",
        type=Path,
        help=(
            "Country–Subsidiary 매핑 JSON 경로. 미지정 시 "
            "config/country_subsidiary_mapping.json 사용"
        ),
    )
    parser.add_argument(
        "--model",
        type=str,
        help=(
            "사용할 Gemini 모델 ID. 미지정 시 GEMINI_MODEL 환경변수 또는 "
            "기본값 gemini-2.5-flash 사용"
        ),
    )
    parser.add_argument(
        "--project",
        type=str,
        help=(
            "Google Cloud 프로젝트 ID. 미지정 시 GOOGLE_CLOUD_PROJECT "
            "환경변수 또는 코드 기본값 사용"
        ),
    )
    parser.add_argument(
        "--location",
        type=str,
        help=(
            "Vertex AI location. 미지정 시 GOOGLE_CLOUD_LOCATION 환경변수 "
            "또는 기본값 global 사용"
        ),
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        help="실제 API 호출 최대 행 수. 샘플 테스트에 사용",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help=(
            "API 호출 및 formatted 파일 쓰기 없이 입력 파일, 미디어, "
            "프롬프트, URL 매핑만 검증"
        ),
    )

    return parser.parse_args()


def resolve_execution_paths(
    args: argparse.Namespace,
) -> tuple[
    Path,
    Path,
    Path,
    Path | None,
    str,
    Path | None,
]:
    input_date = (
        validate_input_date(args.input_date)
        if args.input_date
        else None
    )

    explicit_input_excel = (
        resolve_path_from_project(args.input_excel)
        if args.input_excel is not None
        else None
    )
    explicit_target_excel = (
        resolve_path_from_project(args.target_excel)
        if args.target_excel is not None
        else None
    )

    if input_date is None and explicit_input_excel is not None:
        input_date = infer_input_date_from_path(
            explicit_input_excel
        )

    if input_date is None and explicit_target_excel is not None:
        input_date = infer_input_date_from_path(
            explicit_target_excel
        )

    defaults_required = (
        explicit_input_excel is None
        or explicit_target_excel is None
        or args.llm_log_excel is None
        or (
            not args.dry_run
            and args.output_excel is None
        )
    )

    (
        execution_output_dir,
        directory_date,
        _run_number,
    ) = resolve_execution_output_dir(
        cli_output_dir=args.output_dir,
        input_date=input_date,
        required=defaults_required,
    )

    if input_date is None:
        input_date = directory_date

    if input_date is None and defaults_required:
        input_date = validate_input_date(
            input(
                "조회 날짜를 입력하세요 (YYMMDD): "
            ).strip()
        )

        (
            execution_output_dir,
            directory_date,
            _run_number,
        ) = resolve_execution_output_dir(
            cli_output_dir=args.output_dir,
            input_date=input_date,
            required=True,
        )

    if input_date is None:
        raise ValueError(
            "input_date를 파일명 또는 실행 output 폴더에서 "
            "확정할 수 없습니다. --input-date를 지정하세요."
        )

    input_excel_path = (
        explicit_input_excel
        if explicit_input_excel is not None
        else build_default_input_path(
            input_date=input_date,
            output_dir=execution_output_dir,
        )
    )

    target_excel_path = (
        explicit_target_excel
        if explicit_target_excel is not None
        else build_default_target_excel_path(
            input_date=input_date,
            output_dir=execution_output_dir,
        )
    )

    llm_log_excel_path = (
        resolve_path_from_project(
            args.llm_log_excel
        )
        if args.llm_log_excel is not None
        else build_default_llm_log_path(
            input_path=input_excel_path,
            output_dir=execution_output_dir,
        )
    )

    output_excel_path: Path | None

    if args.dry_run:
        output_excel_path = None
    else:
        output_excel_path = (
            resolve_path_from_project(
                args.output_excel
            )
            if args.output_excel is not None
            else build_default_output_path(
                target_excel_path=target_excel_path,
                output_dir=execution_output_dir,
            )
        )

    return (
        input_excel_path,
        target_excel_path,
        llm_log_excel_path,
        output_excel_path,
        input_date,
        execution_output_dir,
    )


def build_temporary_result_path(
    final_path: Path,
) -> Path:
    return final_path.with_name(
        f".{final_path.stem}.partial{final_path.suffix}"
    )


def prepare_result_artifacts(
    llm_log_excel_path: Path,
    output_excel_path: Path | None,
    overwrite_results: bool,
) -> tuple[Path, Path | None]:
    """
    기존 결과를 직접 수정하지 않고 임시 결과 파일 경로를 준비한다.
    """

    final_paths = [
        llm_log_excel_path,
    ]

    if output_excel_path is not None:
        final_paths.append(
            output_excel_path
        )

    existing_paths = [
        path
        for path in final_paths
        if path.exists()
    ]

    if existing_paths and not overwrite_results:
        raise FileExistsError(
            "동일한 실행 차수의 LLM 결과 파일이 이미 존재합니다.\n"
            + "\n".join(
                f"- {path}"
                for path in existing_paths
            )
            + "\n기존 결과를 새 결과로 교체하려면 "
            "--overwrite-results 옵션을 명시하세요."
        )

    temporary_llm_log_path = build_temporary_result_path(
        llm_log_excel_path
    )

    temporary_output_path = (
        build_temporary_result_path(
            output_excel_path
        )
        if output_excel_path is not None
        else None
    )

    for temporary_path in (
        temporary_llm_log_path,
        temporary_output_path,
    ):
        if (
            temporary_path is not None
            and temporary_path.exists()
        ):
            temporary_path.unlink()

    return (
        temporary_llm_log_path,
        temporary_output_path,
    )


def commit_result_artifacts(
    temporary_llm_log_path: Path,
    llm_log_excel_path: Path,
    temporary_output_path: Path | None,
    output_excel_path: Path | None,
) -> None:
    """
    LLM 로그와 최종 formatted Excel이 모두 완성된 경우에만
    기존 결과 파일을 교체한다.
    """

    artifact_pairs: list[tuple[Path, Path]] = [
        (
            temporary_llm_log_path,
            llm_log_excel_path,
        )
    ]

    if (
        temporary_output_path is not None
        and output_excel_path is not None
    ):
        artifact_pairs.append(
            (
                temporary_output_path,
                output_excel_path,
            )
        )

    for temporary_path, _ in artifact_pairs:
        if not temporary_path.is_file():
            raise FileNotFoundError(
                "최종 반영할 임시 결과 파일이 없습니다: "
                f"{temporary_path}"
            )

    backup_pairs: list[tuple[Path, Path]] = []

    try:
        for _, final_path in artifact_pairs:
            backup_path = final_path.with_name(
                f".{final_path.name}.backup"
            )

            if backup_path.exists():
                raise RuntimeError(
                    "이전 실행의 결과 백업 파일이 남아 있습니다.\n"
                    f"백업 파일: {backup_path}\n"
                    "자동으로 삭제하지 않습니다. 기존 결과를 확인한 뒤 "
                    "수동으로 복구 또는 정리하세요."
                )

            if final_path.exists():
                final_path.rename(
                    backup_path
                )
                backup_pairs.append(
                    (
                        final_path,
                        backup_path,
                    )
                )

        for temporary_path, final_path in artifact_pairs:
            os.replace(
                temporary_path,
                final_path,
            )

    except Exception:
        for _, final_path in artifact_pairs:
            if final_path.exists():
                final_path.unlink()

        for final_path, backup_path in reversed(
            backup_pairs
        ):
            if backup_path.exists():
                backup_path.rename(
                    final_path
                )

        raise

    for _, backup_path in backup_pairs:
        if backup_path.exists():
            backup_path.unlink()


def cleanup_temporary_results(
    temporary_llm_log_path: Path,
    temporary_output_path: Path | None,
) -> None:
    for temporary_path in (
        temporary_llm_log_path,
        temporary_output_path,
    ):
        if (
            temporary_path is not None
            and temporary_path.exists()
        ):
            temporary_path.unlink()


def main() -> None:
    global GEMINI_MODEL, GOOGLE_CLOUD_PROJECT, GOOGLE_CLOUD_LOCATION

    args = parse_args()

    if args.model:
        GEMINI_MODEL = args.model.strip()
    if args.project:
        GOOGLE_CLOUD_PROJECT = args.project.strip()
    if args.location:
        GOOGLE_CLOUD_LOCATION = args.location.strip()

    if args.max_rows is not None and args.max_rows < 1:
        raise ValueError(
            "--max-rows는 1 이상의 정수여야 합니다."
        )

    if (
        args.target_header_row is not None
        and args.target_header_row < 1
    ):
        raise ValueError(
            "--target-header-row는 1 이상의 정수여야 합니다."
        )

    (
        input_excel_path,
        target_excel_path,
        llm_log_excel_path,
        output_excel_path,
        input_date,
        execution_output_dir,
    ) = resolve_execution_paths(args)

    if not input_excel_path.is_file():
        raise FileNotFoundError(
            "media_extractor 결과 Excel을 찾을 수 없습니다: "
            f"{input_excel_path}"
        )

    if not target_excel_path.is_file():
        raise FileNotFoundError(
            f"formatted Excel을 찾을 수 없습니다: {target_excel_path}"
        )

    (
        temporary_llm_log_path,
        temporary_output_path,
    ) = prepare_result_artifacts(
        llm_log_excel_path=llm_log_excel_path,
        output_excel_path=output_excel_path,
        overwrite_results=args.overwrite_results,
    )

    country_mapping_path = (
        resolve_path_from_project(
            args.country_mapping_file
        )
        if args.country_mapping_file is not None
        else DEFAULT_COUNTRY_SUBSIDIARY_MAPPING_PATH
    )
    country_to_subsidiary = load_country_subsidiary_mapping(
        country_mapping_path
    )
    validate_special_account_mappings(
        country_to_subsidiary
    )
    country_mapping_summary = summarize_country_subsidiary_mapping(
        country_to_subsidiary
    )

    prompt_bundle = load_prompt_bundle(
        prompts_dir=PROMPTS_DIR,
        user_prompt_path=(
            resolve_path_from_project(
                args.prompt_file
            )
            if args.prompt_file is not None
            else None
        ),
        system_prompt_path=(
            resolve_path_from_project(
                args.system_prompt_file
            )
            if args.system_prompt_file is not None
            else None
        ),
        response_schema_path=(
            resolve_path_from_project(
                args.response_schema_file
            )
            if args.response_schema_file is not None
            else None
        ),
    )

    print(
        f"실행 Python 파일: "
        f"{Path(__file__).resolve()}"
    )
    print(
        f"조회 날짜: "
        f"{input_date}"
    )

    if execution_output_dir is not None:
        print(
            f"실행 output 폴더: "
            f"{execution_output_dir}"
        )

    print(
        f"LLM 입력 Excel: "
        f"{input_excel_path}"
    )
    print(
        f"결과 입력 대상 Excel: "
        f"{target_excel_path}"
    )
    print(
        f"대상 시트: "
        f"{args.target_sheet}"
    )
    print(
        f"사용자 프롬프트: "
        f"{prompt_bundle.user_prompt_path}"
    )
    print(
        f"Google Cloud 프로젝트: "
        f"{GOOGLE_CLOUD_PROJECT}"
    )
    print(
        f"Vertex AI location: "
        f"{GOOGLE_CLOUD_LOCATION}"
    )
    print(
        f"모델: "
        f"{GEMINI_MODEL}"
    )

    if output_excel_path is not None:
        print(
            f"최종 formatted Excel: "
            f"{output_excel_path}"
        )
    else:
        print(
            "DRY-RUN: formatted Excel 파일은 "
            "생성하지 않습니다."
        )

    try:
        result_dataframe = run_llm_pipeline(
            result_excel_path=input_excel_path,
            prompt_bundle=prompt_bundle,
            country_to_subsidiary=country_to_subsidiary,
            max_rows=args.max_rows,
            dry_run=args.dry_run,
        )

        result_dataframe = map_and_write_formatted_excel(
            target_excel_path=target_excel_path,
            output_excel_path=(
                temporary_output_path
                if not args.dry_run
                else None
            ),
            result_dataframe=result_dataframe,
            country_to_subsidiary=country_to_subsidiary,
            target_sheet_name=args.target_sheet,
            target_header_row=args.target_header_row,
            target_url_column=args.target_url_column,
            target_campaign_id_column=(
                args.target_campaign_id_column
            ),
            overwrite_existing=args.overwrite_existing,
            dry_run=args.dry_run,
        )

        save_llm_result_excel(
            output_excel_path=temporary_llm_log_path,
            result_dataframe=result_dataframe,
        )

        commit_result_artifacts(
            temporary_llm_log_path=temporary_llm_log_path,
            llm_log_excel_path=llm_log_excel_path,
            temporary_output_path=temporary_output_path,
            output_excel_path=output_excel_path,
        )

    except Exception:
        cleanup_temporary_results(
            temporary_llm_log_path=temporary_llm_log_path,
            temporary_output_path=temporary_output_path,
        )
        raise

    api_status_counts = (
        result_dataframe[
            "api_status"
        ]
        .value_counts(
            dropna=False
        )
        .to_dict()
    )
    mapping_status_counts = (
        result_dataframe[
            "target_mapping_status"
        ]
        .value_counts(
            dropna=False
        )
        .to_dict()
    )

    print(
        f"LLM 로그 Excel 저장 완료: "
        f"{llm_log_excel_path}"
    )

    if output_excel_path is not None:
        print(
            f"최종 formatted Excel 저장 완료: "
            f"{output_excel_path}"
        )

    print(
        f"API 처리 상태 요약: "
        f"{api_status_counts}"
    )
    print(
        f"URL 매핑 상태 요약: "
        f"{mapping_status_counts}"
    )


if __name__ == "__main__":
    main()
