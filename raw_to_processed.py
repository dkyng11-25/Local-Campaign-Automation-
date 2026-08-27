# ===============================================================================================
# Sprinklr Raw Excel -> Processed Excel
#
# 주요 처리:
# 1. Raw Data_원문 / Raw Data_전략법인 시트를 읽는다.
# 2. 로컬 캠페인 리스트_QHB8 시트를 새로 생성한다.
# 3. URL 컬럼에는 게시물 유형 헤더 없이 실제 Permalink만 저장한다.
# 4. Influencer 및 Subsidiary 컬럼은 이후 LLM 분석 단계에서 채우도록 비워 둔다.
# 5. Sprinklr Export 단계에서 추가한 Sender Profile 컬럼은 Raw 시트에 그대로 보존된다.
#
# 경로:
# run_pipeline.py가 전달한 실행 output 폴더를 사용한다.
# 모듈 단독 실행 시에는 --output-dir로 기존 차수 폴더를 명시한다.
# 이 모듈은 새로운 차수 폴더를 생성하지 않는다.
# ===============================================================================================

# =========================================================
# 0. Import library & Set variables
# =========================================================

import argparse
import os
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from comment_extractor import (
    CommentExtractorSession,
    build_url_cell_value,
)


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_BASE_DIR = BASE_DIR / "output"

ENV_INPUT_DATE = "LOCAL_CAMPAIGN_INPUT_DATE"
ENV_RUN_NUMBER = "LOCAL_CAMPAIGN_RUN_NUMBER"
ENV_OUTPUT_DIR = "LOCAL_CAMPAIGN_OUTPUT_DIR"

RAW_SHEET_NAMES = (
    "Raw Data_원문",
    "Raw Data_전략법인",
)

# 최근 Sprinklr Export 파일은 헤더가 1행에 생성된다.
# 과거 파일처럼 헤더가 2행에 있는 경우도 처리하기 위해
# 실제 실행 시 find_header_row()로 헤더 행을 자동 탐색한다.
HEADER_SEARCH_MAX_ROW = 10

OUTPUT_COLUMNS = [
    "#",
    "Campaign Date",
    "Campaign Image",
    "Subsidiary (Country) / Influencer (Subsidiary)",
    "Campaign Name",
    "Product",
    "CXP Product Feature",
    "Description",
    "Buzz Volume",
    "Channel",
    "Giveaway",
    "Influencer",
    "HTR/DE",
    "Conv.Card",
    "Hashtags",
    "URL",
    "비고",
    "Query",
]

NEW_SHEET_NAME = "로컬 캠페인 리스트_QHB8"

# 이 컬럼들은 이후 LLM 분석 단계에서 채운다.
LLM_RESULT_COLUMNS = {
    "Subsidiary (Country) / Influencer (Subsidiary)",
    "Influencer",
}


# =========================================================
# 1. Excel sheet 생성
# =========================================================

def make_new_sheet(
    workbook,
    new_sheet_name: str,
    output_columns: list[str],
):
    """
    새 시트를 만들고 첫 번째 행에 결과 컬럼명을 작성한다.

    이미 같은 이름의 시트가 있으면 삭제 후 다시 생성한다.
    Raw Data 시트는 삭제하거나 수정하지 않는다.
    """

    if new_sheet_name in workbook.sheetnames:
        old_ws = workbook[new_sheet_name]
        workbook.remove(old_ws)

    new_ws = workbook.create_sheet(new_sheet_name)

    for col_idx, col_name in enumerate(
        output_columns,
        start=1,
    ):
        new_ws.cell(
            row=1,
            column=col_idx,
            value=col_name,
        )

    return new_ws


# =========================================================
# 2. 원본 헤더 및 column index 추출
# =========================================================

def normalize_header(value: object) -> str:
    """
    Excel 헤더 비교를 위해 문자열 앞뒤 공백을 제거한다.
    """

    if value is None:
        return ""

    return str(value).strip()


def find_header_row(
    ws,
    required_columns: set[str],
    max_search_row: int = HEADER_SEARCH_MAX_ROW,
) -> int | None:
    """
    시트 상단에서 필수 컬럼들이 모두 존재하는 헤더 행을 찾는다.

    최근 파일의 1행 헤더와 과거 파일의 2행 헤더를 모두 처리한다.
    """

    last_row_to_check = min(
        ws.max_row,
        max_search_row,
    )

    for row_idx in range(1, last_row_to_check + 1):
        row_headers = {
            normalize_header(cell.value)
            for cell in ws[row_idx]
            if normalize_header(cell.value)
        }

        if required_columns.issubset(row_headers):
            return row_idx

    return None


def find_column_index(
    ws,
    column_name: str,
    header_row: int,
) -> int | None:
    """
    지정된 헤더 행에서 column_name과 일치하는 컬럼 번호를 반환한다.
    openpyxl 컬럼 번호는 1부터 시작한다.
    """

    for cell in ws[header_row]:
        if normalize_header(cell.value) == column_name:
            return cell.column

    return None


# =========================================================
# 3. 결과본 column index 추출
# =========================================================

def get_output_column_index(
    output_columns: list[str],
    column_name: str,
) -> int:
    """
    OUTPUT_COLUMNS에서 원하는 컬럼의 Excel 컬럼 번호를 반환한다.
    """

    if column_name not in output_columns:
        raise ValueError(
            f"{column_name} is not in OUTPUT_COLUMNS."
        )

    return output_columns.index(column_name) + 1


# =========================================================
# 4. Formatting Functions
# =========================================================


def date_formatting(value) -> datetime | None:
    """
    Created Time 값을 Python datetime 객체로 변환한다.

    Excel 셀에는 datetime 타입으로 저장하고,
    셀의 number_format을 통해 '00월 00일'로 표시한다.

    처리 가능 예:
        datetime 객체
        date 객체
        2026-08-04 17:00:19
        2026-08-04T17:00:19
        2026-08-04 17:00:19.000
        2026-08-04 17-00-19
        2026-08-04
        Jul 08, 2026, 08:00:25 PM
    """

    if value is None:
        return None

    # openpyxl이 이미 datetime으로 읽어온 경우
    if isinstance(value, datetime):
        return value

    # 날짜만 있는 객체인 경우 00:00:00을 붙여 datetime으로 변환
    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time(),
        )

    value_str = str(value).strip()

    if not value_str:
        return None

    # Excel 또는 원천 데이터에 포함될 수 있는 특수 공백 정리
    value_str = (
        value_str
        .replace("\u00a0", " ")  # non-breaking space
        .replace("\u200b", "")   # zero-width space
        .replace("\ufeff", "")   # BOM
        .strip()
    )

    # 연속된 공백을 하나로 정리
    value_str = " ".join(value_str.split())

    # ISO 형식을 우선 처리
    # 예: 2026-08-04 17:00:19
    #     2026-08-04T17:00:19
    #     2026-08-04 17:00:19.000
    iso_value = value_str.replace("Z", "+00:00")

    try:
        parsed_datetime = datetime.fromisoformat(
            iso_value
        )

        # BigQuery DATETIME은 timezone이 없는 날짜/시간 타입이므로
        # timezone 정보가 있으면 동일한 시각 값만 유지
        if parsed_datetime.tzinfo is not None:
            parsed_datetime = parsed_datetime.replace(
                tzinfo=None
            )

        return parsed_datetime

    except ValueError:
        pass

    possible_formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H-%M-%S",
        "%Y-%m-%d",
        "%b %d, %Y, %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p",
    ]

    for fmt in possible_formats:
        try:
            return datetime.strptime(
                value_str,
                fmt,
            )
        except ValueError:
            continue

    print(
        "[WARNING] Date formatting failed: "
        f"value={value!r}, "
        f"type={type(value).__name__}"
    )

    return None


def mapping_channel(value) -> str | None:
    """
    snType을 Channel 약어로 변환한다.

    YOUTUBE   -> YT
    INSTAGRAM -> IG
    FACEBOOK  -> FB
    TWITTER   -> X
    TIKTOK    -> TT
    """

    if value is None:
        return None

    channel_dict = {
        "YOUTUBE": "YT",
        "INSTAGRAM": "IG",
        "FACEBOOK": "FB",
        "TWITTER": "X",
        "TIKTOK": "TT"
    }

    normalized_value = str(value).strip().upper()

    if not normalized_value:
        return None

    if normalized_value not in channel_dict:
        raise ValueError(
            f"{normalized_value} is invalid channel."
        )

    return channel_dict[normalized_value]

def htr_de_conv_card_mapping(value) -> tuple[str | None, str | None]:

    if value is None:
        return None, None

    if value == "X":
        return None, None

    return "No", "No"



def extract_hashtags(value) -> str | None:
    """
    Conversation Stream에서 해시태그만 추출하고 줄바꿈으로 연결한다.
    """

    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    tokens = text.split()
    hashtags: list[str] = []

    for token in tokens:
        if not token.startswith("#"):
            continue

        cleaned_tag = token.strip().rstrip(
            ".,!?;:)]}」』。،"
        )

        if cleaned_tag:
            if cleaned_tag.casefold() == "#samsung":
                continue

            hashtags.append(cleaned_tag)

    if not hashtags:
        return "N/A"

    return "\n".join(hashtags)


def normalize_url(value) -> str | None:
    """
    URL 앞뒤 공백을 제거한다.

    게시물 유형 헤더는 붙이지 않고 실제 URL만 반환한다.
    """

    if value is None:
        return None

    url = str(value).strip()

    if not url:
        return None

    return url


def prepare_unique_url(
    value,
    seen_urls: set[str],
) -> str | None:
    """
    URL을 정규화하고 중복 여부를 검사한다.

    반환값에는 다음과 같은 헤더를 붙이지 않는다.
        [당사 게시글]
        [인플루언서 게시글]
        [소비자 반응]

    당사/인플루언서 여부는 이후 LLM 분석 단계에서 결정한다.
    """

    normalized_url = normalize_url(value)

    if normalized_url is None:
        return None

    if normalized_url in seen_urls:
        return None

    seen_urls.add(normalized_url)

    return normalized_url


def extract_query(value) -> str | None:
    """
    URL을 기반으로 Query문을 생성한다.
    """

    normalized_url = normalize_url(value)

    if normalized_url is None:
        return None

    url_tokens = normalized_url.split("/")
    lower_tokens = [
        token.lower()
        for token in url_tokens
    ]

    if (
        "www.instagram.com" in lower_tokens
        or "instagram.com" in lower_tokens
    ):
        non_empty_tokens = [
            token
            for token in url_tokens
            if token
        ]

        if len(non_empty_tokens) < 2:
            return None

        query = non_empty_tokens[-1]
        return f"inUrl: {query}"

    if (
        "www.facebook.com" in lower_tokens
        or "facebook.com" in lower_tokens
    ):
        non_empty_tokens = [
            token
            for token in url_tokens
            if token
        ]

        if len(non_empty_tokens) < 2:
            return None

        query = non_empty_tokens[-1]
        return f"inUrl: {query}"

    if (
        "www.youtube.com" in lower_tokens
        or "youtube.com" in lower_tokens
        or "youtu.be" in lower_tokens
    ):
        query = url_tokens[-1].replace(
            "watch?v=",
            "",
        )

        return f"inUrl: {query}"

    query = url_tokens[-1]

    if not query:
        non_empty_tokens = [
            token
            for token in url_tokens
            if token
        ]

        query = (
            non_empty_tokens[-1]
            if non_empty_tokens
            else ""
        )

    if not query:
        return None

    return f"engagingWithGuid: {query}"


# =========================================================
# 5. 원본 sheet 하나 처리
# =========================================================

def process_one_sheet(
    source_ws,
    target_ws,
    output_columns: list[str],
    start_output_row: int,
    seen_urls: set[str],
    comment_session: CommentExtractorSession,
) -> int:
    """
    Raw Data 시트 하나를 읽어 결과 시트에 작성한다.

    Influencer 및 Subsidiary 결과는 이 단계에서 추정하지 않는다.
    해당 컬럼은 이후 LLM 분석 단계에서 채운다.

    처리 후 다음에 작성할 output row 번호를 반환한다.
    """

    required_source_columns = {
        "Created Time",
        "snType column",
        "Conversation Stream",
        "Permalink",
    }

    header_row = find_header_row(
        ws=source_ws,
        required_columns=required_source_columns,
    )

    if header_row is None:
        print(
            f"[SKIP] {source_ws.title}: "
            "required header row not found"
        )
        return start_output_row

    print(
        f"[INFO] {source_ws.title}: "
        f"header row = {header_row}"
    )

    created_time_col_idx = find_column_index(
        ws=source_ws,
        column_name="Created Time",
        header_row=header_row,
    )

    sn_type_col_idx = find_column_index(
        ws=source_ws,
        column_name="snType column",
        header_row=header_row,
    )

    conversation_stream_col_idx = find_column_index(
        ws=source_ws,
        column_name="Conversation Stream",
        header_row=header_row,
    )

    permalink_col_idx = find_column_index(
        ws=source_ws,
        column_name="Permalink",
        header_row=header_row,
    )

    source_column_indexes = {
        "Created Time": created_time_col_idx,
        "snType column": sn_type_col_idx,
        "Conversation Stream": conversation_stream_col_idx,
        "Permalink": permalink_col_idx,
    }

    missing_source_columns = [
        column_name
        for column_name, column_idx
        in source_column_indexes.items()
        if column_idx is None
    ]

    if missing_source_columns:
        print(
            f"[SKIP] {source_ws.title}: "
            f"columns not found: {missing_source_columns}"
        )
        return start_output_row

    number_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="#",
    )

    campaign_date_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="Campaign Date",
    )

    channel_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="Channel",
    )

    htr_de_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name = "HTR/DE"
    )

    conv_card_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name = "Conv.Card"
    )

    hashtags_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="Hashtags",
    )

    url_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="URL",
    )

    query_output_col_idx = get_output_column_index(
        output_columns=output_columns,
        column_name="Query",
    )

    output_row = start_output_row
    written_count = 0
    skipped_duplicate_count = 0

    for row_idx in range(
        header_row + 1,
        source_ws.max_row + 1,
    ):
        raw_created_time = source_ws.cell(
            row=row_idx,
            column=created_time_col_idx,
        ).value

        if raw_created_time is None:
            continue

        campaign_date = date_formatting(
            raw_created_time
        )

        if campaign_date is None:
            continue

        raw_sn_type = source_ws.cell(
            row=row_idx,
            column=sn_type_col_idx,
        ).value

        if raw_sn_type is None:
            continue

        channel = mapping_channel(
            raw_sn_type
        )

        if channel is None:
            continue

        htr_de, conv_card = htr_de_conv_card_mapping(channel)

        raw_conversation_stream = source_ws.cell(
            row=row_idx,
            column=conversation_stream_col_idx,
        ).value

        if raw_conversation_stream is None:
            continue

        hashtags = extract_hashtags(
            raw_conversation_stream
        )

        permalink = source_ws.cell(
            row=row_idx,
            column=permalink_col_idx,
        ).value

        if permalink is None:
            continue

        plain_url = prepare_unique_url(
            value=permalink,
            seen_urls=seen_urls,
        )

        if plain_url is None:
            skipped_duplicate_count += 1
            continue

        query = extract_query(
            plain_url
        )

        url_cell_value = build_url_cell_value(
            channel=channel,
            post_url=plain_url,
            session=comment_session,
            raise_on_error=False
        )

        target_ws.cell(
            row=output_row,
            column=number_output_col_idx,
            value=output_row - 1,
        )

        target_ws.cell(
            row=output_row,
            column=campaign_date_output_col_idx,
            value=campaign_date,
        )

        target_ws.cell(
            row=output_row,
            column=channel_output_col_idx,
            value=channel,
        )

        target_ws.cell(
            row=output_row,
            column=htr_de_output_col_idx,
            value=htr_de
        )

        target_ws.cell(
            row=output_row,
            column=conv_card_output_col_idx,
            value=conv_card
        )

        target_ws.cell(
            row=output_row,
            column=hashtags_output_col_idx,
            value=hashtags,
        )

        target_ws.cell(
            row=output_row,
            column=url_output_col_idx,
            value=url_cell_value,
        )

        target_ws.cell(
            row=output_row,
            column=query_output_col_idx,
            value=query,
        )

        # 다음 두 컬럼은 LLM 분석 결과가 들어갈 자리이므로
        # 이 단계에서는 명시적으로 값을 쓰지 않는다.
        #
        # - Subsidiary (Country) / Influencer (Subsidiary)
        # - Influencer

        output_row += 1
        written_count += 1

    print(
        f"[DONE] {source_ws.title} -> {target_ws.title} "
        f"| written={written_count:,} "
        f"| duplicate_skipped={skipped_duplicate_count:,}"
    )

    return output_row


# =========================================================
# 6. 실행 인자 및 경로 결정
# =========================================================

def parse_arguments() -> argparse.Namespace:
    """
    단독 실행 시 사용할 선택 인자를 읽는다.

    전체 파이프라인에서는 run_pipeline.py가
    LOCAL_CAMPAIGN_OUTPUT_DIR 환경변수로 경로를 전달한다.
    """

    parser = argparse.ArgumentParser(
        description=(
            "Sprinklr Raw Excel을 같은 실행 차수 폴더 안에서 "
            "Processed Excel로 변환합니다."
        )
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        help=(
            "기존 실행 output 폴더. "
            "예: output/260805_2차. "
            "단독 실행 시 반드시 지정해야 합니다."
        ),
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "동일한 formatted Excel이 이미 있을 때, "
            "새 결과가 완전히 생성된 후 기존 파일을 교체합니다."
        ),
    )

    return parser.parse_args()


def resolve_path_from_project(
    path: Path,
) -> Path:
    """
    상대경로는 현재 터미널 위치가 아니라 프로젝트 루트 기준으로 해석한다.
    """

    expanded_path = path.expanduser()

    if not expanded_path.is_absolute():
        expanded_path = BASE_DIR / expanded_path

    return expanded_path.resolve()


def validate_output_directory_name(
    output_dir: Path,
    input_date: str,
) -> None:
    """
    output 폴더명이 입력 날짜와 일치하는지 검증한다.

    허용 예:
        260805
        260805_1차
        260805_2차
    """

    allowed_pattern = re.compile(
        rf"^{re.escape(input_date)}(?:_\d+차)?$"
    )

    if allowed_pattern.fullmatch(output_dir.name) is None:
        raise ValueError(
            "지정된 output 폴더명이 입력 날짜와 일치하지 않습니다.\n"
            f"입력 날짜: {input_date}\n"
            f"지정된 output 폴더: {output_dir}\n"
            "허용 형식 예: "
            f"{input_date}, {input_date}_2차"
        )


def resolve_output_directory(
    input_date: str,
    cli_output_dir: Path | None,
) -> Path:
    """
    이번 모듈이 사용할 기존 실행 output 폴더를 확정한다.

    우선순위:
        1. --output-dir
        2. LOCAL_CAMPAIGN_OUTPUT_DIR
        3. 둘 다 없으면 오류

    이 함수는 차수 폴더를 생성하지 않는다.
    """

    pipeline_input_date = os.getenv(
        ENV_INPUT_DATE
    )

    if (
        pipeline_input_date
        and pipeline_input_date != input_date
    ):
        raise ValueError(
            "run_pipeline.py에서 전달된 작업 날짜와 "
            "입력 날짜가 일치하지 않습니다.\n"
            f"전달된 작업 날짜: {pipeline_input_date}\n"
            f"입력한 날짜: {input_date}"
        )

    environment_output_dir_text = os.getenv(
        ENV_OUTPUT_DIR
    )

    environment_output_dir = (
        resolve_path_from_project(
            Path(environment_output_dir_text)
        )
        if environment_output_dir_text
        else None
    )

    resolved_cli_output_dir = (
        resolve_path_from_project(
            cli_output_dir
        )
        if cli_output_dir is not None
        else None
    )

    if (
        resolved_cli_output_dir is not None
        and environment_output_dir is not None
        and resolved_cli_output_dir != environment_output_dir
    ):
        raise ValueError(
            "--output-dir과 run_pipeline.py가 전달한 "
            "output 경로가 서로 다릅니다.\n"
            f"--output-dir: {resolved_cli_output_dir}\n"
            f"환경변수 경로: {environment_output_dir}"
        )

    output_dir = (
        resolved_cli_output_dir
        or environment_output_dir
    )

    if output_dir is None:
        raise RuntimeError(
            "실행 output 폴더가 지정되지 않았습니다.\n"
            "전체 실행은 run_pipeline.py를 통해 시작하세요.\n"
            "기존 차수에서 이 모듈만 단독 실행하는 경우에는 "
            "다음처럼 기존 폴더를 명시하세요.\n"
            "python raw_to_processed.py "
            '--output-dir "output\\260805_2차"'
        )

    if not output_dir.exists():
        raise FileNotFoundError(
            "지정된 실행 output 폴더를 찾을 수 없습니다.\n"
            f"경로: {output_dir}\n"
            "이 모듈은 차수 폴더를 생성하지 않습니다."
        )

    if not output_dir.is_dir():
        raise NotADirectoryError(
            "지정된 output 경로가 폴더가 아닙니다.\n"
            f"경로: {output_dir}"
        )

    validate_output_directory_name(
        output_dir=output_dir,
        input_date=input_date,
    )

    run_number = os.getenv(
        ENV_RUN_NUMBER
    )

    if resolved_cli_output_dir is not None:
        print(
            "[INFO] --output-dir로 지정된 기존 실행 "
            "폴더를 사용합니다."
        )
    else:
        print(
            "[INFO] run_pipeline.py에서 전달된 "
            "실행 output 폴더를 사용합니다."
        )

    if run_number:
        print(
            f"[INFO] 실행 차수: {run_number}차"
        )

    print(
        f"[INFO] 실행 output 폴더: {output_dir}"
    )

    return output_dir


def build_excel_paths(
    input_date: str,
    output_dir: Path,
) -> tuple[Path, Path]:
    """
    확정된 실행 output 폴더 안에서 입력 및 출력 Excel 경로를 생성한다.

    입력:
        {실행 output 폴더}/
        {YYMMDD}_SLCC_SOV_Local Campaign Tracking_{월}월_v01.xlsx

    출력:
        같은 실행 output 폴더 안의 *_formatted.xlsx
    """

    try:
        input_date_obj = datetime.strptime(
            input_date,
            "%y%m%d",
        )
    except ValueError as exc:
        raise ValueError(
            "날짜는 YYMMDD 형식으로 입력해야 합니다. "
            "예시: 260724"
        ) from exc

    input_month = input_date_obj.month

    input_path = output_dir / (
        f"{input_date}_SLCC_SOV_Local Campaign Tracking_"
        f"{input_month}월_v01.xlsx"
    )

    output_path = input_path.with_name(
        input_path.stem
        + "_formatted.xlsx"
    )

    if not input_path.is_file():
        raise FileNotFoundError(
            "Input Excel file not found: "
            f"{input_path}"
        )

    return input_path, output_path


def prepare_temporary_output_path(
    output_path: Path,
    overwrite: bool,
) -> Path:
    """
    기존 formatted 파일에 직접 저장하지 않고 임시 파일 경로를 준비한다.
    """

    if output_path.exists() and not overwrite:
        raise FileExistsError(
            "동일한 실행의 formatted Excel이 이미 존재합니다.\n"
            f"파일: {output_path}\n"
            "기존 결과를 새 결과로 교체하려면 "
            "--overwrite 옵션을 명시하세요."
        )

    temporary_output_path = output_path.with_name(
        f".{output_path.stem}.partial.xlsx"
    )

    if temporary_output_path.exists():
        temporary_output_path.unlink()

    return temporary_output_path


def cleanup_temporary_output(
    temporary_output_path: Path,
) -> None:
    """
    실패한 실행에서 남은 임시 Excel을 삭제한다.
    """

    if temporary_output_path.exists():
        temporary_output_path.unlink()


# =========================================================
# 7. Main Function
# =========================================================

def main() -> None:
    args = parse_arguments()

    input_date = input().strip()

    # 날짜 형식은 경로를 선택하기 전에 먼저 검증한다.
    try:
        datetime.strptime(
            input_date,
            "%y%m%d",
        )
    except ValueError as exc:
        raise ValueError(
            "날짜는 YYMMDD 형식으로 입력해야 합니다. "
            "예시: 260724"
        ) from exc

    output_dir = resolve_output_directory(
        input_date=input_date,
        cli_output_dir=args.output_dir,
    )

    input_path, output_path = build_excel_paths(
        input_date=input_date,
        output_dir=output_dir,
    )

    temporary_output_path = prepare_temporary_output_path(
        output_path=output_path,
        overwrite=args.overwrite,
    )

    print(f"Input file: {input_path}")
    print(f"Output file: {output_path}")

    # X / FB / TT의 브라우저 context를 전체 Raw 처리 동안
    # 하나의 session에서 재사용한다. 실제 Playwright/브라우저 연결은
    # 해당 플랫폼이 처음 등장할 때 lazy하게 생성된다.
    comment_session = CommentExtractorSession()

    try:
        workbook = load_workbook(
            input_path
        )

        target_ws = make_new_sheet(
            workbook=workbook,
            new_sheet_name=NEW_SHEET_NAME,
            output_columns=OUTPUT_COLUMNS,
        )

        output_row = 2
        seen_urls: set[str] = set()
        processed_sheet_count = 0

        for sheet_name in RAW_SHEET_NAMES:
            if sheet_name not in workbook.sheetnames:
                print(
                    f"[WARNING] Source sheet not found: "
                    f"{sheet_name}"
                )
                continue

            source_ws = workbook[sheet_name]

            output_row = process_one_sheet(
                source_ws=source_ws,
                target_ws=target_ws,
                output_columns=OUTPUT_COLUMNS,
                start_output_row=output_row,
                seen_urls=seen_urls,
                comment_session=comment_session,
            )

            processed_sheet_count += 1

        if processed_sheet_count == 0:
            raise ValueError(
                "No Raw Data sheets were found. "
                f"Expected one of: {RAW_SHEET_NAMES}"
            )

        # 새 결과를 임시 파일에 완성한 뒤 최종 파일로 교체한다.
        workbook.save(
            temporary_output_path
        )

        os.replace(
            temporary_output_path,
            output_path,
        )

    except Exception:
        cleanup_temporary_output(
            temporary_output_path
        )
        raise

    finally:
        # 전체 Raw 시트 처리가 끝난 뒤 X / FB persistent context와
        # TikTok CDP 연결, Playwright를 한 번만 정리한다.
        comment_session.close()

    total_output_rows = max(
        output_row - 2,
        0,
    )

    print("Done")
    print(
        f"Processed rows: "
        f"{total_output_rows:,}"
    )
    print(
        "Sender Profile columns remain preserved "
        "in the Raw Data sheets."
    )
    print(f"Output file: {output_path}")


if __name__ == "__main__":
    main()
