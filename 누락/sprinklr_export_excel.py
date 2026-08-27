# ============================================================
# 기존 Sprinklr to Excel Extraction에서 Profile URL column 추가
# ============================================================

import argparse
import json
import os
import re
import shutil
from pathlib import Path
from typing import Any

from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import requests

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from dotenv import load_dotenv

# =========================================================
# User Guide
#
# 이 파일은 누락 전용 Sprinklr Export 모듈입니다.
#
# 기본 실행:
#   python sprinklr_export_excel.py
#
# 기본 출력:
#   누락/output_누락/{YYMMDD}_누락/
#
# 특정 출력 폴더를 직접 지정:
#   python sprinklr_export_excel.py ^
#     --output-dir "output_누락\260811_누락"
#
# 동일 날짜 결과를 새 결과로 교체:
#   python sprinklr_export_excel.py ^
#     --overwrite
#
# 날짜/시간 입력 형식:
# start datetime: 2026-07-08 18:00:00
# end datetime:   2026-07-09 18:00:59
# =========================================================

# =========================================================
# 0. Project path & Widget configuration
# =========================================================

# 파일 위치:
#   Local_Campaign_Automation_version4/누락/sprinklr_export_excel.py
#
# 공통 payload:
#   Local_Campaign_Automation_version4/payload/
#
# 누락 output:
#   Local_Campaign_Automation_version4/누락/output_누락/{YYMMDD}_누락/
BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent

PAYLOAD_DIR = PROJECT_ROOT / "payload"
OUTPUT_BASE_DIR = BASE_DIR / "output_누락"

WIDGET_CONFIGS = [
    {
        "widget_name": "6.1. 누락",
        "payload_path": PAYLOAD_DIR / "payload_6_1_누락건.json",
    },
]

# =========================================================
# Raw Data 시트 컬럼 구성
# =========================================================

SENDER_PROFILE_COLUMNS = [
    "Sender Profile Available",
    "Sender Screen Name",
    "Sender Follower Count",
    "Sender Location",
    "Sender Detailed Location",
    "Sender Bio",
    "Sender Website",
    "Sender Verified",
    "Sender Verified Type",
    "Sender Profile Tags",
]

RAW_DATA_ORIGINAL_COLUMNS = [
    # 기존 컬럼
    "Conversation Stream",
    "Campaign ID",
    "Profile URL",
    "User Name",
    "Permalink",
    "Created Time",
    "snType column",
    "Media Type",
    "Media URL",
    "source_widget",
    "data_cut_start",
    "data_cut_end",
    "extracted_at",

    # 신규 Sender Profile 컬럼
    *SENDER_PROFILE_COLUMNS,
]

RAW_DATA_SUBSIDIARY_COLUMNS = [
    # 기존 컬럼
    "Conversation Stream",
    "Campaign ID",
    "Profile URL",
    "User Name",
    "Permalink",
    "Created Time",
    "snType column",
    "Author Screen Name",
    "Media Type",
    "Media URL",
    "source_widget",
    "data_cut_start",
    "data_cut_end",
    "extracted_at",

    # 신규 Sender Profile 컬럼
    *SENDER_PROFILE_COLUMNS,
]

RAW_DATA_SHEET_COLUMNS = {
    "Raw Data_원문": RAW_DATA_ORIGINAL_COLUMNS,
    "Raw Data_전략법인": RAW_DATA_SUBSIDIARY_COLUMNS,
}

# =========================================================
# 1. 사용자 설정값 (URL, API Key, Access Token)
# =========================================================

"""사용자 설정 필요"""
SPRINKLR_BASE_URL = "https://api3.sprinklr.com/prod"
ENDPOINT = "/api/v2/reports/query"

# 보안상 실제 값은 코드에 직접 쓰기보다 환경변수로 관리
load_dotenv()
API_KEY = os.getenv("SPRINKLR_API_KEY")
ACCESS_TOKEN = os.getenv("SPRINKLR_ACCESS_TOKEN")

def parse_arguments() -> argparse.Namespace:
    """
    누락 모듈 단독 실행 시 사용할 선택 인자를 읽는다.

    --output-dir을 생략하면 자동으로:
        누락/output_누락/{YYMMDD}_누락
    폴더를 사용한다.
    """

    parser = argparse.ArgumentParser(
        description=(
            "누락 건 Sprinklr 데이터를 추출하여 "
            "누락/output_누락 폴더에 Excel로 저장합니다."
        )
    )

    parser.add_argument(
        "--output-dir",
        type=Path,
        help=(
            "누락 결과를 저장할 폴더. "
            "예: output_누락/260811_누락. "
            "생략하면 종료 날짜 기준 폴더를 자동 사용합니다."
        ),
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "동일 날짜의 최종 Excel이 이미 있을 때 "
            "새 결과가 완전히 생성된 후 기존 파일을 교체합니다."
        ),
    )

    return parser.parse_args()


def resolve_path_from_missing_module(
    path: Path,
) -> Path:
    """
    상대경로는 현재 터미널 위치가 아니라 누락 모듈 폴더 기준으로 해석한다.
    """

    expanded_path = path.expanduser()

    if not expanded_path.is_absolute():
        expanded_path = BASE_DIR / expanded_path

    return expanded_path.resolve()


def validate_output_directory_name(
    output_dir: Path,
    naming_date: str,
) -> None:
    """
    누락 output 폴더명이 {YYMMDD}_누락 형식인지 확인한다.
    """

    expected_name = f"{naming_date}_누락"

    if output_dir.name != expected_name:
        raise ValueError(
            "누락 output 폴더명이 종료 날짜와 일치하지 않습니다.\n"
            f"종료 날짜 기준 작업일: {naming_date}\n"
            f"지정된 output 폴더: {output_dir}\n"
            f"기대 폴더명: {expected_name}"
        )


def resolve_output_directory(
    naming_date: str,
    cli_output_dir: Path | None,
) -> Path:
    """
    누락 결과를 저장할 output 폴더를 확정한다.

    우선순위:
        1. --output-dir
        2. 기본값: 누락/output_누락/{YYMMDD}_누락

    기존 누락 파이프라인과 동일하게 필요한 폴더는 자동 생성한다.
    """

    if cli_output_dir is not None:
        output_dir = resolve_path_from_missing_module(
            cli_output_dir
        )
        print(
            "[INFO] --output-dir로 지정된 누락 output 폴더를 사용합니다."
        )
    else:
        output_dir = (
            OUTPUT_BASE_DIR
            / f"{naming_date}_누락"
        ).resolve()
        print(
            "[INFO] 종료 날짜 기준 누락 output 폴더를 자동 사용합니다."
        )

    validate_output_directory_name(
        output_dir=output_dir,
        naming_date=naming_date,
    )

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    print(
        f"[INFO] 누락 output 폴더: {output_dir}"
    )

    return output_dir


def create_empty_workbook() -> Workbook:
    """
    기존 파일을 불러오지 않고 새로운 빈 Workbook을 생성한다.

    동일한 데이터 컷 재실행 시 기존 행 아래에 중복 추가되는 것을
    방지하기 위해 Sprinklr Export는 항상 새 Workbook에서 시작한다.
    """

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    return workbook


def prepare_output_artifacts(
    output_excel_path: Path,
    overwrite: bool,
) -> tuple[Path, Path, Path]:
    """
    최종 파일을 바로 수정하지 않고 임시 산출물 경로를 준비한다.

    반환:
        temporary_excel_path
        temporary_response_dir
        final_response_dir
    """

    if output_excel_path.exists() and not overwrite:
        raise FileExistsError(
            "동일 날짜의 누락 최종 Excel 파일이 이미 존재합니다.\n"
            f"파일: {output_excel_path}\n"
            "중복 적재를 방지하기 위해 실행을 중단합니다.\n"
            "기존 결과를 새 결과로 교체하려면 "
            "--overwrite 옵션을 명시하세요."
        )

    temporary_excel_path = output_excel_path.with_name(
        f".{output_excel_path.stem}.partial.xlsx"
    )

    temporary_response_dir = (
        output_excel_path.parent
        / ".sprinklr_response_samples.partial"
    )

    final_response_dir = (
        output_excel_path.parent
        / "sprinklr_response_samples"
    )

    # 이전 실패 실행에서 남은 임시 산출물만 정리한다.
    if temporary_excel_path.exists():
        temporary_excel_path.unlink()

    if temporary_response_dir.exists():
        shutil.rmtree(temporary_response_dir)

    temporary_response_dir.mkdir(
        parents=False,
        exist_ok=False,
    )

    return (
        temporary_excel_path,
        temporary_response_dir,
        final_response_dir,
    )


def commit_output_artifacts(
    temporary_excel_path: Path,
    output_excel_path: Path,
    temporary_response_dir: Path,
    final_response_dir: Path,
) -> None:
    """
    모든 Widget 처리가 성공한 경우에만 임시 산출물을 최종 위치로 반영한다.

    응답 샘플 폴더를 먼저 교체하고, 마지막에 Excel을 os.replace로
    반영한다. Excel 교체가 실패하면 응답 샘플 폴더도 이전 상태로
    되돌리도록 rollback을 시도한다.
    """

    if not temporary_excel_path.exists():
        raise FileNotFoundError(
            "최종 반영할 임시 Excel 파일이 없습니다: "
            f"{temporary_excel_path}"
        )

    if not temporary_response_dir.is_dir():
        raise FileNotFoundError(
            "최종 반영할 임시 응답 샘플 폴더가 없습니다: "
            f"{temporary_response_dir}"
        )

    response_backup_dir = (
        final_response_dir.parent
        / ".sprinklr_response_samples.backup"
    )

    if response_backup_dir.exists():
        shutil.rmtree(response_backup_dir)

    previous_response_backed_up = False
    new_response_committed = False

    try:
        if final_response_dir.exists():
            final_response_dir.rename(
                response_backup_dir
            )
            previous_response_backed_up = True

        temporary_response_dir.rename(
            final_response_dir
        )
        new_response_committed = True

        # 최종 Excel은 마지막에 교체한다. 기존 파일이 있어도
        # 새 Excel이 완전히 저장된 뒤에만 교체된다.
        os.replace(
            temporary_excel_path,
            output_excel_path,
        )

    except Exception:
        if new_response_committed and final_response_dir.exists():
            shutil.rmtree(final_response_dir)

        if previous_response_backed_up and response_backup_dir.exists():
            response_backup_dir.rename(
                final_response_dir
            )

        raise

    else:
        if response_backup_dir.exists():
            shutil.rmtree(response_backup_dir)


def cleanup_temporary_artifacts(
    temporary_excel_path: Path,
    temporary_response_dir: Path,
) -> None:
    """
    실패한 실행의 임시 파일과 임시 응답 폴더를 정리한다.
    """

    if temporary_excel_path.exists():
        temporary_excel_path.unlink()

    if temporary_response_dir.exists():
        shutil.rmtree(temporary_response_dir)


# =========================================================
# 2. 날짜/시간 문자열 변환 (milliseconds)
# =========================================================

def datetime_to_milliseconds(
        datetime_str: str,
        timezone_str: str = "Asia/Seoul"
        ) -> int:
    """
    사용자가 입력한 날짜/시간 문자열을 milliseconds로 변환

    입력예시: 
        "2024-06-01 00:00:00"

    출력예시:
        "1711920000000"
    """
    
    try:
        dt = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(
            f"Invalid datetime format: {datetime_str}."
            "Expected format is YYYY-MM-DD HH:MM:SS, e.g. 2026-07-01 00:00:00"
        )

    # 사용자가 입력한 시간을 지정 timezone 기준 시간으로 해석
    dt = dt.replace(tzinfo=ZoneInfo(timezone_str))

    # epoch seconds → milliseconds 변환
    epoch_ms = int(dt.timestamp() * 1000)

    return epoch_ms

def build_time_range_from_datetimes(
    start_datetime_str: str,
    end_datetime_str: str,
    timezone_str: str = "Asia/Seoul"
) -> tuple[int, int, str, str]:
    """
    사용자가 입력한 시작/종료 datetime 문자열을
    Sprinklr payload용 startTime/endTime milliseconds로 변환한다.

    입력 형식:
        YYYY-MM-DD HH:MM:SS

    입력 예시:
        start_datetime_str = "2026-07-07 18:30:00"
        end_datetime_str   = "2026-07-08 16:45:59"
    """

    try:
        datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S")
        datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise ValueError(
            "Invalid datetime format. Expected format is YYYY-MM-DD HH:MM:SS, "
            "e.g. 2026-07-07 18:30:00"
        )

    start_time_ms = datetime_to_milliseconds(
        start_datetime_str,
        timezone_str=timezone_str
    )

    end_time_ms = datetime_to_milliseconds(
        end_datetime_str,
        timezone_str=timezone_str
    )

    if start_time_ms >= end_time_ms:
        raise ValueError(
            "Invalid time range. startTime must be earlier than endTime. "
            f"start={start_datetime_str}, end={end_datetime_str}"
        )

    return start_time_ms, end_time_ms, start_datetime_str, end_datetime_str

# =========================================================
# 3. Excel load or create
# =========================================================

def load_or_create_excel(excel_path: str | Path) -> Workbook:
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)  # Ensure parent directories exist

    # Excel 파일이 이미 있으면 열기
    if path.exists():
        workbook = load_workbook(path)
    else:
        # 없으면 새 workbook 생성
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
    
    return workbook


def get_raw_data_sheet_columns(
    sheet_name: str,
) -> list[str]:
    """Raw Data 시트별 최종 컬럼 순서를 반환한다."""

    if sheet_name not in RAW_DATA_SHEET_COLUMNS:
        raise ValueError(
            f"Unsupported Raw Data sheet: {sheet_name}"
        )

    return list(RAW_DATA_SHEET_COLUMNS[sheet_name])


def ensure_raw_data_sheets(
    workbook: Workbook,
) -> None:
    """
    Raw Data_원문과 Raw Data_전략법인 시트를 항상 생성한다.

    해당 기간에 조회된 게시글이 0건이어도 두 시트는 유지되며,
    각 시트의 1행에는 최종 컬럼 헤더가 기록된다.
    """

    for sheet_name, expected_headers in RAW_DATA_SHEET_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(
                title=sheet_name
            )
            sheet.append(list(expected_headers))

            print(
                f"Created empty sheet with headers: "
                f"{sheet_name}"
            )
            continue

        sheet = workbook[sheet_name]

        existing_headers = [
            sheet.cell(
                row=1,
                column=column_index,
            ).value
            for column_index in range(
                1,
                len(expected_headers) + 1,
            )
        ]

        header_is_blank = all(
            value is None
            or (
                isinstance(value, str)
                and not value.strip()
            )
            for value in existing_headers
        )

        # 완전히 비어 있는 기존 시트에는 헤더만 작성한다.
        if header_is_blank and sheet.max_row == 1:
            for column_index, header in enumerate(
                expected_headers,
                start=1,
            ):
                sheet.cell(
                    row=1,
                    column=column_index,
                    value=header,
                )

            print(
                f"Initialized empty sheet headers: "
                f"{sheet_name}"
            )
            continue

        extra_headers = [
            sheet.cell(
                row=1,
                column=column_index,
            ).value
            for column_index in range(
                len(expected_headers) + 1,
                sheet.max_column + 1,
            )
        ]

        has_extra_headers = any(
            value is not None
            and not (
                isinstance(value, str)
                and not value.strip()
            )
            for value in extra_headers
        )

        if (
            existing_headers != list(expected_headers)
            or has_extra_headers
        ):
            raise ValueError(
                "Existing Raw Data sheet schema does not "
                "match the expected schema.\n"
                f"Sheet: {sheet_name}\n"
                f"Existing headers: {existing_headers}\n"
                f"Expected headers: {list(expected_headers)}"
            )


# =========================================================
# 4. Payload load & open
# =========================================================

def load_payload(payload_path: str | Path) -> dict:
    path = Path(payload_path)

    if not path.exists():
        raise FileNotFoundError(f"Payload file not found: {payload_path}")
    
    if path.suffix.lower() != ".json":
        raise ValueError(f"Payload file must be a JSON file: {payload_path}")
    
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if not isinstance(payload, dict):
        raise ValueError(f"Payload JSON must be an dict but got {type(payload)}")

    return payload

# =========================================================
# 5. save modified payload (backup: optional)
# =========================================================

def save_payload(
    payload: dict,
    payload_path: str | Path,
    make_backup: bool = False,
) -> None:
    """
    수정된 payload dict를 원래 JSON 파일에 저장

    make_backup=True이면 기존 payload 파일을 .bak 파일로 백업한 뒤 저장
    """

    path = Path(payload_path)

    if make_backup and path.exists():
        backup_path = path.with_suffix(path.suffix + ".bak")
        with open(path, "r", encoding="utf-8") as src:
            original_text = src.read()

        with open(backup_path, "w", encoding="utf-8") as bak:
            bak.write(original_text)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

# =========================================================
# 6. Payload 날짜 변경 (사용자 지정) 
# =========================================================

def update_payload_time_range(
        payload: dict,
        start_time_ms: int,
        end_time_ms: int
    ) -> dict:
    
    payload["startTime"] = start_time_ms
    payload["endTime"] = end_time_ms

    return payload

# =========================================================
# 7. Sprinklr API 호출
# =========================================================

def fetch_sprinklr_data(
    base_url: str,
    endpoint: str,
    api_key: str,
    access_token: str,
    payload: dict
) -> dict:
    if not api_key:
        raise ValueError("API_KEY is missing. Check SPRINKLR_API_KEY environment variable.")

    if not access_token:
        raise ValueError("ACCESS_TOKEN is missing. Check SPRINKLR_ACCESS_TOKEN environment variable.")

    url = f"{base_url}{endpoint}"

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Key": api_key,
        "Content-Type": "application/json",
    }

    response = requests.post(
        url,
        headers=headers,
        json=payload,
        timeout=(10, 180)
    )

    # HTTP error 발생 시 여기서 에러 발생
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError:
        print("HTTP status:", response.status_code)
        print("Request URL:", url)
        print("Response text:")
        print(response.text[:3000])
        raise

    return response.json()

# =========================================================
# 8. Sprinklr response → DataFrame 변환
# =========================================================

def get_expected_columns_from_payload(payload: dict) -> list[str]:
    """
    payload의 groupBys/projections에서 예상 column 이름을 가져옴

    예:
        groupBys: POST_ID, ACCOUNT_ID
        projections: TOTAL_ENGAGEMENT

    결과:
        ["POST_ID", "ACCOUNT_ID", "TOTAL_ENGAGEMENT"]
    """
    columns = []

    for group_by in payload.get("groupBys", []):
        column_name = (
            group_by.get("heading")
            or group_by.get("dimensionName")
        )
        if column_name:
            columns.append(column_name)

    for projection in payload.get("projections", []):
        column_name = (
            projection.get("heading")
            or projection.get("measurementName")
        )
        if column_name:
            columns.append(column_name)

    return columns

# =========================================================
# 9. Sprinklr response에서 필요 row / heading 추출
# =========================================================
def find_rows_and_headings_in_response(response_json: dict):
    headings = None

    # Case 1. response_json["data"]가 dict인 경우
    # 예:
    # {
    #   "data": {
    #       "headings": [...],
    #       "rows": [...]
    #   }
    # }
    if isinstance(response_json, dict) and isinstance(response_json.get("data"), dict):
        data = response_json["data"]

        if isinstance(data.get("headings"), list):
            headings = data["headings"]

        if isinstance(data.get("rows"), list):
            return data["rows"], headings, "data.rows"

        if isinstance(data.get("results"), list):
            return data["results"], headings, "data.results"

        if isinstance(data.get("values"), list):
            return data["values"], headings, "data.values"

        if isinstance(data.get("data"), list):
            return data["data"], headings, "data.data"

        if headings is not None:
            return [], headings, "data.headings_only"

    # Case 2. response_json["data"] 자체가 list인 경우
    if isinstance(response_json, dict) and isinstance(response_json.get("data"), list):
        return response_json["data"], headings, "data"

    # Case 3. top-level rows
    if isinstance(response_json, dict) and isinstance(response_json.get("rows"), list):
        return response_json["rows"], headings, "rows"

    # Case 4. top-level results
    if isinstance(response_json, dict) and isinstance(response_json.get("results"), list):
        return response_json["results"], headings, "results"

    # Case 5. top-level headings only
    if isinstance(response_json, dict) and isinstance(response_json.get("headings"), list):
        headings = response_json["headings"]
        return [], headings, "headings_only"

    return None, headings, None

# =========================================================
# 10. 추출된 Created time + 9h
# =========================================================
def add_9_hours_to_created_time(value) -> str:
    """
    Sprinklr에서 추출된 Created Time 값에 9시간을 더한 뒤
    YYYY-MM-DD HH:MM:SS 형식으로 변환한다.

    처리 가능 입력:
        1. milliseconds timestamp
           예: 1783504943000

        2. Sprinklr 날짜 문자열
           예: "Jul 08, 2026, 08:00:25 PM"

    예:
        "Jul 08, 2026, 08:00:25 PM"
        -> "2026-07-09 05:00:25"
    """

    if value is None:
        return None

    # -----------------------------------------------------
    # Case 1. milliseconds timestamp인 경우
    # 예: 1783504943000
    # -----------------------------------------------------
    try:
        value_int = int(value)

        # milliseconds 기준으로 9시간 더하기
        adjusted_ms = value_int + (9 * 60 * 60 * 1000)

        dt = datetime.fromtimestamp(adjusted_ms / 1000, tz=ZoneInfo("UTC"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")

    except (ValueError, TypeError):
        pass

    # -----------------------------------------------------
    # Case 2. 문자열 날짜인 경우
    # 예: "Jul 08, 2026, 08:00:25 PM"
    # -----------------------------------------------------
    value_str = str(value).strip()

    possible_formats = [
        "%b %d, %Y, %I:%M:%S %p",  # Jul 08, 2026, 08:00:25 PM
        "%B %d, %Y, %I:%M:%S %p",  # July 08, 2026, 08:00:25 PM
    ]

    for fmt in possible_formats:
        try:
            dt = datetime.strptime(value_str, fmt)
            dt = dt + timedelta(hours=9)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    # 어떤 형식으로도 변환이 안 되면 원본 반환
    return value

# =========================================================
# 11. Sender Profile 필드 추출
# =========================================================

def optional_text(value: object) -> str | None:
    """
    값을 공백이 제거된 문자열로 변환한다.

    None 또는 빈 문자열이면 None을 반환한다.
    """

    if value is None:
        return None

    text = str(value).strip()
    return text or None


def extract_tiktok_account_from_permalink(
    permalink: object,
) -> tuple[str | None, str | None]:
    """
    TikTok 게시물 permalink에서 계정명과 Profile URL을 추출한다.

    예:
        https://www.tiktok.com/@bluebrywow/video/123
        -> ("bluebrywow", "https://www.tiktok.com/@bluebrywow")

    TikTok 전용 fallback이며 다른 플랫폼에는 사용하지 않는다.
    """

    permalink_text = optional_text(permalink)
    if not permalink_text:
        return None, None

    marker = "tiktok.com/@"
    marker_index = permalink_text.lower().find(marker)

    if marker_index < 0:
        return None, None

    username_start = marker_index + len(marker)
    username_part = permalink_text[username_start:]

    username = (
        username_part
        .split("/", 1)[0]
        .split("?", 1)[0]
        .split("#", 1)[0]
        .strip()
    )

    if not username:
        return None, None

    profile_url = f"https://www.tiktok.com/@{username}"

    return username, profile_url


def join_text_values(value: object) -> str | None:
    """
    문자열 또는 리스트 값을 Excel 셀에 저장할 수 있는
    줄바꿈 문자열로 변환한다.
    """

    if value is None:
        return None

    if isinstance(value, (list, tuple, set)):
        cleaned_values: list[str] = []

        for item in value:
            if item is None:
                continue

            if isinstance(item, dict):
                item_text = json.dumps(
                    item,
                    ensure_ascii=False,
                )
            else:
                item_text = str(item).strip()

            if item_text:
                cleaned_values.append(item_text)

        return "\n".join(cleaned_values) or None

    if isinstance(value, dict):
        return json.dumps(
            value,
            ensure_ascii=False,
        )

    return optional_text(value)


def extract_profile_tag_names(
    profile_tags: object,
) -> str | None:
    """profileTags에서 tagName만 추출한다."""

    if isinstance(profile_tags, dict):
        profile_tag_items = [profile_tags]
    elif isinstance(profile_tags, list):
        profile_tag_items = profile_tags
    else:
        return optional_text(profile_tags)

    tag_names: list[str] = []

    for profile_tag in profile_tag_items:
        if not isinstance(profile_tag, dict):
            tag_text = optional_text(profile_tag)
            if tag_text:
                tag_names.append(tag_text)
            continue

        tag_name = optional_text(
            profile_tag.get("tagName")
        )
        if tag_name:
            tag_names.append(tag_name)

    return "\n".join(tag_names) or None


def extract_profile_websites(
    sender_profile: dict[str, Any],
) -> str | None:
    """
    contactInfo.website을 우선 사용하고, 값이 없으면
    urlEntities.bio의 expanded_url 등을 보조로 사용한다.
    """

    contact_info = sender_profile.get("contactInfo")
    if not isinstance(contact_info, dict):
        contact_info = {}

    website = join_text_values(
        contact_info.get("website")
    )
    if website:
        return website

    url_entities = sender_profile.get("urlEntities")
    if not isinstance(url_entities, dict):
        return None

    bio_url_entities = url_entities.get("bio")
    if isinstance(bio_url_entities, dict):
        bio_url_entities = [bio_url_entities]

    if not isinstance(bio_url_entities, list):
        return None

    expanded_urls: list[str] = []

    for url_entity in bio_url_entities:
        if not isinstance(url_entity, dict):
            continue

        expanded_url = (
            url_entity.get("expanded_url")
            or url_entity.get("normalized_url")
            or url_entity.get("url")
        )
        expanded_url_text = optional_text(expanded_url)

        if expanded_url_text:
            expanded_urls.append(expanded_url_text)

    return join_text_values(expanded_urls)


def extract_sender_profile_fields(
    message_obj: dict[str, Any],
) -> dict[str, Any]:
    """
    senderProfile에서 게시자 분류와 국가 판별에 필요한
    핵심 필드만 추출한다.

    senderProfile 또는 일부 key가 없어도 항상 동일한
    Excel 컬럼 구조를 반환한다.
    """

    sender_profile = message_obj.get("senderProfile")

    if not isinstance(sender_profile, dict):
        return {
            "Profile URL": None,
            "User Name": None,
            "Sender Profile Available": False,
            "Sender Screen Name": None,
            "Sender Follower Count": None,
            "Sender Location": None,
            "Sender Detailed Location": None,
            "Sender Bio": None,
            "Sender Website": None,
            "Sender Verified": None,
            "Sender Verified Type": None,
            "Sender Profile Tags": None,
        }

    demographics_additional = sender_profile.get(
        "demographicsAdditional"
    )
    if not isinstance(demographics_additional, dict):
        demographics_additional = {}

    profile_url = (
        sender_profile.get("profileUrl")
        or sender_profile.get("permalink")
    )

    detailed_location = (
        demographics_additional.get("LOCATION_DETAILS")
        or demographics_additional.get("locationDetails")
    )

    return {
        # 기존 컬럼
        "Profile URL": optional_text(profile_url),
        "User Name": optional_text(
            sender_profile.get("name")
        ),

        # 신규 Sender Profile 컬럼
        "Sender Profile Available": True,
        "Sender Screen Name": optional_text(
            sender_profile.get("screenName")
        ),
        "Sender Follower Count": optional_text(
            sender_profile.get("followers")
        ),
        "Sender Location": optional_text(
            sender_profile.get("location")
        ),
        "Sender Detailed Location": optional_text(
            detailed_location
        ),
        "Sender Bio": optional_text(
            sender_profile.get("bio")
        ),
        "Sender Website": extract_profile_websites(
            sender_profile
        ),
        "Sender Verified": sender_profile.get(
            "verified"
        ),
        "Sender Verified Type": optional_text(
            sender_profile.get("verifiedType")
        ),
        "Sender Profile Tags": extract_profile_tag_names(
            sender_profile.get("profileTags")
        ),
    }


# =========================================================
# 12. Conversation stream 관련 필드 추출
# =========================================================
def make_conversation_stream_dataframe(
    response_json: dict,
    target_sheet_name: str,
) -> pd.DataFrame:
    """
    Sprinklr response에서 Conversation Stream 데이터를 추출하여
    DataFrame으로 변환한다.

    공통 추출 정보:
        - Conversation Stream / Campaign ID
        - Profile URL / User Name
        - Sender Profile 식별자, 계정명, 위치, Bio, Website
        - Sender 인증 정보 및 Sprinklr Profile Tags
        - Permalink / Created Time / Platform
        - Media Type / Media URL

    Raw Data_전략법인:
        - 공통 컬럼
        - Author Screen Name
    """

    rows, _, _ = find_rows_and_headings_in_response(
        response_json
    )

    if rows is None:
        raise ValueError(
            "Could not find rows in Sprinklr response. "
            "Please inspect saved response sample."
        )

    records: list[dict] = []

    for row in rows:
        message_obj = None

        # -------------------------------------------------
        # row에서 실제 message object 추출
        # -------------------------------------------------
        if isinstance(row, list):
            for cell in row:
                if isinstance(cell, dict):
                    message_obj = cell
                    break

        elif isinstance(row, dict):
            message_obj = row

        if not isinstance(message_obj, dict):
            continue

        # -------------------------------------------------
        # Sender Profile 추출
        # -------------------------------------------------
        sender_profile_fields = extract_sender_profile_fields(
            message_obj
        )

        # 기존 전략법인 컬럼과의 호환성을 위해
        # Author Screen Name에는 기존처럼 profile name을 사용한다.
        author_screen_name = sender_profile_fields.get(
            "User Name"
        )

        # -------------------------------------------------
        # 공통 record 생성
        # -------------------------------------------------
        sender_profile_obj = message_obj.get(
            "senderProfile"
        )

        sender_platform = None
        if isinstance(sender_profile_obj, dict):
            sender_platform = sender_profile_obj.get(
                "snType"
            )

        record = {
            "Conversation Stream": message_obj.get(
                "message"
            ),
            "Campaign ID": message_obj.get(
                "snMsgId"
            ),
            "Permalink": message_obj.get(
                "permalink"
            ),
            "Created Time": add_9_hours_to_created_time(
                message_obj.get("snCreatedTime")
            ),
            "snType column": (
                message_obj.get("snType")
                or sender_platform
            ),
            "Media Type": None,
            "Media URL": None,
        }

        # Profile URL, User Name 및 신규 Sender Profile 필드 추가
        record.update(sender_profile_fields)

        sn_type = str(
            message_obj.get("snType") or ""
        ).strip().upper()

        # -------------------------------------------------
        # TikTok 계정 정보 fallback
        # -------------------------------------------------
        # TikTok senderProfile이 Anonymous User로 내려오는 경우에만
        # 게시물 permalink의 @username을 이용해 계정 정보를 보완한다.
        # 다른 플랫폼의 Sender Profile 처리 로직에는 영향을 주지 않는다.
        if sn_type == "TIKTOK":
            tiktok_username, tiktok_profile_url = (
                extract_tiktok_account_from_permalink(
                    record.get("Permalink")
                )
            )

            anonymous_profile_values = {
                "ANONYMOUS",
                "ANONYMOUS USER",
            }

            current_user_name = optional_text(
                record.get("User Name")
            )
            current_screen_name = optional_text(
                record.get("Sender Screen Name")
            )

            if (
                tiktok_username
                and (
                    not current_user_name
                    or current_user_name.upper()
                    in anonymous_profile_values
                )
            ):
                record["User Name"] = tiktok_username

            if (
                tiktok_username
                and (
                    not current_screen_name
                    or current_screen_name.upper()
                    in anonymous_profile_values
                )
            ):
                record["Sender Screen Name"] = (
                    tiktok_username
                )

            if (
                tiktok_profile_url
                and not optional_text(
                    record.get("Profile URL")
                )
            ):
                record["Profile URL"] = (
                    tiktok_profile_url
                )

            # 전략법인 시트의 Author Screen Name도
            # TikTok에서 보완된 User Name을 사용한다.
            author_screen_name = record.get(
                "User Name"
            )

        # -------------------------------------------------
        # mediaList 정규화
        # dict이면 list로 변환하고,
        # list이면 그대로 사용하며,
        # 그 외에는 빈 리스트로 처리
        # -------------------------------------------------
        media_list = message_obj.get("mediaList")

        if isinstance(media_list, dict):
            media_items = [media_list]

        elif isinstance(media_list, list):
            media_items = media_list

        else:
            media_items = []

        # -------------------------------------------------
        # 추출된 미디어 저장
        # -------------------------------------------------
        media_pairs: list[tuple[str, str]] = []

        # URL 중복 제거용
        seen_urls: set[str] = set()

        def add_media(
            media_type_value,
            source_value,
        ) -> None:
            """
            media type과 URL을 media_pairs에 추가한다.

            source_value가 없거나 빈 문자열이면 추가하지 않는다.
            동일 URL이 이미 추가되어 있으면 중복 추가하지 않는다.
            """

            if not source_value:
                return

            source_text = str(source_value).strip()

            if not source_text:
                return

            if source_text in seen_urls:
                return

            media_type_text = str(
                media_type_value or "UNKNOWN"
            ).strip().upper()

            seen_urls.add(source_text)

            media_pairs.append(
                (
                    media_type_text,
                    source_text,
                )
            )

        # =================================================
        # Twitter
        # =================================================
        if sn_type == "TWITTER":
            for media_item in media_items:
                if not isinstance(media_item, dict):
                    continue

                raw_media_type = media_item.get("type")

                # mediaList 항목에 type 키가 없거나
                # type 값이 비어 있으면 해당 항목은 사용하지 않는다.
                #
                # 전체 mediaList 처리 후 유효 미디어가 하나도 없으면
                # UNKNOWN + Permalink fallback이 실행된다.
                if not raw_media_type:
                    continue

                media_type = str(
                    raw_media_type
                ).strip().upper()

                media_url = None

                # -----------------------------------------
                # Twitter 이미지
                # -----------------------------------------
                if media_type == "PHOTO":
                    media_url = (
                        media_item.get("picture")
                        or media_item.get("source")
                    )

                # -----------------------------------------
                # Twitter 비디오
                # -----------------------------------------
                elif media_type == "VIDEO":
                    media_url = media_item.get("source")

                    if not media_url: 
                        additional = media_item.get("additional")

                        if isinstance(additional, dict):
                            media_url = additional.get("orgSMUrl")

                # -----------------------------------------
                # 알 수 없는 Twitter 미디어 타입
                # -----------------------------------------
                else:
                    # 현재 단계에서는 타입을 추정하지 않는다.
                    # media_extractor 단계에서 게시물 URL을 기반으로
                    # 실제 미디어 타입을 판단하게 한다.
                    continue

                add_media(
                    media_type_value=media_type,
                    source_value=media_url,
                )

        # =================================================
        # Instagram
        # =================================================
        elif sn_type == "INSTAGRAM":
            for media_item in media_items:
                if not isinstance(media_item, dict):
                    continue

                child_medias = media_item.get("childMedias")

                # -----------------------------------------
                # Instagram Carousel
                # -----------------------------------------
                if (
                    isinstance(child_medias, list)
                    and child_medias
                ):
                    for child_media in child_medias:
                        if not isinstance(child_media, dict):
                            continue

                        child_media_type = child_media.get(
                            "type"
                        )

                        # child media에 type이 없으면
                        # 해당 항목은 저장하지 않는다.
                        if not child_media_type:
                            continue

                        child_media_url = (
                            child_media.get("source")
                            or child_media.get("picture")
                        )

                        add_media(
                            media_type_value=child_media_type,
                            source_value=child_media_url,
                        )

                # -----------------------------------------
                # Instagram 단일 이미지 또는 영상
                # -----------------------------------------
                else:
                    instagram_media_type = media_item.get(
                        "type"
                    )

                    # media item에 type이 없으면
                    # 전체 처리 후 Permalink fallback을 사용한다.
                    if not instagram_media_type:
                        continue

                    instagram_media_url = (
                        media_item.get("source")
                        or media_item.get("picture")
                    )

                    add_media(
                        media_type_value=instagram_media_type,
                        source_value=instagram_media_url,
                    )
            
        # =================================================
        # Facebook
        # =================================================
        # 확인된 Sprinklr 응답 구조:
        #
        # "mediaList": [
        #     {
        #         "type": "PHOTO",
        #         "picture": "https://..."
        #     }
        # ]
        #
        # PHOTO는 picture를 우선 사용한다.
        # VIDEO 및 기타 타입은 source를 우선 사용하고,
        # 값이 없으면 picture 또는 additional.orgSMUrl을 사용한다.
        elif sn_type in {"FACEBOOK", "FB"}:
            for media_item in media_items:
                if not isinstance(media_item, dict):
                    continue

                raw_facebook_media_type = media_item.get(
                    "type"
                )

                # type이 없으면 유효한 미디어로 저장하지 않고,
                # 전체 처리 후 Permalink fallback을 사용한다.
                if not raw_facebook_media_type:
                    continue

                facebook_media_type = str(
                    raw_facebook_media_type
                ).strip().upper()

                facebook_media_url = None

                # -----------------------------------------
                # Facebook 이미지
                # -----------------------------------------
                if facebook_media_type == "PHOTO":
                    facebook_media_url = (
                        media_item.get("picture")
                        or media_item.get("source")
                    )

                # -----------------------------------------
                # Facebook 영상
                # -----------------------------------------
                elif facebook_media_type == "VIDEO":
                    facebook_media_url = (
                        media_item.get("source")
                        or media_item.get("picture")
                    )

                # -----------------------------------------
                # 기타 Facebook 미디어 타입
                # -----------------------------------------
                else:
                    facebook_media_url = (
                        media_item.get("source")
                        or media_item.get("picture")
                    )

                # media item 자체에 URL이 없을 때만
                # additional 값을 보조 fallback으로 사용한다.
                if not facebook_media_url:
                    additional = media_item.get(
                        "additional"
                    )

                    if isinstance(additional, dict):
                        facebook_media_url = (
                            additional.get("orgSMUrl")
                            or additional.get("url")
                        )

                add_media(
                    media_type_value=facebook_media_type,
                    source_value=facebook_media_url,
                )

        # =================================================
        # Youtube
        # =================================================
        elif sn_type == "YOUTUBE":
            for media_item in media_items:
                if not isinstance(media_item, dict):
                    continue

                youtube_media_type = media_item.get(
                        "type"
                )

                if not youtube_media_type:
                    continue
            
                youtube_media_url = message_obj.get("permalink")

                add_media(
                    media_type_value=youtube_media_type,
                    source_value=youtube_media_url,
                )

        # =================================================
        # TikTok
        # =================================================
        elif sn_type == "TIKTOK":
            for media_item in media_items:
                if not isinstance(media_item, dict):
                    continue

                tiktok_media_type = media_item.get(
                    "type"
                )

                if not tiktok_media_type:
                    continue

                # TikTok 응답의 source는 게시물 URL이며,
                # source가 없을 때만 기존 Permalink를 사용한다.
                tiktok_media_url = (
                    media_item.get("source")
                    or record.get("Permalink")
                )

                add_media(
                    media_type_value=tiktok_media_type,
                    source_value=tiktok_media_url,
                )

        # TikTok mediaList에서 유효한 미디어를 찾지 못한 경우에만
        # 게시물 Permalink를 UNKNOWN 타입으로 저장한다.
        if sn_type == "TIKTOK" and not media_pairs:
            add_media(
                media_type_value="UNKNOWN",
                source_value=record.get("Permalink"),
            )


        # =================================================
        # Twitter / Instagram / Facebook 공통 fallback
        # =================================================
        # 다음 경우에 실행:
        #
        # 1. mediaList가 []
        # 2. mediaList가 None
        # 3. mediaList 내부에 type 키가 없음
        # 4. type은 있지만 source/picture URL이 없음
        # 5. 유효한 미디어 URL을 하나도 추출하지 못함
        #
        # 결과:
        # Media Type = UNKNOWN
        # Media URL = 게시물 Permalink
        # =================================================
        if (
            sn_type in {
                "TWITTER",
                "INSTAGRAM",
                "FACEBOOK",
                "FB",
            }
            and not media_pairs
        ):
            # record 생성 시 이미 추출한 Permalink 사용
            permalink_url = record.get("Permalink")

            # permalink가 없는 경우에만
            # additional.orgSMUrl을 보조 fallback으로 사용
            if not permalink_url:
                additional = message_obj.get("additional")

                if isinstance(additional, dict):
                    permalink_url = additional.get(
                        "orgSMUrl"
                    )

            add_media(
                media_type_value="UNKNOWN",
                source_value=permalink_url,
            )

        # -------------------------------------------------
        # 추출된 Media Type / URL을 Excel cell 값으로 변환
        # 복수 미디어는 줄바꿈으로 구분
        # -------------------------------------------------
        if media_pairs:
            record["Media Type"] = "\n".join(
                media_type
                for media_type, _ in media_pairs
            )

            record["Media URL"] = "\n".join(
                media_url
                for _, media_url in media_pairs
            )

        # -------------------------------------------------
        # 전략법인 sheet 전용 컬럼
        # -------------------------------------------------
        if target_sheet_name == "Raw Data_전략법인":
            record["Author Screen Name"] = (
                author_screen_name
            )

        records.append(record)

    # -----------------------------------------------------
    # DataFrame 생성 단계의 컬럼 순서
    # 기존 컬럼을 먼저 배치하고 Sender Profile 컬럼은 뒤에 둔다.
    # source_widget 등의 실행 메타데이터는 main()에서 추가한다.
    # -----------------------------------------------------
    base_columns = [
        "Conversation Stream",
        "Campaign ID",
        "Profile URL",
        "User Name",
        "Permalink",
        "Created Time",
        "snType column",
    ]

    if target_sheet_name == "Raw Data_전략법인":
        columns = (
            base_columns
            + [
                "Author Screen Name",
                "Media Type",
                "Media URL",
            ]
            + SENDER_PROFILE_COLUMNS
        )
    else:
        columns = (
            base_columns
            + [
                "Media Type",
                "Media URL",
            ]
            + SENDER_PROFILE_COLUMNS
        )

    return pd.DataFrame(
        records,
        columns=columns,
    )

# =========================================================
# 12. Sprinklr response에서 필요한 data 추출 
# =========================================================
def parse_sprinklr_response(response_json: dict, payload: dict | None = None) -> pd.DataFrame:
    print("Top-level response keys:", list(response_json.keys()))

    if "data" in response_json:
        print("response_json['data'] type:", type(response_json["data"]))
        if isinstance(response_json["data"], dict):
            print("data keys:", list(response_json["data"].keys()))

    if "errors" in response_json:
        print("Response errors:", response_json["errors"])

    rows, headings, found_path = find_rows_and_headings_in_response(response_json)

    if rows is None:
        raise ValueError(
            "Could not find row data or headings in Sprinklr response. "
            "Please inspect saved response sample."
        )

    print(f"Found response data at: {found_path}")
    print("headings:", headings)
    print("number of rows:", len(rows))

    # -----------------------------------------------------
    # Case 0. headings만 있고 rows가 없는 경우
    # -----------------------------------------------------
    if len(rows) == 0:
        if headings:
            print("No row data found. Returning empty DataFrame with headings.")
            return pd.DataFrame(columns=headings)

        print("No row data and no headings found. Returning empty DataFrame.")
        return pd.DataFrame()

    first_row = rows[0]
    print("first row type:", type(first_row))
    print("first row preview:", first_row)

    # -----------------------------------------------------
    # Case 1. rows = [{...}, {...}]
    # -----------------------------------------------------
    if isinstance(first_row, dict):
        df = pd.json_normalize(rows)

    # -----------------------------------------------------
    # Case 2 or 3. rows = [[...], [...]]
    # -----------------------------------------------------
    elif isinstance(first_row, list):

        # Case 3. rows = [[{...}], [{...}]]
        if len(first_row) == 1 and isinstance(first_row[0], dict):
            extracted_rows = []

            for row in rows:
                if isinstance(row, list) and len(row) == 1 and isinstance(row[0], dict):
                    extracted_rows.append(row[0])
                else:
                    extracted_rows.append({"value": row})

            df = pd.json_normalize(extracted_rows)

        # Case 2. rows = [["a", "b", 1], ["c", "d", 2]]
        else:
            df = pd.DataFrame(rows)

            if headings and len(headings) == len(df.columns):
                df.columns = headings
            elif payload is not None:
                expected_columns = get_expected_columns_from_payload(payload)

                if len(expected_columns) == len(df.columns):
                    df.columns = expected_columns
                else:
                    print(
                        "Warning: column count does not match. "
                        f"Headings: {len(headings) if headings else 0}, "
                        f"Payload columns: {len(expected_columns)}, "
                        f"Response columns: {len(df.columns)}."
                    )
            else:
                print("Warning: no headings or payload columns available.")

    else:
        df = pd.DataFrame({"value": rows})

    df = make_dataframe_excel_safe(df)

    return df

# =========================================================
# 13. Widget 이름 기반으로 excel sheet 이름 결정  
# =========================================================

def get_target_sheet_name(widget_name: str) -> str:
    """
    누락 widget은 Raw Data_원문 시트에 저장한다.
    """

    widget_name = widget_name.strip()

    if widget_name.startswith("6."):
        return "Raw Data_원문"

    raise ValueError(
        f"Cannot determine target sheet for widget: {widget_name}. "
        "누락 widget name must start with '6.'."
    )


# =========================================================
# 14. Dict/list -> JSON string  
# =========================================================

def make_dataframe_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def convert_value(value):
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    return df.map(convert_value)

# =========================================================
# 15. DataFrame을 Excel 특정 sheet에 저장
# =========================================================

def append_dataframe_to_excel(
    workbook: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
) -> None:
    """
    DataFrame을 지정된 Excel 시트에 저장한다.

    처리 규칙:
    1. 시트가 없으면 즉시 새로 생성하고 헤더와 데이터를 기록한다.
    2. 이미 존재하는 시트가 완전히 비어 있으면 새 시트로 교체한 뒤 기록한다.
    3. 기존 시트에 헤더가 있으면 신규 DataFrame 헤더와 비교한다.
    4. 헤더가 같으면 데이터만 아래에 이어서 추가한다.

    주의:
    새로 생성한 빈 시트에 iter_rows()를 실행하면 openpyxl이 A1 셀을
    사용된 셀처럼 만들 수 있으므로, 새 시트는 검사하지 않고 바로 기록한다.
    """

    if not sheet_name:
        raise ValueError(
            "sheet_name cannot be empty."
        )

    incoming_headers = list(df.columns)

    def is_blank_value(value: object) -> bool:
        """None 또는 공백 문자열인지 확인한다."""

        if value is None:
            return True

        if isinstance(value, str):
            return not value.strip()

        return False

    def write_dataframe_with_header(sheet) -> None:
        """헤더를 포함하여 DataFrame 전체를 시트에 기록한다."""

        for excel_row in dataframe_to_rows(
            df,
            index=False,
            header=True,
        ):
            sheet.append(excel_row)

    # -----------------------------------------------------
    # Case 1. 시트가 아직 없는 경우
    # -----------------------------------------------------
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(
            title=sheet_name
        )

        write_dataframe_with_header(sheet)
        return

    # -----------------------------------------------------
    # Case 2. 시트가 이미 존재하는 경우
    # -----------------------------------------------------
    sheet = workbook[sheet_name]

    # 신규 DataFrame 컬럼 수만큼 1행 헤더를 읽는다.
    existing_headers = [
        sheet.cell(
            row=1,
            column=column_index,
        ).value
        for column_index in range(
            1,
            len(incoming_headers) + 1,
        )
    ]

    header_is_blank = all(
        is_blank_value(value)
        for value in existing_headers
    )

    # -----------------------------------------------------
    # 미리 생성된 빈 시트인 경우
    # -----------------------------------------------------
    if header_is_blank and sheet.max_row == 1:
        sheet_index = workbook.sheetnames.index(
            sheet_name
        )

        workbook.remove(sheet)

        sheet = workbook.create_sheet(
            title=sheet_name,
            index=sheet_index,
        )

        write_dataframe_with_header(sheet)
        return

    # -----------------------------------------------------
    # 1행은 비어 있는데 2행 이하에 값이 있는 비정상 상태
    # -----------------------------------------------------
    if header_is_blank:
        raise ValueError(
            "The worksheet header row is blank, "
            "but rows exist below it.\n"
            f"Sheet: {sheet_name}\n"
            f"max_row: {sheet.max_row}\n"
            f"max_column: {sheet.max_column}\n"
            "This indicates that data was written below "
            "an empty first row."
        )

    # -----------------------------------------------------
    # 신규 DataFrame 컬럼보다 오른쪽에 추가 헤더가 있는지 확인
    # -----------------------------------------------------
    extra_headers = [
        sheet.cell(
            row=1,
            column=column_index,
        ).value
        for column_index in range(
            len(incoming_headers) + 1,
            sheet.max_column + 1,
        )
    ]

    has_extra_headers = any(
        not is_blank_value(value)
        for value in extra_headers
    )

    # -----------------------------------------------------
    # 기존 헤더와 신규 헤더가 다른 경우 중단
    # -----------------------------------------------------
    if (
        existing_headers != incoming_headers
        or has_extra_headers
    ):
        full_existing_headers = [
            sheet.cell(
                row=1,
                column=column_index,
            ).value
            for column_index in range(
                1,
                sheet.max_column + 1,
            )
        ]

        raise ValueError(
            "Existing Excel sheet schema does not match "
            "the new DataFrame schema.\n"
            f"Sheet: {sheet_name}\n"
            f"Existing headers: {full_existing_headers}\n"
            f"Incoming headers: {incoming_headers}"
        )

    # -----------------------------------------------------
    # 헤더가 같으면 헤더 없이 데이터만 아래에 추가
    # -----------------------------------------------------
    for excel_row in dataframe_to_rows(
        df,
        index=False,
        header=False,
    ):
        sheet.append(excel_row)

# =========================================================
# 16. Srpinklr response 저장 파일 생성
# =========================================================
def save_response_sample(
    response_json: dict,
    widget_name: str,
    output_dir: str | Path,
) -> None:
    safe_widget_name = (
        widget_name
        .replace(" ", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
    )

    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)

    file_path = path / f"{safe_widget_name}_response.json"

    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(response_json, f, ensure_ascii=False, indent=2)

    return file_path

# =========================================================
# 17. 실행 함수
# =========================================================

def main() -> None:
    args = parse_arguments()

    # 1. User input 받기 및 가공
    print("=== 누락 Sprinklr Export to Excel ===")
    print("Please provide the following inputs:")

    start_datetime = input(
        "Enter start datetime (YYYY-MM-DD HH:MM:SS): "
    ).strip()

    end_datetime = input(
        "Enter end datetime (YYYY-MM-DD HH:MM:SS): "
    ).strip()

    (
        start_time_ms,
        end_time_ms,
        data_cut_start,
        data_cut_end,
    ) = build_time_range_from_datetimes(
        start_datetime_str=start_datetime,
        end_datetime_str=end_datetime,
        timezone_str="Asia/Seoul",
    )

    # 종료 날짜 기준으로 누락 작업 폴더/파일명을 결정한다.
    end_datetime_obj = datetime.strptime(
        end_datetime,
        "%Y-%m-%d %H:%M:%S",
    )

    naming_date = end_datetime_obj.strftime(
        "%y%m%d"
    )

    naming_month = end_datetime_obj.month

    output_date_dir = resolve_output_directory(
        naming_date=naming_date,
        cli_output_dir=args.output_dir,
    )

    output_excel_path = output_date_dir / (
        f"{naming_date}_SLCC_SOV_Local Campaign Tracking_"
        f"{naming_month}월_v01.xlsx"
    )

    (
        temporary_excel_path,
        temporary_response_dir,
        final_response_dir,
    ) = prepare_output_artifacts(
        output_excel_path=output_excel_path,
        overwrite=args.overwrite,
    )

    # 동일 데이터 컷 재실행 시 기존 행 아래에 중복 추가하지 않도록
    # 항상 새 Workbook에서 시작한다.
    workbook = create_empty_workbook()

    # 조회 결과가 0건이어도 Raw Data 시트가 유지되도록
    # 두 시트를 먼저 생성한다.
    ensure_raw_data_sheets(workbook)

    try:
        for widget_config in WIDGET_CONFIGS:
            widget_name = widget_config["widget_name"]
            payload_path = widget_config["payload_path"]

            print(
                f"Processing widget: {widget_name}"
            )

            print(
                "Loading and Updating Sprinklr payload..."
            )
            payload = load_payload(
                payload_path
            )

            payload = update_payload_time_range(
                payload=payload,
                start_time_ms=start_time_ms,
                end_time_ms=end_time_ms,
            )

            save_payload(
                payload=payload,
                payload_path=payload_path,
                make_backup=False,
            )

            print(
                "Calling Sprinklr API..."
            )
            response_json = fetch_sprinklr_data(
                base_url=SPRINKLR_BASE_URL,
                endpoint=ENDPOINT,
                api_key=API_KEY,
                access_token=ACCESS_TOKEN,
                payload=payload,
            )

            # 응답 sample도 해당 누락 날짜 output 폴더에 함께 관리한다.
            save_response_sample(
                response_json=response_json,
                widget_name=widget_name,
                output_dir=temporary_response_dir,
            )

            target_sheet_name = get_target_sheet_name(
                widget_name
            )

            print(
                "Converting response to DataFrame..."
            )
            df = make_conversation_stream_dataframe(
                response_json=response_json,
                target_sheet_name=target_sheet_name,
            )

            df = make_dataframe_excel_safe(
                df
            )

            df["source_widget"] = widget_name

            # 최신 정상본과 동일하게 사람이 읽을 수 있는 datetime 문자열 저장
            # (milliseconds가 아님)
            df["data_cut_start"] = data_cut_start
            df["data_cut_end"] = data_cut_end

            df["extracted_at"] = datetime.now(
                ZoneInfo("Asia/Seoul")
            ).strftime(
                "%Y-%m-%d %H:%M:%S"
            )

            final_columns = get_raw_data_sheet_columns(
                target_sheet_name
            )

            missing_columns = [
                column
                for column in final_columns
                if column not in df.columns
            ]

            if missing_columns:
                raise ValueError(
                    "Required Raw Data columns are missing.\n"
                    f"Sheet: {target_sheet_name}\n"
                    f"Missing columns: {missing_columns}\n"
                    f"Actual columns: {list(df.columns)}"
                )

            df = df[
                final_columns
            ].copy()

            print(
                "Writing data to Excel... "
                f"sheet={target_sheet_name}, "
                f"rows={len(df)}"
            )

            append_dataframe_to_excel(
                workbook=workbook,
                df=df,
                sheet_name=target_sheet_name,
            )

        # 모든 처리가 끝난 뒤 임시 Excel을 완성한다.
        workbook.save(
            temporary_excel_path
        )

        # Excel + response sample을 최종 위치에 atomic하게 반영한다.
        commit_output_artifacts(
            temporary_excel_path=temporary_excel_path,
            output_excel_path=output_excel_path,
            temporary_response_dir=temporary_response_dir,
            final_response_dir=final_response_dir,
        )

    except Exception:
        cleanup_temporary_artifacts(
            temporary_excel_path=temporary_excel_path,
            temporary_response_dir=temporary_response_dir,
        )
        raise

    print("Done.")
    print(
        f"Output file: {output_excel_path}"
    )

if __name__ == "__main__":
    main()
